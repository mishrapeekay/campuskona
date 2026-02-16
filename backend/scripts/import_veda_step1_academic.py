"""
Veda Vidyalaya Data Import - Step 1: Academic Structure

Imports:
- Academic Years (2)
- Board (1 - CBSE)
- Classes (14)
- Sections (35)
- Subjects (18)
"""

import os
import sys
import django
from pathlib import Path
from datetime import datetime, date

# Fix Windows console encoding
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Setup Django
sys.path.append(str(Path(__file__).parent))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.development')
django.setup()

from django.db import connection
from apps.tenants.models import School
from apps.academics.models import AcademicYear, Board, Class, Section, Subject
from uuid import UUID
import openpyxl

# Configuration
TENANT_SUBDOMAIN = "vedatest"
EXCEL_FILE = "G:/School Mgmt System/mock_data/Veda_files/veda_vidyalaya_complete_data.xlsx"

print("="*80)
print("VEDA VIDYALAYA - STEP 1: ACADEMIC STRUCTURE IMPORT")
print("="*80)
print()

# Helper functions
def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except:
            return None
    return None

def parse_uuid(value):
    if value is None:
        return None
    if isinstance(value, UUID):
        return value
    try:
        return UUID(str(value))
    except:
        return None

# Find tenant
print("[1/6] Finding tenant...")
tenant = School.objects.get(subdomain=TENANT_SUBDOMAIN)
print(f"   ✅ {tenant.name} ({tenant.schema_name})")

# Switch schema
print("\n[2/6] Switching to tenant schema...")
with connection.cursor() as cursor:
    cursor.execute(f'SET search_path TO "{tenant.schema_name}", "public"')
print(f"   ✅ Schema: {tenant.schema_name}")

# Load Excel
print("\n[3/6] Loading Excel file...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print(f"   ✅ Loaded {len(wb.sheetnames)} sheets")

# Clear existing data
print("\n[4/6] Clearing existing academic data...")
Section.objects.all().delete()
Class.objects.all().delete()
Subject.objects.all().delete()
Board.objects.all().delete()
AcademicYear.objects.all().delete()
print("   ✅ Cleared")

# Import Academic Years
print("\n[5/6] Importing academic years...")
sheet = wb['Academic_Years']
headers = [cell.value for cell in sheet[1]]
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        AcademicYear.objects.create(
            id=parse_uuid(row[0]),
            name=row[1],
            start_date=parse_date(row[2]),
            end_date=parse_date(row[3]),
            is_current=row[4] if len(row) > 4 else False
        )
print(f"   ✅ Imported {AcademicYear.objects.count()} academic years")

# Create Board
print("\n[5/6] Creating CBSE board...")
board = Board.objects.create(
    id=UUID('fdfcf85f-0c8d-4256-84b3-f8242c3af3f1'),
    board_name='Central Board of Secondary Education',
    board_code='CBSE',
    board_type='NATIONAL',
    grading_system='PERCENTAGE',
    minimum_passing_percentage=33.0,
    is_active=True
)
print(f"   ✅ Created {board.board_name}")

# Import Classes
print("\n[5/6] Importing classes...")
sheet = wb['Classes']
class_order_map = {
    'LKG': 1, 'UKG': 2,
    'Class 1': 3, 'Class 2': 4, 'Class 3': 5, 'Class 4': 6, 'Class 5': 7,
    'Class 6': 8, 'Class 7': 9, 'Class 8': 10, 'Class 9': 11, 'Class 10': 12,
    'Class 11': 13, 'Class 12': 14
}

for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        class_name = row[1]
        Class.objects.create(
            id=parse_uuid(row[0]),
            name=class_name,
            display_name=class_name,
            class_order=class_order_map.get(class_name, 0),
            board=board,
            is_active=True
        )
print(f"   ✅ Imported {Class.objects.count()} classes")

# Import Sections
print("\n[5/6] Importing sections...")
sheet = wb['Sections']
current_ay = AcademicYear.objects.filter(is_current=True).first()

for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        class_obj = Class.objects.filter(id=parse_uuid(row[2])).first()
        if class_obj:
            Section.objects.create(
                id=parse_uuid(row[0]),
                name=row[1],
                class_instance=class_obj,
                academic_year=current_ay,
                max_students=40,
                is_active=True
            )
print(f"   ✅ Imported {Section.objects.count()} sections")

# Import Subjects
print("\n[5/6] Importing subjects...")
sheet = wb['Subjects']
subject_codes_used = set()
subject_counter = {}

for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        subject_name = row[1]
        base_code = (row[2] if len(row) > 2 and row[2] else subject_name[:3].upper())

        # Make code unique if duplicate
        code = base_code
        if code in subject_codes_used:
            if code not in subject_counter:
                subject_counter[code] = 1
            subject_counter[code] += 1
            code = f"{base_code}{subject_counter[code]}"

        subject_codes_used.add(code)

        Subject.objects.create(
            id=parse_uuid(row[0]),
            name=subject_name,
            code=code,
            subject_type=row[3] if len(row) > 3 and row[3] else 'THEORY',
            board=board,
            theory_max_marks=100,
            is_active=True
        )
print(f"   ✅ Imported {Subject.objects.count()} subjects")

# Summary
print("\n" + "="*80)
print("STEP 1 COMPLETE - ACADEMIC STRUCTURE IMPORTED!")
print("="*80)
print(f"\n📊 Summary:")
print(f"   Academic Years: {AcademicYear.objects.count()}")
print(f"   Boards: {Board.objects.count()}")
print(f"   Classes: {Class.objects.count()}")
print(f"   Sections: {Section.objects.count()}")
print(f"   Subjects: {Subject.objects.count()}")
print(f"\n✅ Total: {AcademicYear.objects.count() + Board.objects.count() + Class.objects.count() + Section.objects.count() + Subject.objects.count()} records")
print("\n📝 Next: Run import_veda_step2_students.py to import student data")
print("="*80)
