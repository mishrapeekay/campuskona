"""
Veda Vidyalaya - Import Remaining Data (Transport, Library, Communication, Examinations)

This script imports additional data after academic structure is already in place.
Fixed version with correct column mappings from Excel.
"""

import os
import sys
import django
from pathlib import Path
from datetime import datetime, date, timedelta
import random

# Fix Windows console encoding
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Setup Django
sys.path.append(str(Path(__file__).parent))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.development')
django.setup()

from django.db import connection, transaction
from apps.tenants.models import School
from apps.academics.models import AcademicYear
from uuid import UUID
import openpyxl

# Configuration
TENANT_SUBDOMAIN = "vedatest"
EXCEL_FILE = "G:/School Mgmt System/mock_data/Veda_files/veda_vidyalaya_complete_data.xlsx"

print("="*80)
print("VEDA VIDYALAYA - IMPORT REMAINING DATA (FIXED)")
print("="*80)
print()

# Helper functions
def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            # Try parsing ISO format with timezone
            dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
            return dt.date()
        except:
            try:
                return datetime.strptime(value, '%Y-%m-%d').date()
            except:
                return None
    return None

def parse_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            # Try parsing ISO format
            return datetime.fromisoformat(value.replace('Z', '+00:00'))
        except:
            return None
    return None

def parse_uuid(value):
    if value is None:
        return None
    if isinstance(value, UUID):
        return value
    try:
        return UUID(str(value))
    except:
        return None

def parse_time(value):
    """Parse time value from Excel"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%H:%M:%S').time()
        except:
            try:
                return datetime.strptime(value, '%H:%M').time()
            except:
                return None
    return None

# Statistics
stats = {
    'vehicles': 0,
    'routes': 0,
    'stops': 0,
    'authors': 0,
    'books': 0,
    'notices': 0,
    'exams': 0,
}

# Find tenant
print("[1/6] Finding tenant...")
tenant = School.objects.get(subdomain=TENANT_SUBDOMAIN)
print(f"   ✅ {tenant.name} ({tenant.schema_name})")

# Switch schema
print("\n[2/6] Switching to tenant schema...")
with connection.cursor() as cursor:
    cursor.execute(f'SET search_path TO "{tenant.schema_name}", "public"')
print(f"   ✅ Schema: {tenant.schema_name}")

# Load Excel
print("\n[3/6] Loading Excel file...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print(f"   ✅ Loaded {len(wb.sheetnames)} sheets")

# Get current academic year
current_ay = AcademicYear.objects.filter(is_current=True).first()

# Import Transport Data
print("\n[4/6] Importing transport data...")
try:
    from apps.transport.models import Vehicle, Route, Stop

    # Clear existing transport data
    Stop.objects.all().delete()
    Route.objects.all().delete()
    Vehicle.objects.all().delete()

    # Import Vehicles
    # Excel columns: id, registration_number, capacity, driver_id, driver_name, vehicle_type, status
    print("   → Vehicles...")
    sheet = wb['Transport_Vehicles']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            Vehicle.objects.create(
                id=parse_uuid(row[0]),
                registration_number=row[1] if row[1] else f"VEH-{str(row[0])[:8]}",
                model=row[5] if len(row) > 5 and row[5] else 'Bus',  # vehicle_type column
                capacity=int(row[2]) if row[2] else 40,
                status=row[6] if len(row) > 6 and row[6] else 'ACTIVE',
                insurance_expiry=date.today() + timedelta(days=365),
                last_service_date=date.today()
            )
            stats['vehicles'] += 1
    print(f"      ✅ Imported {stats['vehicles']} vehicles")

    # Import Routes
    # Excel columns: id, route_name, start_location, end_location, distance_km, estimated_duration
    print("   → Routes...")
    sheet = wb['Transport_Routes']
    vehicles = list(Vehicle.objects.all())
    vehicle_index = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            # Assign vehicles to routes in round-robin fashion
            vehicle = vehicles[vehicle_index % len(vehicles)] if vehicles else None
            vehicle_index += 1

            Route.objects.create(
                id=parse_uuid(row[0]),
                name=row[1] if row[1] else f"Route-{str(row[0])[:8]}",
                start_point=row[2] if len(row) > 2 and row[2] else 'Start Point',
                end_point=row[3] if len(row) > 3 and row[3] else 'School',
                fare=1000.00,
                vehicle=vehicle
            )
            stats['routes'] += 1
    print(f"      ✅ Imported {stats['routes']} routes")

    # Import Stops
    # Excel columns: id, stop_name, route_id, sequence, pickup_time, drop_time, fee
    print("   → Stops...")
    sheet = wb['Transport_Stops']

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[2]:  # Must have id and route_id
            route_id = parse_uuid(row[2])
            try:
                route = Route.objects.get(id=route_id)

                Stop.objects.create(
                    id=parse_uuid(row[0]),
                    route=route,
                    name=row[1] if row[1] else f"Stop-{str(row[0])[:8]}",
                    sequence_order=int(row[3]) if len(row) > 3 and row[3] else 1,
                    arrival_time=parse_time(row[4]) if len(row) > 4 else None,
                    pickup_fare=float(row[6]) if len(row) > 6 and row[6] else 500.00
                )
                stats['stops'] += 1
            except Route.DoesNotExist:
                continue  # Skip stops with non-existent routes

    print(f"      ✅ Imported {stats['stops']} stops")
    print(f"   ✅ Transport data imported successfully")

except Exception as e:
    print(f"   ⚠️  Transport import error: {e}")
    import traceback
    traceback.print_exc()

# Import Library Data
print("\n[5/6] Importing library data...")
try:
    from apps.library.models import Author, Book

    # Clear existing library data
    Book.objects.all().delete()
    Author.objects.all().delete()

    # Create a dictionary to track authors
    authors_dict = {}

    # Import Books
    # Excel columns: id, title, author, isbn, category, publisher, publication_year, total_copies, available_copies, shelf_location
    print("   → Books with Authors...")
    sheet = wb['Library_Books']

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            author_name = row[2] if len(row) > 2 and row[2] else 'Unknown Author'

            # Get or create author
            if author_name not in authors_dict:
                author = Author.objects.create(
                    name=author_name,
                    bio=f"Biography of {author_name}"
                )
                authors_dict[author_name] = author
                stats['authors'] += 1
            else:
                author = authors_dict[author_name]

            Book.objects.create(
                id=parse_uuid(row[0]),
                title=row[1] if row[1] else 'Untitled Book',
                author=author,  # ForeignKey to Author
                isbn=row[3] if len(row) > 3 else None,
                category=row[4] if len(row) > 4 else 'General',
                publication_year=int(row[6]) if len(row) > 6 and row[6] else 2020,
                quantity=int(row[7]) if len(row) > 7 and row[7] else 1,
                available_copies=int(row[8]) if len(row) > 8 and row[8] else 1,
                location=row[9] if len(row) > 9 and row[9] else 'Library'
            )
            stats['books'] += 1

    print(f"      ✅ Created {stats['authors']} unique authors")
    print(f"      ✅ Imported {stats['books']} books")
    print(f"   ✅ Library data imported successfully")

except Exception as e:
    print(f"   ⚠️  Library import error: {e}")
    import traceback
    traceback.print_exc()

# Import Communication & Examination Data
print("\n[6/6] Importing communication and examination data...")
try:
    from apps.communication.models import Notice
    from apps.examinations.models import Examination
    from apps.authentication.models import User

    # Get an admin user to assign as posted_by
    admin_user = User.objects.filter(user_type='SCHOOL_ADMIN').first()

    # Clear existing data
    Notice.objects.all().delete()
    Examination.objects.all().delete()

    # Import Notices
    # Excel columns: id, title, content, target_audience, published_at, created_by
    print("   → Notices...")
    sheet = wb['Notices']

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            Notice.objects.create(
                id=parse_uuid(row[0]),
                title=row[1] if row[1] else 'Notice',
                content=row[2] if len(row) > 2 and row[2] else 'Notice content',
                target_audience=row[3] if len(row) > 3 and row[3] else 'ALL',
                priority='MEDIUM',
                posted_by=admin_user,
                is_published=True,
                display_until=date.today() + timedelta(days=30)
            )
            stats['notices'] += 1

    print(f"      ✅ Imported {stats['notices']} notices")

    # Import Examinations
    # Excel columns: id, exam_name, exam_type, academic_year_id, start_date, end_date
    print("   → Examinations...")
    sheet = wb['Exams']

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            Examination.objects.create(
                id=parse_uuid(row[0]),
                name=row[1] if row[1] else 'Examination',
                exam_type=row[2] if len(row) > 2 and row[2] else 'TERM',
                academic_year=current_ay,
                grade_scale='PERCENTAGE',
                start_date=parse_date(row[4]) if len(row) > 4 else date.today(),
                end_date=parse_date(row[5]) if len(row) > 5 else date.today() + timedelta(days=7),
                result_date=date.today() + timedelta(days=14),
                status='UPCOMING',
                description='',
                is_published=True
            )
            stats['exams'] += 1

    print(f"      ✅ Imported {stats['exams']} examinations")
    print(f"   ✅ Communication & Examination data imported successfully")

except Exception as e:
    print(f"   ⚠️  Communication/Examination import error: {e}")
    import traceback
    traceback.print_exc()

# Summary
print("\n" + "="*80)
print("IMPORT COMPLETE!")
print("="*80)

total = sum(stats.values())
print(f"\n📊 New Records Imported: {total:,}")
print()

for key, value in stats.items():
    if value > 0:
        print(f"   {key.replace('_', ' ').title()}: {value:,}")

print("\n" + "="*80)
print("COMBINED TOTALS (Including Academic Structure from Step 1)")
print("="*80)
print()
print("✅ Academic Structure (Step 1): 79 records")
print("   - Academic Years: 2")
print("   - Education Board: 1")
print("   - Classes: 14")
print("   - Sections: 35")
print("   - Subjects: 27")
print()
print(f"✅ Additional Data (Step 2): {total} records")
print(f"   - Transport: {stats['vehicles'] + stats['routes'] + stats['stops']} records")
print(f"   - Library: {stats['authors'] + stats['books']} records")
print(f"   - Communication: {stats['notices']} records")
print(f"   - Examinations: {stats['exams']} records")
print()
print(f"✅ GRAND TOTAL: {79 + total} records")
print()
print("="*80)
print("READY FOR TESTING!")
print("="*80)
print()
print("📝 You can now test the complete system:")
print()
print("1. Academic Management:")
print("   - Classes, Sections, Subjects")
print("   - Academic Years")
print()
print("2. Transport Management:")
print(f"   - {stats['vehicles']} Vehicles")
print(f"   - {stats['routes']} Routes")
print(f"   - {stats['stops']} Stops")
print()
print("3. Library Management:")
print(f"   - {stats['authors']} Authors")
print(f"   - {stats['books']} Books")
print()
print("4. Communication:")
print(f"   - {stats['notices']} Notices")
print()
print("5. Examinations:")
print(f"   - {stats['exams']} Exams")
print()
print("Login at: http://localhost:3000")
print("Admin: admin@veda.com / admin123")
print()
print("="*80)
