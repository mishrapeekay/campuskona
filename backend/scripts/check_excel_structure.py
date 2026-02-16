import openpyxl

wb = openpyxl.load_workbook('G:/School Mgmt System/mock_data/Veda_files/veda_vidyalaya_complete_data.xlsx', data_only=True)

# Check Transport_Vehicles
print("=" * 60)
print("Transport_Vehicles sheet:")
sheet = wb['Transport_Vehicles']
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=3, values_only=True)):
    print(f"Row {i}: {row}")

print("\n" + "=" * 60)
print("Library_Books sheet:")
sheet = wb['Library_Books']
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=3, values_only=True)):
    print(f"Row {i}: {row}")

print("\n" + "=" * 60)
print("Notices sheet:")
sheet = wb['Notices']
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=3, values_only=True)):
    print(f"Row {i}: {row}")
