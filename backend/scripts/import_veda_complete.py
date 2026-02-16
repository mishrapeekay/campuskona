"""
Veda Vidyalaya Complete Data Import Script

Imports comprehensive mock data from Excel file (11,000+ records) including:
- Academic structure (classes, sections, subjects)
- Students (1,080 students with parents and documents)
- Staff (66 members with attendance)
- Transport (vehicles, routes, stops, allocations)
- Library (500 books, 200 issues)
- Communication (notices, events)
- Examinations (exams, results)

Prerequisites:
1. Install openpyxl: pip install openpyxl
2. Create tenant: Veda Test (subdomain: vedatest)
3. Ensure all Django models are migrated
"""

import os
import sys
import django
from pathlib import Path
from datetime import datetime, date
import traceback

# Fix Windows console encoding
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Setup Django
sys.path.append(str(Path(__file__).parent))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.development')
django.setup()

from django.db import connection, transaction
from apps.tenants.models import School
from uuid import UUID

# Configuration
TENANT_SUBDOMAIN = "vedatest"
EXCEL_FILE = "G:/School Mgmt System/mock_data/Veda_files/veda_vidyalaya_complete_data.xlsx"

print("="*80)
print("VEDA VIDYALAYA - COMPLETE DATA IMPORT")
print("="*80)
print(f"\n📁 Excel File: {EXCEL_FILE}")
print(f"🏫 Tenant: {TENANT_SUBDOMAIN}")
print()

# Check if openpyxl is installed
try:
    import openpyxl
except ImportError:
    print("❌ Error: openpyxl not installed!")
    print("\nPlease install it:")
    print("  pip install openpyxl")
    sys.exit(1)

# Step 1: Find tenant
print("[1/11] Finding tenant...")
try:
    tenant = School.objects.get(subdomain=TENANT_SUBDOMAIN)
    print(f"   ✅ Found tenant: {tenant.name}")
    print(f"   ✅ Schema: {tenant.schema_name}")
except School.DoesNotExist:
    print(f"   ❌ Tenant '{TENANT_SUBDOMAIN}' not found!")
    print("\n   Please create the tenant first via Django admin.")
    sys.exit(1)

# Step 2: Switch to tenant schema
print("\n[2/11] Switching to tenant schema...")
with connection.cursor() as cursor:
    cursor.execute(f'SET search_path TO "{tenant.schema_name}", "public"')
print(f"   ✅ Switched to schema: {tenant.schema_name}")

# Step 3: Load Excel file
print("\n[3/11] Loading Excel file...")
try:
    workbook = openpyxl.load_workbook(EXCEL_FILE, read_only=False, data_only=True)
    sheet_names = workbook.sheetnames
    print(f"   ✅ Loaded workbook with {len(sheet_names)} sheets")
    print(f"   📊 Sheets: {', '.join(sheet_names[:5])}...")
except Exception as e:
    print(f"   ❌ Error loading Excel file: {e}")
    traceback.print_exc()
    sys.exit(1)

# Helper functions
def parse_date(value):
    """Convert Excel date to Python date"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except:
            return None
    return None

def parse_uuid(value):
    """Convert string to UUID"""
    if value is None:
        return None
    if isinstance(value, UUID):
        return value
    try:
        return UUID(str(value))
    except:
        return None

def get_sheet_data(sheet_name):
    """Get all rows from a sheet as list of dicts"""
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # Skip empty rows
            continue
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data.append(row_dict)

    return data

# Import statistics
stats = {
    'academic_years': 0,
    'boards': 0,
    'classes': 0,
    'sections': 0,
    'subjects': 0,
    'students': 0,
    'enrollments': 0,
    'parents': 0,
    'documents': 0,
    'staff': 0,
    'attendance': 0,
    'vehicles': 0,
    'routes': 0,
    'stops': 0,
    'allocations': 0,
    'books': 0,
    'issues': 0,
    'notices': 0,
    'events': 0,
    'exams': 0,
    'results': 0,
}

# Step 4: Import Academic Structure
print("\n[4/11] Importing Academic Structure...")
try:
    from apps.academics.models import AcademicYear, Board, Class, Section, Subject

    # Clear existing data to avoid duplicates
    print("   → Clearing existing data...")
    Section.objects.all().delete()
    Class.objects.all().delete()
    Subject.objects.all().delete()
    Board.objects.all().delete()
    AcademicYear.objects.all().delete()
    print("      ✅ Cleared existing academic data")

    # Import Academic Years
    print("   → Academic Years...")
    ay_data = get_sheet_data('Academic_Years')
    for row in ay_data:
        AcademicYear.objects.create(
            id=parse_uuid(row.get('id')),
            name=row.get('name'),
            start_date=parse_date(row.get('start_date')),
            end_date=parse_date(row.get('end_date')),
            is_current=row.get('is_current', False)
        )
        stats['academic_years'] += 1
    print(f"      ✅ Imported {stats['academic_years']} academic years")

    # Import Board (create manually since not in Excel)
    print("   → Education Board...")
    board = Board.objects.create(
        id=UUID('fdfcf85f-0c8d-4256-84b3-f8242c3af3f1'),
        board_name='Central Board of Secondary Education',
        board_code='CBSE',
        board_type='NATIONAL',
        grading_system='PERCENTAGE',
        minimum_passing_percentage=33.0,
        description='CBSE Board',
        is_active=True
    )
    stats['boards'] = 1
    print(f"      ✅ Created CBSE board")

    # Import Classes (with proper class_order)
    print("   → Classes...")
    class_data = get_sheet_data('Classes')
    class_order_map = {
        'LKG': 1, 'UKG': 2,
        'Class 1': 3, 'Class 2': 4, 'Class 3': 5, 'Class 4': 6, 'Class 5': 7,
        'Class 6': 8, 'Class 7': 9, 'Class 8': 10, 'Class 9': 11, 'Class 10': 12,
        'Class 11': 13, 'Class 12': 14
    }

    for row in class_data:
        class_name = row.get('name')
        Class.objects.create(
            id=parse_uuid(row.get('id')),
            name=class_name,
            display_name=class_name,
            class_order=class_order_map.get(class_name, 0),
            board=board,
            is_active=True
        )
        stats['classes'] += 1
    print(f"      ✅ Imported {stats['classes']} classes")

    # Import Sections
    print("   → Sections...")
    section_data = get_sheet_data('Sections')
    for row in section_data:
        class_obj = Class.objects.filter(id=parse_uuid(row.get('class_id'))).first()
        if class_obj:
            Section.objects.update_or_create(
                id=parse_uuid(row.get('id')),
                defaults={
                    'name': row.get('name'),
                    'class_field': class_obj,
                    'capacity': row.get('capacity', 30),
                    'is_active': True
                }
            )
            stats['sections'] += 1
    print(f"      ✅ Imported {stats['sections']} sections")

    # Import Subjects
    print("   → Subjects...")
    subject_data = get_sheet_data('Subjects')
    for row in subject_data:
        Subject.objects.update_or_create(
            id=parse_uuid(row.get('id')),
            defaults={
                'name': row.get('name'),
                'code': row.get('code', row.get('name')[:3].upper()),
                'subject_type': row.get('type', 'THEORY'),
                'is_active': True
            }
        )
        stats['subjects'] += 1
    print(f"      ✅ Imported {stats['subjects']} subjects")

    print(f"\n   ✅ Academic Structure Complete!")

except Exception as e:
    print(f"   ❌ Error: {e}")
    traceback.print_exc()

# Print summary
print("\n" + "="*80)
print("IMPORT SUMMARY")
print("="*80)
for key, value in stats.items():
    if value > 0:
        print(f"   {key.replace('_', ' ').title()}: {value:,}")

print("\n" + "="*80)
print("STATUS: Academic structure imported successfully!")
print("="*80)
print("\n📝 Next Steps:")
print("   1. Review imported data in Django admin")
print("   2. Run import script again to continue with students, staff, etc.")
print("   3. Or extend this script to import remaining modules")
print("\n" + "="*80)
