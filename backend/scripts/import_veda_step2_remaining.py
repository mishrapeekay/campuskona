"""
Veda Vidyalaya - Import Remaining Data (Transport, Library, Communication, Examinations)

This script imports additional data after academic structure is already in place.
"""

import os
import sys
import django
from pathlib import Path
from datetime import datetime, date, timedelta
import random

# Fix Windows console encoding
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Setup Django
sys.path.append(str(Path(__file__).parent))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.development')
django.setup()

from django.db import connection, transaction
from apps.tenants.models import School
from apps.academics.models import AcademicYear
from uuid import UUID
import openpyxl

# Configuration
TENANT_SUBDOMAIN = "vedatest"
EXCEL_FILE = "G:/School Mgmt System/mock_data/Veda_files/veda_vidyalaya_complete_data.xlsx"

print("="*80)
print("VEDA VIDYALAYA - IMPORT REMAINING DATA")
print("="*80)
print()

# Helper functions
def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except:
            return None
    return None

def parse_uuid(value):
    if value is None:
        return None
    if isinstance(value, UUID):
        return value
    try:
        return UUID(str(value))
    except:
        return None

def parse_time(value):
    """Parse time value from Excel"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%H:%M:%S').time()
        except:
            try:
                return datetime.strptime(value, '%H:%M').time()
            except:
                return None
    return None

# Statistics
stats = {
    'vehicles': 0,
    'routes': 0,
    'stops': 0,
    'books': 0,
    'notices': 0,
    'events': 0,
    'exams': 0,
}

# Find tenant
print("[1/6] Finding tenant...")
tenant = School.objects.get(subdomain=TENANT_SUBDOMAIN)
print(f"   ✅ {tenant.name} ({tenant.schema_name})")

# Switch schema
print("\n[2/6] Switching to tenant schema...")
with connection.cursor() as cursor:
    cursor.execute(f'SET search_path TO "{tenant.schema_name}", "public"')
print(f"   ✅ Schema: {tenant.schema_name}")

# Load Excel
print("\n[3/6] Loading Excel file...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
print(f"   ✅ Loaded {len(wb.sheetnames)} sheets")

# Get current academic year
current_ay = AcademicYear.objects.filter(is_current=True).first()

# Import Transport Data
print("\n[4/6] Importing transport data...")
try:
    from apps.transport.models import Vehicle, Route, Stop

    # Clear existing transport data
    Stop.objects.all().delete()
    Route.objects.all().delete()
    Vehicle.objects.all().delete()

    # Import Vehicles
    print("   → Vehicles...")
    sheet = wb['Transport_Vehicles']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            Vehicle.objects.create(
                id=parse_uuid(row[0]),
                registration_number=row[1] if len(row) > 1 else f"VEH-{row[0][:8]}",
                model=row[2] if len(row) > 2 else 'Bus',
                capacity=row[3] if len(row) > 3 else 40,
                status='ACTIVE',
                insurance_expiry=parse_date(row[4]) if len(row) > 4 else date.today() + timedelta(days=365),
                last_service_date=parse_date(row[5]) if len(row) > 5 else date.today()
            )
            stats['vehicles'] += 1
    print(f"      ✅ Imported {stats['vehicles']} vehicles")

    # Import Routes
    print("   → Routes...")
    sheet = wb['Transport_Routes']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            # Get a vehicle for this route
            vehicle = Vehicle.objects.first()

            Route.objects.create(
                id=parse_uuid(row[0]),
                name=row[1] if len(row) > 1 else f"Route-{row[0][:8]}",
                start_point=row[2] if len(row) > 2 else 'Start Point',
                end_point=row[3] if len(row) > 3 else 'School',
                fare=row[4] if len(row) > 4 else 1000.00,
                vehicle=vehicle
            )
            stats['routes'] += 1
    print(f"      ✅ Imported {stats['routes']} routes")

    # Import Stops
    print("   → Stops...")
    sheet = wb['Transport_Stops']
    routes = list(Route.objects.all())
    route_index = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            # Assign stops to routes in round-robin fashion
            route = routes[route_index % len(routes)]
            route_index += 1

            Stop.objects.create(
                id=parse_uuid(row[0]),
                route=route,
                name=row[1] if len(row) > 1 else f"Stop-{row[0][:8]}",
                sequence_order=row[2] if len(row) > 2 else route_index,
                arrival_time=parse_time(row[3]) if len(row) > 3 else None,
                pickup_fare=row[4] if len(row) > 4 else 500.00
            )
            stats['stops'] += 1
    print(f"      ✅ Imported {stats['stops']} stops")

    print(f"   ✅ Transport data imported successfully")

except Exception as e:
    print(f"   ⚠️  Transport import error: {e}")
    import traceback
    traceback.print_exc()

# Import Library Data
print("\n[5/6] Importing library data...")
try:
    from apps.library.models import Book

    # Clear existing library data
    Book.objects.all().delete()

    # Import Books
    print("   → Books...")
    sheet = wb['Library_Books']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            Book.objects.create(
                id=parse_uuid(row[0]),
                title=row[1] if len(row) > 1 else 'Untitled Book',
                isbn=row[2] if len(row) > 2 else None,
                author=row[3] if len(row) > 3 else 'Unknown Author',
                category=row[4] if len(row) > 4 else 'General',
                publication_year=row[5] if len(row) > 5 else 2020,
                quantity=row[6] if len(row) > 6 else 1,
                available_copies=row[7] if len(row) > 7 else 1,
                location=row[8] if len(row) > 8 else 'Library'
            )
            stats['books'] += 1
    print(f"      ✅ Imported {stats['books']} books")

    print(f"   ✅ Library data imported successfully")

except Exception as e:
    print(f"   ⚠️  Library import error: {e}")
    import traceback
    traceback.print_exc()

# Import Communication & Examination Data
print("\n[6/6] Importing communication and examination data...")
try:
    from apps.communication.models import Notice
    from apps.examinations.models import Examination
    from apps.authentication.models import User

    # Get an admin user to assign as posted_by
    admin_user = User.objects.filter(user_type='SCHOOL_ADMIN').first()

    # Clear existing data
    Notice.objects.all().delete()
    Examination.objects.all().delete()

    # Import Notices
    print("   → Notices...")
    sheet = wb['Notices']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            Notice.objects.create(
                id=parse_uuid(row[0]),
                title=row[1] if len(row) > 1 else 'Notice',
                content=row[2] if len(row) > 2 else 'Notice content',
                target_audience=row[3] if len(row) > 3 else 'ALL',
                priority=row[4] if len(row) > 4 else 'MEDIUM',
                posted_by=admin_user,
                is_published=True,
                display_until=parse_date(row[5]) if len(row) > 5 else date.today() + timedelta(days=30)
            )
            stats['notices'] += 1
    print(f"      ✅ Imported {stats['notices']} notices")

    # Import Examinations
    print("   → Examinations...")
    sheet = wb['Exams']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            Examination.objects.create(
                id=parse_uuid(row[0]),
                name=row[1] if len(row) > 1 else 'Examination',
                exam_type=row[2] if len(row) > 2 else 'TERM',
                academic_year=current_ay,
                grade_scale=row[3] if len(row) > 3 else 'PERCENTAGE',
                start_date=parse_date(row[4]) if len(row) > 4 else date.today(),
                end_date=parse_date(row[5]) if len(row) > 5 else date.today() + timedelta(days=7),
                result_date=parse_date(row[6]) if len(row) > 6 else date.today() + timedelta(days=14),
                status='UPCOMING',
                description=row[7] if len(row) > 7 else '',
                is_published=True
            )
            stats['exams'] += 1
    print(f"      ✅ Imported {stats['exams']} examinations")

    print(f"   ✅ Communication & Examination data imported successfully")

except Exception as e:
    print(f"   ⚠️  Communication/Examination import error: {e}")
    import traceback
    traceback.print_exc()

# Summary
print("\n" + "="*80)
print("IMPORT COMPLETE!")
print("="*80)

total = sum(stats.values())
print(f"\n📊 Records Imported: {total:,}")
print()

for key, value in stats.items():
    if value > 0:
        print(f"   {key.replace('_', ' ').title()}: {value:,}")

print("\n" + "="*80)
print("COMBINED TOTALS (Including Academic Structure)")
print("="*80)
print()
print("✅ Academic Structure: 79 records")
print(f"✅ Additional Data: {total} records")
print(f"✅ TOTAL: {79 + total} records")
print()
print("📝 You can now test the complete system with:")
print("   - Academic structure (classes, sections, subjects)")
print("   - Transport management (vehicles, routes, stops)")
print("   - Library management (books)")
print("   - Communication (notices)")
print("   - Examinations")
print()
print("="*80)
