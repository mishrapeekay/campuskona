from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from django.utils import timezone

from apps.students.models import (
    Student,
    StudentDocument,
    StudentParent,
    StudentHealthRecord,
    StudentNote
)
from apps.students.serializers import (
    StudentSerializer,
    StudentCreateSerializer,
    StudentUpdateSerializer,
    StudentListSerializer,
    StudentDocumentSerializer,
    StudentParentSerializer,
    StudentParentCreateSerializer,
    StudentHealthRecordSerializer,
    StudentNoteSerializer,
    StudentBulkUploadSerializer
)
from apps.authentication.permissions import (
    IsSuperAdmin,
    IsSchoolAdmin,
    HasPermission
)
from apps.core.models import AuditLog
from apps.privacy.services.data_export import DataExportService


class StudentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Student CRUD operations

    Endpoints:
    - GET /api/v1/students/ - List all students (with filtering, search, pagination)
    - POST /api/v1/students/ - Create new student
    - GET /api/v1/students/{id}/ - Get student details
    - PUT/PATCH /api/v1/students/{id}/ - Update student
    - DELETE /api/v1/students/{id}/ - Soft delete student
    - POST /api/v1/students/bulk_upload/ - Bulk upload students
    - POST /api/v1/students/{id}/approve/ - Approve admission
    - POST /api/v1/students/{id}/reject/ - Reject admission
    - POST /api/v1/students/{id}/transfer/ - Transfer student
    - GET /api/v1/students/{id}/profile/ - Get complete student profile
    """
    permission_classes = [IsAuthenticated] # Managed in get_queryset for granular access
    
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['admission_status', 'gender', 'category', 'religion', 'email', 'phone_number']
    search_fields = [
        'admission_number',
        'first_name',
        'last_name',
        'father_name',
        'mother_name',
        'email',
        'phone_number'
    ]
    ordering_fields = ['admission_number', 'first_name', 'admission_date', 'created_at']
    ordering = ['-admission_date']

    def get_queryset(self):
        """
        Get queryset with:
        1. Soft deletion filtering (is_deleted=False)
        2. Role-based access control (RBAC)
        3. Query param filtering
        4. Optimization (select_related)
        """
        user = self.request.user
        queryset = Student.objects.filter(is_deleted=False)
        
        # --- RBAC Logic ---
        # Super Admin, School Admin, and Principal have access to all
        if user.is_super_admin or user.is_school_admin or user.user_type == 'PRINCIPAL':
            pass
            
        elif user.user_type in ['TEACHER', 'ACCOUNTANT', 'LIBRARIAN', 'TRANSPORT_MANAGER']:
            # Staff can see all students but maybe we should restrict write access elsewhere
            pass
            
        elif user.user_type == 'STUDENT':
            # Students can only see their own profile
            try:
                student = Student.objects.get(user_id=user.id)
                queryset = queryset.filter(id=student.id)
            except Student.DoesNotExist:
                queryset = queryset.none()
            
        elif user.user_type == 'PARENT':
            # Parents can only see their children
            queryset = queryset.filter(parent_links__parent=user)
            
        else:
            # Default deny for others
            queryset = queryset.none()

        # --- Additional filters from query params ---
        class_id = self.request.query_params.get('class_id')
        if class_id:
            queryset = queryset.filter(
                class_enrollments__class_instance__id=class_id,
                class_enrollments__is_active=True
            )

        academic_year = self.request.query_params.get('academic_year')
        if academic_year:
            queryset = queryset.filter(
                class_enrollments__academic_year__id=academic_year,
                class_enrollments__is_active=True
            )

        return queryset.prefetch_related('class_enrollments')

    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """
        Get dashboard statistics for Front Desk / Admissions
        """
        # Admission Status
        total_inquiries = Student.objects.filter(admission_status='INQUIRY').count()
        pending_admissions = Student.objects.filter(admission_status='APPLIED').count()
        new_admissions_today = Student.objects.filter(
            admission_date=timezone.now().date()
        ).count()
        
        return Response({
            'total_inquiries': total_inquiries,
            'pending_admissions': pending_admissions,
            'new_admissions_today': new_admissions_today
        })

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return StudentListSerializer
        elif self.action == 'create':
            return StudentCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return StudentUpdateSerializer
        return StudentSerializer

    def perform_create(self, serializer):
        """Create student with audit log"""
        student = serializer.save()

        # Create audit log
        AuditLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='Student',
            object_id=str(student.id),
            changes=serializer.validated_data,
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    def perform_update(self, serializer):
        """Update student with audit log"""
        old_data = StudentSerializer(serializer.instance).data
        student = serializer.save()

        # Create audit log
        AuditLog.objects.create(
            user=self.request.user,
            action='UPDATE',
            model_name='Student',
            object_id=str(student.id),
            changes={
                'old': old_data,
                'new': StudentSerializer(student).data
            },
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    def perform_destroy(self, instance):
        """Soft delete student"""
        instance.delete(user=self.request.user)

        # Create audit log
        AuditLog.objects.create(
            user=self.request.user,
            action='DELETE',
            model_name='Student',
            object_id=str(instance.id),
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """
        Bulk upload students from Excel/CSV file (Workstream A).
        Supports .xlsx, .xls, .csv files.
        Detects and flags DPDP-sensitive columns (caste, religion, Aadhaar).
        Auto-generates pseudo-emails and admission numbers when absent.
        """
        import io
        import csv
        from openpyxl import load_workbook
        from django.db import transaction
        from django.contrib.auth import get_user_model
        User = get_user_model()

        file_obj = request.FILES.get('file')
        if not file_obj:
            # Fall back to serializer if file not in FILES directly
            serializer = StudentBulkUploadSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            file_obj = serializer.validated_data['file']

        filename = file_obj.name.lower()
        rows = []
        headers = []

        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = load_workbook(file_obj, read_only=True, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    return Response({'error': 'Empty file'}, status=status.HTTP_400_BAD_REQUEST)
                headers = [
                    str(h).strip().lower().replace(' ', '_') if h else ''
                    for h in all_rows[0]
                ]
                rows = [
                    dict(zip(headers, row))
                    for row in all_rows[1:]
                    if any(cell is not None for cell in row)
                ]
            elif filename.endswith('.csv'):
                content = file_obj.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                headers = [h.strip().lower().replace(' ', '_') for h in (reader.fieldnames or [])]
                rows = list(reader)
            else:
                return Response(
                    {'error': 'Unsupported file format. Use .xlsx or .csv'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response({'error': f'File parse error: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # DPDP-sensitive column detection
        DPDP_SENSITIVE = {'aadhar_number', 'aadhaar_number', 'aadhar', 'aadhaar', 'religion', 'caste', 'category'}
        dpdp_flags = [
            {'column': col, 'message': f'Column "{col}" contains sensitive personal data under DPDP Act 2023. Parental consent required.'}
            for col in headers if col in DPDP_SENSITIVE
        ]

        # Field mapping from common column names → Student model fields
        FIELD_MAP = {
            'name': 'first_name', 'student_name': 'first_name', 'first_name': 'first_name',
            'last_name': 'last_name', 'surname': 'last_name',
            'middle_name': 'middle_name',
            'dob': 'date_of_birth', 'date_of_birth': 'date_of_birth', 'birth_date': 'date_of_birth',
            'gender': 'gender',
            'phone': 'phone_number', 'mobile': 'phone_number', 'contact': 'phone_number',
            'phone_number': 'phone_number',
            'father_name': 'father_name', 'father': 'father_name',
            'mother_name': 'mother_name', 'mother': 'mother_name',
            'blood_group': 'blood_group',
            'admission_no': 'admission_number', 'admission_number': 'admission_number',
            'roll_no': 'admission_number', 'roll_number': 'admission_number',
            'email': 'email',
            'address': 'current_address_line1',
            'address_line1': 'current_address_line1', 'current_address_line1': 'current_address_line1',
            'city': 'current_city', 'current_city': 'current_city',
            'state': 'current_state', 'current_state': 'current_state',
            'pincode': 'current_pincode', 'current_pincode': 'current_pincode',
        }

        imported = 0
        errors = []
        warnings = []
        schema = request.tenant.schema_name if hasattr(request, 'tenant') else 'public'

        for row_idx, row in enumerate(rows, start=2):
            mapped = {}
            for col, val in row.items():
                field = FIELD_MAP.get(col)
                if field and val is not None and str(val).strip():
                    mapped[field] = str(val).strip()

            first_name = mapped.get('first_name', '')
            if not first_name:
                errors.append({'row': row_idx, 'error': 'Name is required'})
                continue

            admission_number = mapped.get('admission_number', '')
            if not admission_number:
                import uuid
                admission_number = f'AUTO-{uuid.uuid4().hex[:8].upper()}'
                warnings.append({'row': row_idx, 'warning': f'No admission number, auto-assigned: {admission_number}'})

            try:
                with transaction.atomic():
                    pseudo_email = (
                        mapped.get('email') or
                        f'student.{admission_number.lower().replace(" ", "")}@{schema}.campuskona.internal'
                    )

                    if User.objects.filter(email=pseudo_email).exists():
                        warnings.append({'row': row_idx, 'warning': f'User already exists for {admission_number}, skipped'})
                        continue

                    user = User.objects.create(
                        email=pseudo_email,
                        username=pseudo_email,
                        first_name=first_name,
                        last_name=mapped.get('last_name', ''),
                        is_active=True,
                    )
                    user.set_unusable_password()
                    user.save(update_fields=['password'])

                    student_data = {
                        'user': user,
                        'admission_number': admission_number,
                        'first_name': first_name,
                        'last_name': mapped.get('last_name', ''),
                        'middle_name': mapped.get('middle_name', ''),
                        'father_name': mapped.get('father_name', ''),
                        'mother_name': mapped.get('mother_name', ''),
                        'blood_group': mapped.get('blood_group', ''),
                        'current_address_line1': mapped.get('current_address_line1', ''),
                        'current_city': mapped.get('current_city', ''),
                        'current_state': mapped.get('current_state', ''),
                        'current_pincode': mapped.get('current_pincode', ''),
                        'email': mapped.get('email', '') or '',
                        'admission_date': timezone.now().date(),
                        'admission_status': 'ACTIVE',
                    }

                    # Gender normalization
                    gender_raw = mapped.get('gender', 'M').upper()
                    if gender_raw in ('M', 'MALE', 'BOY'):
                        student_data['gender'] = 'M'
                    elif gender_raw in ('F', 'FEMALE', 'GIRL'):
                        student_data['gender'] = 'F'
                    else:
                        student_data['gender'] = 'O'

                    # Phone number (max 10 digits)
                    if mapped.get('phone_number'):
                        ph = ''.join(filter(str.isdigit, mapped['phone_number']))[-10:]
                        if len(ph) == 10:
                            student_data['phone_number'] = ph

                    # Date of birth parsing
                    if mapped.get('date_of_birth'):
                        from datetime import datetime
                        for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y']:
                            try:
                                student_data['date_of_birth'] = datetime.strptime(
                                    mapped['date_of_birth'], fmt
                                ).date()
                                break
                            except ValueError:
                                continue

                    Student.objects.create(**student_data)
                    imported += 1

            except Exception as e:
                errors.append({'row': row_idx, 'error': str(e)})

        return Response({
            'imported': imported,
            'total_rows': len(rows),
            'warnings': warnings,
            'errors': errors,
            'dpdp_flags': dpdp_flags,
            'dpdp_flag_count': len(dpdp_flags),
            'message': f'Successfully imported {imported} of {len(rows)} students',
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='bulk_template')
    def bulk_template(self, request):
        """Download sample Excel template for bulk student upload."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = 'Students'

        headers = [
            'first_name', 'last_name', 'middle_name', 'date_of_birth',
            'gender', 'blood_group', 'phone_number', 'email',
            'father_name', 'mother_name',
            'admission_number', 'current_address_line1', 'current_city',
            'current_state', 'current_pincode',
        ]
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 20

        # Sample data row
        ws.append([
            'Arjun', 'Sharma', '', '2010-05-15', 'M', 'O+',
            '9876543210', '', 'Rajesh Sharma', 'Priya Sharma',
            'ADM001', '123 MG Road', 'Pune', 'Maharashtra', '411001',
        ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="student_bulk_upload_template.xlsx"'
        return response

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """
        Approve student admission
        """
        student = self.get_object()

        if student.admission_status != 'PENDING':
            return Response({
                'error': 'Only pending admissions can be approved'
            }, status=status.HTTP_400_BAD_REQUEST)

        student.admission_status = 'ACTIVE'
        student.save()

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            model_name='Student',
            object_id=str(student.id),
            changes={'admission_status': 'PENDING -> ACTIVE'},
            ip_address=request.META.get('REMOTE_ADDR')
        )

        return Response({
            'message': 'Student admission approved successfully',
            'student': StudentSerializer(student).data
        })

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """
        Reject student admission
        """
        student = self.get_object()

        if student.admission_status != 'PENDING':
            return Response({
                'error': 'Only pending admissions can be rejected'
            }, status=status.HTTP_400_BAD_REQUEST)

        reason = request.data.get('reason', '')

        student.admission_status = 'REJECTED'
        student.remarks = f"Rejected: {reason}"
        student.save()

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            model_name='Student',
            object_id=str(student.id),
            changes={'admission_status': 'PENDING -> REJECTED', 'reason': reason},
            ip_address=request.META.get('REMOTE_ADDR')
        )

        return Response({
            'message': 'Student admission rejected',
            'student': StudentSerializer(student).data
        })

    @action(detail=True, methods=['post'])
    def transfer(self, request, pk=None):
        """
        Transfer student to another school
        """
        student = self.get_object()

        transfer_date = request.data.get('transfer_date')
        transfer_to_school = request.data.get('transfer_to_school', '')
        reason = request.data.get('reason', '')

        student.admission_status = 'TRANSFERRED'
        student.remarks = f"Transferred to {transfer_to_school} on {transfer_date}. Reason: {reason}"
        student.save()

        # Deactivate current class enrollment
        current_enrollment = student.get_current_class_enrollment()
        if current_enrollment:
            current_enrollment.is_active = False
            current_enrollment.save()

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            model_name='Student',
            object_id=str(student.id),
            changes={
                'admission_status': 'TRANSFERRED',
                'transfer_to': transfer_to_school,
                'transfer_date': transfer_date,
                'reason': reason
            },
            ip_address=request.META.get('REMOTE_ADDR')
        )

        return Response({
            'message': 'Student transferred successfully',
            'student': StudentSerializer(student).data
        })

    @action(detail=True, methods=['get'])
    def profile(self, request, pk=None):
        """
        Get complete student profile with all related data
        """
        student = self.get_object()

        profile_data = StudentSerializer(student).data
        profile_data['documents'] = StudentDocumentSerializer(
            student.documents.filter(is_deleted=False),
            many=True
        ).data
        profile_data['parents'] = StudentParentSerializer(
            student.parent_links.filter(is_deleted=False),
            many=True
        ).data
        profile_data['health_records'] = StudentHealthRecordSerializer(
            student.health_records.filter(is_deleted=False).order_by('-checkup_date')[:5],
            many=True
        ).data

        # Only show notes to staff
        if request.user.user_type not in ['STUDENT', 'PARENT']:
            profile_data['notes'] = StudentNoteSerializer(
                student.notes.filter(is_deleted=False).order_by('-created_at')[:10],
                many=True
            ).data

        return Response(profile_data)

    @action(detail=True, methods=['get'])
    def export_data(self, request, pk=None):
        """
        Export all student data (Right to Access - DPDP Act Section 11)

        GET /api/v1/students/{id}/export_data/?format=json

        Formats: json, csv

        Only accessible by:
        - Parents (for their children)
        - School administrators
        """
        student = self.get_object()
        format = request.query_params.get('format', 'json')

        # Check authorization
        user = request.user

        # Parents can only export their own child's data
        if user.user_type == 'PARENT':
            if not StudentParent.objects.filter(
                student=student,
                parent_user=user
            ).exists():
                return Response(
                    {'error': 'You are not authorized to export data for this student'},
                    status=status.HTTP_403_FORBIDDEN
                )
        # Non-admin staff cannot export
        elif user.user_type not in ['ADMIN', 'SUPERADMIN'] and not user.is_staff:
            return Response(
                {'error': 'Insufficient permissions to export student data'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Generate export using DataExportService
        try:
            response = DataExportService.create_export_response(student, format)

            # Create audit log
            AuditLog.objects.create(
                user=user,
                action='EXPORT',
                model_name='Student',
                object_id=student.id,
                changes={
                    'format': format,
                    'exported_for': user.get_full_name(),
                    'dpdp_compliance': 'Section 11 - Right to Access'
                },
                ip_address=self._get_client_ip(request)
            )

            return response

        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Get student statistics for dashboard
        """
        queryset = self.get_queryset()
        
        total = queryset.count()
        active = queryset.filter(admission_status='ACTIVE').count()
        pending = queryset.filter(admission_status='PENDING').count()
        transferred = queryset.filter(admission_status='TRANSFERRED').count()
        
        # Gender distribution
        male = queryset.filter(gender='M').count()
        female = queryset.filter(gender='F').count()
        other = queryset.filter(gender='O').count()
        
        # Category distribution
        general = queryset.filter(category='GENERAL').count()
        obc = queryset.filter(category='OBC').count()
        sc = queryset.filter(category='SC').count()
        st = queryset.filter(category='ST').count()
        
        return Response({
            'total': total,
            'by_status': {
                'active': active,
                'pending': pending,
                'transferred': transferred,
            },
            'by_gender': {
                'male': male,
                'female': female,
                'other': other,
            },
            'by_category': {
                'general': general,
                'obc': obc,
                'sc': sc,
                'st': st,
            }
        })

    def _get_client_ip(self, request):
        """Extract client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


class StudentDocumentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Student Documents
    """
    serializer_class = StudentDocumentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['student', 'document_type', 'is_verified']
    ordering_fields = ['created_at', 'issue_date']
    ordering = ['-created_at']

    def get_queryset(self):
        """Filter documents based on user type"""
        queryset = StudentDocument.objects.filter(is_deleted=False)

        user = self.request.user
        if user.user_type == 'STUDENT':
            # Students can only see their own documents
            queryset = queryset.filter(student__user=user)
        elif user.user_type == 'PARENT':
            # Parents can see their children's documents
            queryset = queryset.filter(student__parent_links__parent=user)

        return queryset.select_related('student', 'verified_by')

    def perform_create(self, serializer):
        """Create document with audit log"""
        document = serializer.save()

        AuditLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='StudentDocument',
            object_id=str(document.id),
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """
        Verify a student document
        """
        document = self.get_object()

        document.is_verified = True
        document.verified_by = request.user
        document.verified_at = timezone.now()
        document.save()

        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            model_name='StudentDocument',
            object_id=str(document.id),
            changes={'verified': True},
            ip_address=request.META.get('REMOTE_ADDR')
        )

        return Response({
            'message': 'Document verified successfully',
            'document': StudentDocumentSerializer(document).data
        })


class StudentParentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Student-Parent relationships
    """
    serializer_class = StudentParentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['student', 'parent', 'relation', 'is_primary_contact']

    def get_queryset(self):
        """Filter based on user type"""
        queryset = StudentParent.objects.filter(is_deleted=False)

        user = self.request.user
        if user.user_type == 'STUDENT':
            queryset = queryset.filter(student__user=user)
        elif user.user_type == 'PARENT':
            queryset = queryset.filter(parent=user)

        return queryset.select_related('student', 'parent')

    def get_serializer_class(self):
        """Return appropriate serializer"""
        if self.action == 'create_parent':
            return StudentParentCreateSerializer
        return StudentParentSerializer

    @action(detail=False, methods=['post'])
    def create_parent(self, request):
        """
        Create a new parent account and link to student
        """
        serializer = StudentParentCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        student_parent = serializer.save()

        AuditLog.objects.create(
            user=request.user,
            action='CREATE',
            model_name='StudentParent',
            object_id=str(student_parent.id),
            ip_address=request.META.get('REMOTE_ADDR')
        )

        return Response({
            'message': 'Parent account created and linked successfully',
            'data': StudentParentSerializer(student_parent).data
        }, status=status.HTTP_201_CREATED)


class StudentHealthRecordViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Student Health Records
    """
    serializer_class = StudentHealthRecordSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['student']
    ordering_fields = ['checkup_date']
    ordering = ['-checkup_date']

    def get_queryset(self):
        """Filter based on user type"""
        queryset = StudentHealthRecord.objects.filter(is_deleted=False)

        user = self.request.user
        if user.user_type == 'STUDENT':
            queryset = queryset.filter(student__user=user)
        elif user.user_type == 'PARENT':
            queryset = queryset.filter(student__parent_links__parent=user)

        return queryset.select_related('student')

    def perform_create(self, serializer):
        """Create health record with audit log"""
        record = serializer.save()

        AuditLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='StudentHealthRecord',
            object_id=str(record.id),
            ip_address=self.request.META.get('REMOTE_ADDR')
        )


class StudentNoteViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Student Notes
    """
    serializer_class = StudentNoteSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['student', 'note_type', 'is_important']
    ordering_fields = ['created_at']
    ordering = ['-created_at']

    def get_queryset(self):
        """Only staff can access notes"""
        if self.request.user.user_type in ['STUDENT', 'PARENT']:
            return StudentNote.objects.none()

        queryset = StudentNote.objects.filter(is_deleted=False)

        # Filter private notes - only creator can see
        if not self.request.user.is_super_admin:
            queryset = queryset.filter(
                Q(is_private=False) | Q(created_by=self.request.user)
            )

        return queryset.select_related('student', 'created_by')

    def perform_create(self, serializer):
        """Create note with creator"""
        note = serializer.save(created_by=self.request.user)

        AuditLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='StudentNote',
            object_id=str(note.id),
            ip_address=self.request.META.get('REMOTE_ADDR')
        )
