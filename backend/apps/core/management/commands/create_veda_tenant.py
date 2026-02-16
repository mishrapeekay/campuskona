import os
import openpyxl
from datetime import datetime, date
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.db import transaction, connection
from django.conf import settings
from apps.authentication.models import User
from apps.tenants.models import School, Subscription
from apps.students.models import Student, StudentParent
from apps.staff.models import StaffMember
from apps.academics.models import AcademicYear, Board, Class, Section, Subject, StudentEnrollment
from apps.transport.models import Vehicle, Route, Stop, TransportAllocation
from apps.library.models import Book, BookIssue, Category
from apps.communication.models import Notice, Event

class Command(BaseCommand):
    help = 'Create Veda Vidyalaya Tenant and Import Data'

    def handle(self, *args, **options):
        file_path = r"G:\School Mgmt System\mock_data\Veda_files\veda_vidyalaya_complete_data.xlsx"
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        self.stdout.write(self.style.SUCCESS("Starting Veda Vidyalaya Tenant Creation..."))

        # 1. Create Tenant
        school = self.create_tenant()
        
        # 2. Switch Context
        if hasattr(connection, 'set_tenant'):
            connection.set_tenant(school)
        else:
            # Manual schema switch for raw queries or if library not fully active
            with connection.cursor() as cursor:
                 cursor.execute(f'SET search_path TO "{school.schema_name}", public')

        self.stdout.write(f"Switched to schema: {school.schema_name}")

        # Ensure tables exist
        # Tables are created by school.create_schema()
        self.stdout.write("Running migrations for new tenant...")
        # from django.core.management import call_command
        # call_command('migrate', interactive=False)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            with transaction.atomic():
                # 3. Import Academics
                self.import_academics(wb)
                
                # 4. Import Staff
                self.import_staff(wb)
                
                # 5. Import Students & Parents
                self.import_students(wb)
                
                # 6. Import Transport
                self.import_transport(wb)
                
                # 7. Import Library
                self.import_library(wb)
                
                # 8. Import Communication
                self.import_communication(wb)

            self.stdout.write(self.style.SUCCESS("✅ Veda Vidyalaya Setup Complete!"))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Error: {str(e)}"))
            import traceback
            traceback.print_exc()

    def create_tenant(self):
        # Ensure subscription exists
        sub, _ = Subscription.objects.get_or_create(
            name='Premium Plan',
            defaults={
                'price_monthly': 5000,
                'price_yearly': 50000,
                'max_students': 2000,
                'is_active': True
            }
        )
        
        school, created = School.objects.get_or_create(
            code='VV9',
            defaults={
                'name': 'Veda Vidyalaya V9',
                'subdomain': 'veda9',
                'schema_name': 'tenant_veda_v9',
                'email': 'admin@vedavidyalaya.edu.in',
                'phone': '0751-2400001',
                'address': 'City Center',
                'city': 'Gwalior',
                'state': 'Madhya Pradesh',
                'pincode': '474011',
                'primary_board': 'CBSE',
                'subscription': sub,
                'subscription_start_date': date(2024, 4, 1),
                'subscription_end_date': date(2027, 3, 31),
                'is_active': True,
                'auto_create_schema': True
            }
        )
        if created:
            self.stdout.write("Created Tenant: Veda Vidyalaya")
            # Force schema creation if not auto-triggered
            # school.create_schema()
            pass
        else:
            self.stdout.write("Tenant Veda Vidyalaya already exists")
            
        return school

    def _get_sheet_data(self, wb, sheet_name):
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        return data

    def import_academics(self, wb):
        self.stdout.write("Importing Academics...")
        
        # Board
        board, _ = Board.objects.get_or_create(board_code='CBSE', defaults={'board_name': 'Central Board of Secondary Education'})
        
        # Academic Years
        years_data = self._get_sheet_data(wb, 'Academic_Years')
        current_year = None
        for row in years_data:
            ay, _ = AcademicYear.objects.get_or_create(
                name=row['name'],
                defaults={
                    'start_date': row['start_date'],
                    'end_date': row['end_date'],
                    'is_current': row['is_current']
                }
            )
            if row['is_current']:
                current_year = ay

        # Classes
        classes_data = self._get_sheet_data(wb, 'Classes')
        class_map = {}
        order_counter = 1
        for row in classes_data:
            self.stdout.write(f"Processing Class: {row['name']} (Order: {order_counter})")
            c, _ = Class.objects.get_or_create(
                name=row['name'],
                board=board,
                defaults={'class_order': order_counter}
            )
            class_map[row['id']] = c
            order_counter += 1

        # Subjects
        subjects_data = self._get_sheet_data(wb, 'Subjects')
        for row in subjects_data:
            s_type = row.get('type', 'THEORY')
            has_practical = (s_type == 'PRACTICAL')
            
            Subject.objects.get_or_create(
                name=row['name'],
                code=row['code'],
                defaults={
                    'subject_type': 'CORE',
                    'has_practical': has_practical,
                    'theory_max_marks': 100,
                    'practical_max_marks': 100 if has_practical else 0
                }
            )

        # Sections
        sections_data = self._get_sheet_data(wb, 'Sections')
        self.section_map_by_id = {} # Store UUID -> Object for enrollment
        for row in sections_data:
            cls = class_map.get(row['class_id'])
            if cls:
                sec, _ = Section.objects.get_or_create(
                    name=row['name'],
                    class_instance=cls,
                    academic_year=current_year
                )
                self.section_map_by_id[row['id']] = sec

    def import_staff(self, wb):
        self.stdout.write("Importing Staff...")
        staff_data = self._get_sheet_data(wb, 'Staff_Members')
        
        for row in staff_data:
            email = f"{row['employee_id'].lower()}@veda.edu" # Mock email if missing
            
            # Create User (Global)
            user, created = User.objects.get_or_create(
                email=email,
                defaults={
                    'first_name': 'Staff',
                    'last_name': row['designation'],
                    'phone': f"9{row['employee_id'][-9:]}".ljust(10, '0'), # Dummy phone
                    'user_type': 'TEACHER' if 'TEACHER' in row['designation'] else 'SCHOOL_ADMIN',
                    'is_staff': True
                }
            )
            if created:
                user.set_password('Veda@123')
                user.save()

            StaffMember.objects.get_or_create(
                employee_id=row['employee_id'],
                defaults={
                    'user': user,
                    'first_name': 'Staff',
                    'last_name': row['designation'],
                    'designation': row['designation'],
                    'department': row['department'],
                    'employment_type': row['employment_type'],
                    'joining_date': row['joining_date'],
                    'email': email,
                    'phone_number': user.phone,
                    'date_of_birth': date(1990, 1, 1), # Default
                    'gender': 'M',
                    'emergency_contact_number': '9999999999'
                }
            )

    def import_students(self, wb):
        self.stdout.write("Importing Students...")
        students_data = self._get_sheet_data(wb, 'Students')
        enroll_data = self._get_sheet_data(wb, 'Student_Enrollments')
        
        # Create map of student_id -> enrollment row
        enroll_map = {row['student_id']: row for row in enroll_data}
        
        count = 0
        for row in students_data:
            email = f"{row['admission_number'].lower()}@veda.edu"
            
            # Generate unique dummy phone: 99 + 8 digits from count
            dummy_phone = f"99{str(count).zfill(8)}"
            
            f_phone = row.get('Father_Phone', row.get('father_phone', row.get('Father Phone')))
            user_phone = str(f_phone).replace('-', '').replace(' ', '')[:10] if f_phone else dummy_phone
            if len(user_phone) < 10: user_phone = dummy_phone

            user, created = User.objects.get_or_create(
                email=email,
                defaults={
                    'first_name': row.get('first_name', row.get('First_Name')),
                    'last_name': row.get('last_name', row.get('Last_Name')),
                    'phone': user_phone,
                    'user_type': 'STUDENT'
                }
            )
            if created:
                user.set_password('Veda@123')
                user.save()

            student, _ = Student.objects.get_or_create(
                admission_number=row.get('admission_number', row.get('Admission_Number')),
                defaults={
                    'user': user,
                    'first_name': row.get('first_name', row.get('First_Name')),
                    'last_name': row.get('last_name', row.get('Last_Name')),
                    'admission_status': row.get('admission_status', 'ACTIVE'),
                    'admission_date': date(2024, 4, 1),
                    'date_of_birth': row.get('date_of_birth', row.get('Date_of_Birth')),
                    'gender': (row.get('gender') or row.get('Gender', 'O'))[0],
                    'email': email,
                    'emergency_contact_number': '9999999999',
                    'father_name': row.get('Father_Name', 'Father'),
                    'father_phone': '9999999999', 
                    'mother_name': row.get('Mother_Name', 'Mother'),
                    'current_address_line1': 'Gwalior',
                    'current_city': 'Gwalior',
                    'current_state': 'MP',
                    'current_pincode': '474001',
                    'permanent_address_line1': 'Gwalior',
                    'permanent_city': 'Gwalior',
                    'permanent_state': 'MP',
                    'permanent_pincode': '474001',
                }
            )

            # Enrollment
            enr_row = enroll_map.get(row['id'])
            if enr_row and enr_row['section_id'] in self.section_map_by_id:
                section = self.section_map_by_id[enr_row['section_id']]
                StudentEnrollment.objects.get_or_create(
                    student=student,
                    section=section,
                    defaults={
                        'academic_year': section.academic_year,
                        'roll_number': enr_row['roll_number'],
                        'enrollment_status': enr_row['status'],
                        'enrollment_date': date(2024, 4, 1)
                    }
                )
            count += 1

    def import_transport(self, wb):
        self.stdout.write("Importing Transport...")
        vehicles = self._get_sheet_data(wb, 'Transport_Vehicles')
        vehicle_map = {}
        for row in vehicles:
            v, _ = Vehicle.objects.get_or_create(
                registration_number=row['registration_number'],
                defaults={
                    'capacity': row['capacity'],
                    'status': 'ACTIVE',
                    'model': 'Tata Starbus',
                }
            )
            vehicle_map[row['id']] = v

        routes = self._get_sheet_data(wb, 'Transport_Routes')
        for row in routes:
            veh = vehicle_map.get(row['vehicle_id'])
            if veh:
                Route.objects.get_or_create(
                    name=row['name'],
                    defaults={
                        'vehicle': veh,
                        'start_point': 'School',
                        'end_point': 'City',
                        'fare': Decimal('1000.00')
                    }
                )

    def import_library(self, wb):
        self.stdout.write("Importing Library...")
        books = self._get_sheet_data(wb, 'Library_Books')
        from apps.library.models import Author

        for row in books:
            # Create/Get Author
            author_name = row['author'] or 'Unknown Author'
            author, _ = Author.objects.get_or_create(name=author_name)

            Book.objects.get_or_create(
                isbn=row['isbn'],
                defaults={
                    'title': row['title'],
                    'author': author,
                    'quantity': row['total_copies'] or 1,
                    'available_copies': row['available_copies'] or 1
                }
            )

    def import_communication(self, wb):
        self.stdout.write("Importing Communications...")
        notices = self._get_sheet_data(wb, 'Notices')
        for row in notices:
            Notice.objects.get_or_create(
                title=row['title'],
                defaults={
                    'content': row['content'],
                    'target_audience': row['target_audience'],
                    'is_published': True
                }
            )
