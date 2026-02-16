"""
Import remaining Veda Vidyalaya data that is NOT covered by create_veda_tenant or import_veda_data.

This command imports:
  1. Student Parents (2,160 records)
  2. Staff Attendance (~1,980 records)
  3. Transport Stops (~35 records)
  4. Transport Allocations (~650 records)
  5. Library Book Issues (200 records)
  6. Events (10 records)
  7. Exam Schedules + Student Marks (~150 records)

Prerequisites:
  - Veda tenant must exist (run create_veda_tenant first)
  - Core data must be imported (students, staff, academics, transport vehicles/routes, library books)

Usage:
  python manage.py import_veda_remaining
  python manage.py import_veda_remaining --skip parents,attendance
"""

import os
import openpyxl
from datetime import datetime, date, time, timedelta
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.db import connection, transaction
from apps.authentication.models import User
from apps.tenants.models import School
from apps.students.models import Student, StudentParent
from apps.staff.models import StaffMember
from apps.academics.models import AcademicYear, Class, Section, Subject, StudentEnrollment
from apps.attendance.models import StaffAttendance
from apps.transport.models import Vehicle, Route, Stop, TransportAllocation
from apps.library.models import Book, BookIssue, Author, Category
from apps.communication.models import Event
from apps.examinations.models import (
    GradeScale, Grade, ExamType, Examination, ExamSchedule, StudentMark
)


class Command(BaseCommand):
    help = 'Import remaining Veda Vidyalaya data (parents, attendance, stops, allocations, issues, events, exam results)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--skip',
            type=str,
            default='',
            help='Comma-separated list of modules to skip: parents,attendance,stops,allocations,issues,events,exams'
        )

    def handle(self, *args, **options):
        file_path = r"G:\School Mgmt System\mock_data\Veda_files\veda_vidyalaya_complete_data.xlsx"

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        # Get school and switch schema
        try:
            school = School.objects.get(code='VV9')
            self.stdout.write(f"Found school: {school.name}")
        except School.DoesNotExist:
            self.stdout.write(self.style.ERROR(
                "School 'Veda Vidyalaya' not found! Run create_veda_tenant first."
            ))
            return

        # Switch to tenant schema
        from django_tenants.utils import schema_context
        with schema_context(school.schema_name):
            self.stdout.write(f"Switched to schema: {school.schema_name}")
            skip_modules = set(s.strip() for s in options['skip'].split(',') if s.strip())

            try:
                # Disconnect audit signals to prevent issues during bulk import
                try:
                    from django.db.models.signals import post_save
                    from apps.authentication.signals import log_user_creation
                    post_save.disconnect(log_user_creation, sender=User)
                    self.stdout.write("Disconnected audit signals")
                except (ImportError, Exception):
                    pass

                wb = openpyxl.load_workbook(file_path, data_only=True)

                # Build lookup maps from already-imported data
                self.stdout.write("\nBuilding lookup maps from existing data...")
                self._build_maps(wb)

                # Import each module
                if 'parents' not in skip_modules:
                    self.import_student_parents(wb)

                if 'attendance' not in skip_modules:
                    self.import_staff_attendance(wb)

                if 'stops' not in skip_modules:
                    self.import_transport_stops(wb)

                if 'allocations' not in skip_modules:
                    self.import_transport_allocations(wb)

                if 'issues' not in skip_modules:
                    self.import_library_issues(wb)

                if 'events' not in skip_modules:
                    self.import_events(wb)

                if 'exams' not in skip_modules:
                    self.import_exam_results(wb)

                self.stdout.write(self.style.SUCCESS("\n=== Import Complete ==="))
                self.verify_data()

                # Reconnect signals
                try:
                    from django.db.models.signals import post_save
                    from apps.authentication.signals import log_user_creation
                    post_save.connect(log_user_creation, sender=User)
                except (ImportError, Exception):
                    pass

            except Exception as e:
                self.stdout.write(self.style.ERROR(f"\nFatal error: {str(e)}"))
                import traceback
                traceback.print_exc()

    # ─── Helpers ──────────────────────────────────────────────────────────

    def _get_sheet_data(self, wb, sheet_name):
        """Convert an Excel sheet to a list of dicts."""
        try:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))
            return data
        except KeyError:
            self.stdout.write(self.style.WARNING(f"  Sheet '{sheet_name}' not found, skipping"))
            return []

    def _build_maps(self, wb):
        """Build ID-to-object lookup maps from already-imported data."""
        # Student Excel-ID -> Django Student object
        students_data = self._get_sheet_data(wb, 'Students')
        self.student_map = {}  # Excel UUID -> Student
        for row in students_data:
            adm = row.get('admission_number')
            if adm:
                try:
                    student = Student.objects.get(admission_number=adm)
                    self.student_map[row.get('id')] = student
                except Student.DoesNotExist:
                    pass
        self.stdout.write(f"  Students mapped: {len(self.student_map)}")

        # Staff Excel-ID -> Django StaffMember
        staff_data = self._get_sheet_data(wb, 'Staff_Members')
        self.staff_map = {}  # Excel UUID -> StaffMember
        for row in staff_data:
            emp_id = row.get('employee_id')
            if emp_id:
                try:
                    staff = StaffMember.objects.get(employee_id=emp_id)
                    self.staff_map[row.get('id')] = staff
                except StaffMember.DoesNotExist:
                    pass
        self.stdout.write(f"  Staff mapped: {len(self.staff_map)}")

        # Route Excel-ID -> Django Route
        routes_data = self._get_sheet_data(wb, 'Transport_Routes')
        self.route_map = {}  # Excel UUID -> Route
        for row in routes_data:
            name = row.get('name')
            if name:
                try:
                    route = Route.objects.get(name=name)
                    self.route_map[row.get('id')] = route
                except Route.DoesNotExist:
                    pass
        self.stdout.write(f"  Routes mapped: {len(self.route_map)}")

        # Book Excel-ID -> Django Book
        books_data = self._get_sheet_data(wb, 'Library_Books')
        self.book_map = {}  # Excel UUID -> Book
        for row in books_data:
            isbn = row.get('isbn')
            if isbn:
                try:
                    book = Book.objects.get(isbn=isbn)
                    self.book_map[row.get('id')] = book
                except Book.DoesNotExist:
                    pass
        self.stdout.write(f"  Books mapped: {len(self.book_map)}")

        # Exam Excel-ID -> Django Examination
        exams_data = self._get_sheet_data(wb, 'Exams')
        self.exam_map = {}  # Excel UUID -> Examination
        for row in exams_data:
            name = row.get('name')
            if name:
                try:
                    exam = Examination.objects.get(name=name)
                    self.exam_map[row.get('id')] = exam
                except Examination.DoesNotExist:
                    pass
        self.stdout.write(f"  Exams mapped: {len(self.exam_map)}")

        # Subject Excel-ID -> Django Subject
        subjects_data = self._get_sheet_data(wb, 'Subjects')
        self.subject_map = {}  # Excel UUID -> Subject
        for row in subjects_data:
            code = row.get('code')
            if code:
                try:
                    subject = Subject.objects.get(code=code)
                    self.subject_map[row.get('id')] = subject
                except Subject.DoesNotExist:
                    pass
        self.stdout.write(f"  Subjects mapped: {len(self.subject_map)}")

        # Section Excel-ID -> Django Section
        sections_data = self._get_sheet_data(wb, 'Sections')
        self.section_map = {}
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if current_year:
            for row in sections_data:
                sec_name = row.get('name')
                class_id = row.get('class_id')
                if sec_name and class_id:
                    # Need class map too
                    pass  # We'll build this from enrollment data instead

        # Build section map from enrollments
        enrollments_data = self._get_sheet_data(wb, 'Student_Enrollments')
        for row in enrollments_data:
            sid = row.get('student_id')
            student = self.student_map.get(sid)
            if student:
                try:
                    enrollment = StudentEnrollment.objects.filter(student=student).first()
                    if enrollment:
                        self.section_map[row.get('section_id')] = enrollment.section
                except Exception:
                    pass
        self.stdout.write(f"  Sections mapped: {len(self.section_map)}")

    # ─── 1. Student Parents ──────────────────────────────────────────────

    def import_student_parents(self, wb):
        self.stdout.write("\n1/7 Importing Student Parents...")
        parents_data = self._get_sheet_data(wb, 'Student_Parents')
        if not parents_data:
            return

        count = 0
        errors = 0
        for row in parents_data:
            student_excel_id = row.get('student_id')
            student = self.student_map.get(student_excel_id)
            if not student:
                continue

            parent_type = row.get('parent_type', row.get('relationship', 'OTHER'))
            name = row.get('name', row.get('parent_name', ''))
            if not name:
                continue

            # Map parent_type to relation
            relation = 'OTHER'
            if parent_type:
                pt = str(parent_type).upper()
                if 'FATHER' in pt:
                    relation = 'FATHER'
                elif 'MOTHER' in pt:
                    relation = 'MOTHER'
                elif 'GUARDIAN' in pt:
                    relation = 'GUARDIAN'

            phone = row.get('phone', row.get('contact_number', ''))
            phone = str(phone).replace('-', '').replace(' ', '').replace('+91', '')[:10] if phone else ''
            if len(phone) < 10:
                phone = f"97{str(count).zfill(8)}"

            email_local = name.lower().replace(' ', '.').replace("'", '')
            email = row.get('email', f"{email_local}.{student.admission_number}@vedaparent.edu.in")
            if not email or email == 'None':
                email = f"{email_local}.{student.admission_number}@vedaparent.edu.in"

            try:
                # Create parent User in public schema
                name_parts = str(name).split(' ', 1)
                first_name = name_parts[0][:30]
                last_name = name_parts[1][:30] if len(name_parts) > 1 else 'Parent'

                parent_user, created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'phone': phone,
                        'user_type': 'PARENT',
                    }
                )
                if created:
                    parent_user.set_password('Veda@123')
                    parent_user.save()

                # Create StudentParent link
                StudentParent.objects.get_or_create(
                    student=student,
                    parent=parent_user,
                    defaults={
                        'relation': relation,
                        'is_primary_contact': relation == 'FATHER',
                        'is_emergency_contact': relation == 'FATHER',
                        'can_pickup': True,
                    }
                )
                count += 1
            except Exception as e:
                errors += 1
                if errors <= 3:
                    self.stdout.write(self.style.WARNING(f"  Skip parent: {str(e)[:80]}"))

        self.stdout.write(f"  Imported: {count} parent links ({errors} errors)")

    # ─── 2. Staff Attendance ─────────────────────────────────────────────

    def import_staff_attendance(self, wb):
        self.stdout.write("\n2/7 Importing Staff Attendance...")
        att_data = self._get_sheet_data(wb, 'Staff_Attendance')
        if not att_data:
            return

        count = 0
        errors = 0
        for row in att_data:
            staff_excel_id = row.get('staff_id')
            staff = self.staff_map.get(staff_excel_id)
            if not staff:
                continue

            att_date = row.get('attendance_date', row.get('date'))
            if not att_date:
                continue
            if isinstance(att_date, str):
                try:
                    att_date = datetime.strptime(att_date, '%Y-%m-%d').date()
                except ValueError:
                    continue
            elif isinstance(att_date, datetime):
                att_date = att_date.date()

            status = str(row.get('status', 'PRESENT')).upper().strip()
            valid_statuses = {'PRESENT', 'ABSENT', 'HALF_DAY', 'LEAVE', 'HOLIDAY', 'WEEKEND', 'WORK_FROM_HOME'}
            if status not in valid_statuses:
                status = 'PRESENT'

            remarks = row.get('remarks', '') or ''

            try:
                StaffAttendance.objects.get_or_create(
                    staff_member=staff,
                    date=att_date,
                    defaults={
                        'status': status,
                        'remarks': str(remarks),
                    }
                )
                count += 1
            except Exception as e:
                errors += 1
                if errors <= 3:
                    self.stdout.write(self.style.WARNING(f"  Skip attendance: {str(e)[:80]}"))

        self.stdout.write(f"  Imported: {count} attendance records ({errors} errors)")

    # ─── 3. Transport Stops ──────────────────────────────────────────────

    def import_transport_stops(self, wb):
        self.stdout.write("\n3/7 Importing Transport Stops...")
        stops_data = self._get_sheet_data(wb, 'Transport_Stops')
        if not stops_data:
            return

        count = 0
        errors = 0
        self.stop_map = {}  # Excel UUID -> Stop

        for row in stops_data:
            route_excel_id = row.get('route_id')
            route = self.route_map.get(route_excel_id)
            if not route:
                continue

            stop_name = row.get('stop_name', row.get('name', ''))
            if not stop_name:
                continue

            sequence = row.get('sequence', row.get('sequence_order', 1))
            if sequence is None:
                sequence = 1

            pickup_time = row.get('pickup_time', row.get('arrival_time'))
            if isinstance(pickup_time, str):
                try:
                    pickup_time = datetime.strptime(pickup_time, '%H:%M').time()
                except ValueError:
                    pickup_time = time(7, 30)
            elif isinstance(pickup_time, datetime):
                pickup_time = pickup_time.time()
            elif not isinstance(pickup_time, time):
                pickup_time = time(7, 30)

            fare = row.get('fare', row.get('pickup_fare'))
            if fare is not None:
                try:
                    fare = Decimal(str(fare))
                except Exception:
                    fare = None

            try:
                stop, created = Stop.objects.get_or_create(
                    route=route,
                    sequence_order=int(sequence),
                    defaults={
                        'name': str(stop_name),
                        'arrival_time': pickup_time,
                        'pickup_fare': fare,
                    }
                )
                self.stop_map[row.get('id')] = stop
                if created:
                    count += 1
            except Exception as e:
                errors += 1
                if errors <= 3:
                    self.stdout.write(self.style.WARNING(f"  Skip stop: {str(e)[:80]}"))

        self.stdout.write(f"  Imported: {count} stops ({errors} errors)")

    # ─── 4. Transport Allocations ────────────────────────────────────────

    def import_transport_allocations(self, wb):
        self.stdout.write("\n4/7 Importing Transport Allocations...")
        alloc_data = self._get_sheet_data(wb, 'Transport_Allocations')
        if not alloc_data:
            return

        count = 0
        errors = 0
        skipped_existing = 0

        for row in alloc_data:
            student_excel_id = row.get('student_id')
            student = self.student_map.get(student_excel_id)
            if not student:
                continue

            route_excel_id = row.get('route_id')
            route = self.route_map.get(route_excel_id)
            if not route:
                continue

            # OneToOneField — skip if student already has an allocation
            if TransportAllocation.objects.filter(student=student).exists():
                skipped_existing += 1
                continue

            stop = None
            stop_excel_id = row.get('stop_id')
            if stop_excel_id and hasattr(self, 'stop_map'):
                stop = self.stop_map.get(stop_excel_id)

            try:
                TransportAllocation.objects.create(
                    student=student,
                    route=route,
                    stop=stop,
                    start_date=date(2024, 4, 1),
                    is_active=True,
                )
                count += 1
            except Exception as e:
                errors += 1
                if errors <= 3:
                    self.stdout.write(self.style.WARNING(f"  Skip allocation: {str(e)[:80]}"))

        self.stdout.write(
            f"  Imported: {count} allocations ({skipped_existing} already existed, {errors} errors)"
        )

    # ─── 5. Library Book Issues ──────────────────────────────────────────

    def import_library_issues(self, wb):
        self.stdout.write("\n5/7 Importing Library Book Issues...")
        issues_data = self._get_sheet_data(wb, 'Library_Issues')
        if not issues_data:
            return

        count = 0
        errors = 0

        for row in issues_data:
            book_excel_id = row.get('book_id')
            book = self.book_map.get(book_excel_id)
            if not book:
                continue

            student_excel_id = row.get('student_id')
            student = self.student_map.get(student_excel_id)
            # BookIssue requires either student or staff
            if not student:
                continue

            issue_date = row.get('issue_date')
            if isinstance(issue_date, str):
                try:
                    issue_date = datetime.strptime(issue_date, '%Y-%m-%d').date()
                except ValueError:
                    issue_date = date.today()
            elif isinstance(issue_date, datetime):
                issue_date = issue_date.date()
            elif not isinstance(issue_date, date):
                issue_date = date.today()

            return_date = row.get('return_date')
            if isinstance(return_date, str):
                try:
                    return_date = datetime.strptime(return_date, '%Y-%m-%d').date()
                except ValueError:
                    return_date = None
            elif isinstance(return_date, datetime):
                return_date = return_date.date()
            elif not isinstance(return_date, date):
                return_date = None

            due_date = row.get('due_date')
            if isinstance(due_date, str):
                try:
                    due_date = datetime.strptime(due_date, '%Y-%m-%d').date()
                except ValueError:
                    due_date = issue_date + timedelta(days=14)
            elif isinstance(due_date, datetime):
                due_date = due_date.date()
            elif not isinstance(due_date, date):
                due_date = issue_date + timedelta(days=14)

            status = str(row.get('status', 'ISSUED')).upper().strip()
            valid_statuses = {'ISSUED', 'RETURNED', 'OVERDUE', 'LOST'}
            if status not in valid_statuses:
                status = 'ISSUED'
            # If returned, set status
            if return_date and status == 'ISSUED':
                status = 'RETURNED'

            fine_amount = row.get('fine_amount', 0) or 0
            try:
                fine_amount = Decimal(str(fine_amount))
            except Exception:
                fine_amount = Decimal('0')

            try:
                BookIssue.objects.create(
                    book=book,
                    student=student,
                    issue_date=issue_date,
                    due_date=due_date,
                    return_date=return_date,
                    status=status,
                    fine_amount=fine_amount,
                    remarks=str(row.get('remarks', '') or ''),
                )
                count += 1
            except Exception as e:
                errors += 1
                if errors <= 3:
                    self.stdout.write(self.style.WARNING(f"  Skip issue: {str(e)[:80]}"))

        self.stdout.write(f"  Imported: {count} book issues ({errors} errors)")

    # ─── 6. Events ───────────────────────────────────────────────────────

    def import_events(self, wb):
        self.stdout.write("\n6/7 Importing Events...")
        events_data = self._get_sheet_data(wb, 'Events')
        if not events_data:
            return

        count = 0
        errors = 0

        for row in events_data:
            title = row.get('name', row.get('title', ''))
            if not title:
                continue

            event_type = str(row.get('event_type', 'OTHER')).upper().strip()
            valid_types = {'ACADEMIC', 'HOLIDAY', 'EXAM', 'MEETING', 'SPORTS', 'CULTURAL', 'OTHER'}
            if event_type not in valid_types:
                event_type = 'OTHER'

            event_date = row.get('event_date', row.get('start_date'))
            if isinstance(event_date, str):
                try:
                    event_date = datetime.strptime(event_date, '%Y-%m-%d')
                except ValueError:
                    continue
            elif isinstance(event_date, date) and not isinstance(event_date, datetime):
                event_date = datetime.combine(event_date, time(9, 0))
            elif not isinstance(event_date, datetime):
                continue

            end_date = row.get('end_date')
            if isinstance(end_date, str):
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d')
                except ValueError:
                    end_date = event_date + timedelta(hours=8)
            elif isinstance(end_date, date) and not isinstance(end_date, datetime):
                end_date = datetime.combine(end_date, time(17, 0))
            elif not isinstance(end_date, datetime):
                end_date = event_date + timedelta(hours=8)

            description = row.get('description', '') or ''

            try:
                Event.objects.get_or_create(
                    title=str(title),
                    event_type=event_type,
                    defaults={
                        'description': str(description),
                        'start_date': event_date,
                        'end_date': end_date,
                        'is_public': True,
                    }
                )
                count += 1
            except Exception as e:
                errors += 1
                if errors <= 3:
                    self.stdout.write(self.style.WARNING(f"  Skip event: {str(e)[:80]}"))

        self.stdout.write(f"  Imported: {count} events ({errors} errors)")

    # ─── 7. Exam Results ─────────────────────────────────────────────────

    def import_exam_results(self, wb):
        self.stdout.write("\n7/7 Importing Exam Results...")
        results_data = self._get_sheet_data(wb, 'Exam_Results')
        if not results_data:
            return

        if not self.exam_map:
            self.stdout.write(self.style.WARNING("  No exams found, skipping results"))
            return

        academic_year = AcademicYear.objects.filter(is_current=True).first()
        if not academic_year:
            self.stdout.write(self.style.WARNING("  No current academic year, skipping"))
            return

        # Ensure grade scale and grades exist
        grade_scale = GradeScale.objects.filter(name='CBSE Standard').first()
        if not grade_scale:
            self.stdout.write(self.style.WARNING("  No grade scale found, skipping results"))
            return

        count = 0
        errors = 0
        schedules_created = 0

        for row in results_data:
            exam_excel_id = row.get('exam_id')
            exam = self.exam_map.get(exam_excel_id)
            if not exam:
                continue

            student_excel_id = row.get('student_id')
            student = self.student_map.get(student_excel_id)
            if not student:
                continue

            subject_excel_id = row.get('subject_id')
            subject = self.subject_map.get(subject_excel_id)
            if not subject:
                continue

            marks = row.get('marks_obtained')
            if marks is None:
                continue
            try:
                marks = Decimal(str(marks))
            except Exception:
                continue

            # Get student's enrollment to find class/section
            enrollment = StudentEnrollment.objects.filter(student=student).first()
            if not enrollment:
                continue

            section = enrollment.section
            class_obj = section.class_instance

            # Get or create ExamSchedule for this exam+class+section+subject
            try:
                schedule, sch_created = ExamSchedule.objects.get_or_create(
                    examination=exam,
                    class_obj=class_obj,
                    section=section,
                    subject=subject,
                    defaults={
                        'exam_date': exam.start_date,
                        'start_time': time(9, 0),
                        'end_time': time(12, 0),
                        'duration_minutes': 180,
                        'max_marks': Decimal('100'),
                        'min_passing_marks': Decimal('33'),
                    }
                )
                if sch_created:
                    schedules_created += 1
            except Exception as e:
                errors += 1
                if errors <= 3:
                    self.stdout.write(self.style.WARNING(f"  Schedule error: {str(e)[:80]}"))
                continue

            # Create StudentMark
            try:
                StudentMark.objects.get_or_create(
                    exam_schedule=schedule,
                    student=student,
                    defaults={
                        'marks_obtained': marks,
                        'status': 'PRESENT',
                    }
                )
                count += 1
            except Exception as e:
                errors += 1
                if errors <= 3:
                    self.stdout.write(self.style.WARNING(f"  Mark error: {str(e)[:80]}"))

        self.stdout.write(
            f"  Imported: {count} marks, {schedules_created} schedules created ({errors} errors)"
        )

    # ─── Verification ────────────────────────────────────────────────────

    def verify_data(self):
        self.stdout.write("\n--- Data Verification ---")
        checks = [
            ('Students', Student.objects.count()),
            ('Staff', StaffMember.objects.count()),
            ('Student Parents', StudentParent.objects.count()),
            ('Staff Attendance', StaffAttendance.objects.count()),
            ('Transport Stops', Stop.objects.count()),
            ('Transport Allocations', TransportAllocation.objects.count()),
            ('Library Books', Book.objects.count()),
            ('Book Issues', BookIssue.objects.count()),
            ('Events', Event.objects.count()),
        ]
        try:
            checks.append(('Exam Schedules', ExamSchedule.objects.count()))
            checks.append(('Student Marks', StudentMark.objects.count()))
        except Exception:
            pass

        total = 0
        for label, cnt in checks:
            icon = 'OK' if cnt > 0 else '--'
            self.stdout.write(f"  [{icon}] {label:.<30} {cnt:>6}")
            total += cnt
        self.stdout.write(f"  {'Total':.<34} {total:>6}")
