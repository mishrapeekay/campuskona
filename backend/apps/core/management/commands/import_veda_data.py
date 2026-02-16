"""
Import Veda Vidyalaya data directly from Excel using Django ORM
"""
import os
import openpyxl
from datetime import datetime, date
from django.core.management.base import BaseCommand
from django.db import transaction, connection
from apps.authentication.models import User
from apps.tenants.models import School
from apps.students.models import Student
from apps.staff.models import StaffMember
from apps.academics.models import AcademicYear, Board, Class, Section, Subject, StudentEnrollment
from apps.transport.models import Vehicle, Route
from apps.library.models import Book, Author
from apps.communication.models import Notice

class Command(BaseCommand):
    help = 'Import Veda Vidyalaya data from Excel file'

    def handle(self, *args, **options):
        file_path = r"G:\School Mgmt System\mock_data\Veda_files\veda_vidyalaya_complete_data.xlsx"
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        # Get school and switch schema
        try:
            school = School.objects.get(code='VV')
            self.stdout.write(f"Found school: {school.name}")
        except School.DoesNotExist:
            self.stdout.write(self.style.ERROR("School 'Veda Vidyalaya' not found!"))
            return

        # Switch to tenant schema
        with connection.cursor() as cursor:
            cursor.execute(f'SET search_path TO "{school.schema_name}"')
        
        self.stdout.write(f"Switched to schema: {school.schema_name}")

        try:
            # Disconnect signals to prevent audit log issues during import
            from django.db.models.signals import post_save
            from apps.authentication.models import User
            from apps.authentication.signals import log_user_creation
            
            post_save.disconnect(log_user_creation, sender=User)
            self.stdout.write("Disconnected signals")

            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Import without transaction to allow partial success
            self.import_academics(wb)
            self.import_staff(wb)
            self.import_students(wb)
            self.import_transport(wb)
            self.import_library(wb)
            self.import_communication(wb)
            self.import_exams(wb)

            self.stdout.write(self.style.SUCCESS("\n✅ Veda Vidyalaya Import Complete!"))
            self.verify_data()
            
            # Reconnect signals
            post_save.connect(log_user_creation, sender=User)
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Error: {str(e)}"))
            import traceback
            traceback.print_exc()

    def _get_sheet_data(self, wb, sheet_name):
        """Helper to convert sheet to list of dicts"""
        try:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))
            return data
        except KeyError:
            self.stdout.write(self.style.WARNING(f"Sheet '{sheet_name}' not found"))
            return []

    def import_academics(self, wb):
        self.stdout.write("\n📚 Importing Academics...")
        
        # Board - Try to get existing or skip if creation fails
        try:
            board = Board.objects.filter(board_code='CBSE').first()
            if not board:
                board = Board(
                    board_code='CBSE',
                    board_name='Central Board of Secondary Education',
                    board_type='CBSE'
                )
                board.save()
                self.stdout.write("  ✓ Created CBSE board")
            else:
                self.stdout.write("  ✓ Using existing CBSE board")
        except Exception as e:
            self.stdout.write(self.style.WARNING(f"  ⚠️  Board creation skipped: {str(e)}"))
            # Try to get any existing board as fallback
            board = Board.objects.first()
            if not board:
                self.stdout.write(self.style.ERROR("  ❌ No board available, cannot continue"))
                return
        
        # Academic Years
        years_data = self._get_sheet_data(wb, 'Academic_Years')
        for row in years_data:
            if row.get('name'):
                try:
                    year, created = AcademicYear.objects.get_or_create(
                        name=row['name'],
                        defaults={
                            'start_date': row.get('start_date') or date(2024, 4, 1),
                            'end_date': row.get('end_date') or date(2025, 3, 31),
                            'is_current': row.get('is_current', False)
                        }
                    )
                except Exception as e:
                    self.stdout.write(self.style.WARNING(f"  ⚠️  Skipped year {row['name']}: {str(e)}"))
        self.stdout.write(f"  ✓ Academic Years: {AcademicYear.objects.count()}")
        
        # Classes
        classes_data = self._get_sheet_data(wb, 'Classes')
        class_map = {}
        for idx, row in enumerate(classes_data, 1):
            if row.get('name'):
                try:
                    c, _ = Class.objects.get_or_create(
                        name=row['name'],
                        board=board,
                        defaults={
                            'display_name': row.get('display_name', row['name']),
                            'class_order': idx
                        }
                    )
                    class_map[row.get('id')] = c
                except Exception as e:
                    self.stdout.write(self.style.WARNING(f"  ⚠️  Skipped class {row['name']}: {str(e)}"))
        self.stdout.write(f"  ✓ Classes: {Class.objects.count()}")
        
        # Subjects
        subjects_data = self._get_sheet_data(wb, 'Subjects')
        for row in subjects_data:
            if row.get('name') and row.get('code'):
                try:
                    Subject.objects.get_or_create(
                        code=row['code'],
                        defaults={
                            'name': row['name'],
                            'subject_type': 'CORE',
                            'board': board
                        }
                    )
                except Exception as e:
                    self.stdout.write(self.style.WARNING(f"  ⚠️  Skipped subject {row['code']}: {str(e)}"))
        self.stdout.write(f"  ✓ Subjects: {Subject.objects.count()}")
        
        # Sections
        sections_data = self._get_sheet_data(wb, 'Sections')
        current_year = AcademicYear.objects.filter(is_current=True).first()
        self.section_map = {}
        
        for row in sections_data:
            if row.get('name') and row.get('class_id'):
                cls = class_map.get(row['class_id'])
                if cls and current_year:
                    sec, _ = Section.objects.get_or_create(
                        name=row['name'],
                        class_instance=cls,
                        academic_year=current_year,
                        defaults={'max_students': 40}
                    )
                    self.section_map[row.get('id')] = sec
        self.stdout.write(f"  ✓ Sections: {Section.objects.count()}")

    def import_staff(self, wb):
        self.stdout.write("\n👥 Importing Staff...")
        staff_data = self._get_sheet_data(wb, 'Staff_Members')
        
        count = 0
        for row in staff_data:
            emp_id = row.get('employee_id')
            if not emp_id:
                continue
                
            if StaffMember.objects.filter(employee_id=emp_id).exists():
                continue
            
            email = f"{emp_id.lower()}@vedavidyalaya.edu.in"
            
            # Create User
            user, created = User.objects.get_or_create(
                email=email,
                defaults={
                    'first_name': row.get('designation', 'Staff')[:30],
                    'last_name': 'Member',
                    'phone': f"98{str(count).zfill(8)}",
                    'user_type': 'TEACHER' if 'TEACHER' in row.get('designation', '') else 'SCHOOL_ADMIN',
                    'is_staff': True
                }
            )
            if created:
                user.set_password('Veda@123')
                user.save()

            # Create Staff
            StaffMember.objects.get_or_create(
                employee_id=emp_id,
                defaults={
                    'user': user,
                    'first_name': row.get('designation', 'Staff')[:30],
                    'last_name': 'Member',
                    'designation': row.get('designation', 'OTHER'),
                    'department': row.get('department', ''),
                    'employment_type': row.get('employment_type', 'PERMANENT'),
                    'joining_date': row.get('joining_date') or date(2020, 1, 1),
                    'email': email,
                    'phone_number': user.phone,
                    'date_of_birth': date(1990, 1, 1),
                    'gender': 'M',
                    'emergency_contact_number': '9999999999'
                }
            )
            count += 1
            
        self.stdout.write(f"  ✓ Staff Members: {StaffMember.objects.count()}")

    def import_students(self, wb):
        self.stdout.write("\n🎓 Importing Students...")
        students_data = self._get_sheet_data(wb, 'Students')
        enroll_data = self._get_sheet_data(wb, 'Student_Enrollments')
        
        enroll_map = {row.get('student_id'): row for row in enroll_data if row.get('student_id')}
        
        count = 0
        for row in students_data:
            adm_no = row.get('admission_number')
            if not adm_no:
                continue
                
            if Student.objects.filter(admission_number=adm_no).exists():
                continue
            
            email = f"{adm_no.lower()}@vedavidyalaya.edu.in"
            phone = f"99{str(count).zfill(8)}"
            
            # Create User
            user, created = User.objects.get_or_create(
                email=email,
                defaults={
                    'first_name': row.get('first_name', 'Student')[:30],
                    'last_name': row.get('last_name', str(count))[:30],
                    'phone': phone,
                    'user_type': 'STUDENT'
                }
            )
            if created:
                user.set_password('Veda@123')
                user.save()

            # Create Student
            student = Student.objects.create(
                user=user,
                admission_number=adm_no,
                first_name=row.get('first_name', 'Student')[:30],
                last_name=row.get('last_name', str(count))[:30],
                admission_status='ACTIVE',
                admission_date=date(2024, 4, 1),
                date_of_birth=row.get('date_of_birth') or date(2015, 1, 1),
                gender=(row.get('gender') or 'O')[0],
                email=email,
                emergency_contact_number='9999999999',
                father_name='Father',
                father_phone='9999999999',
                mother_name='Mother',
                current_address_line1='Gwalior',
                current_city='Gwalior',
                current_state='MP',
                current_pincode='474001',
                permanent_address_line1='Gwalior',
                permanent_city='Gwalior',
                permanent_state='MP',
                permanent_pincode='474001'
            )

            # Create Enrollment
            enr_row = enroll_map.get(row.get('id'))
            if enr_row and enr_row.get('section_id') in self.section_map:
                section = self.section_map[enr_row['section_id']]
                StudentEnrollment.objects.create(
                    student=student,
                    section=section,
                    academic_year=section.academic_year,
                    roll_number=enr_row.get('roll_number', ''),
                    enrollment_status='ENROLLED',
                    enrollment_date=date(2024, 4, 1)
                )
            
            count += 1
            if count % 100 == 0:
                self.stdout.write(f"  ... {count} students imported")
            
        self.stdout.write(f"  ✓ Students: {Student.objects.count()}")
        self.stdout.write(f"  ✓ Enrollments: {StudentEnrollment.objects.count()}")

    def import_transport(self, wb):
        self.stdout.write("\n🚌 Importing Transport...")
        vehicles = self._get_sheet_data(wb, 'Transport_Vehicles')
        
        for row in vehicles:
            if row.get('registration_number'):
                Vehicle.objects.get_or_create(
                    registration_number=row['registration_number'],
                    defaults={
                        'capacity': row.get('capacity', 40),
                        'status': 'ACTIVE',
                        'model': 'Tata Starbus'
                    }
                )
        
        self.stdout.write(f"  ✓ Vehicles: {Vehicle.objects.count()}")
        
        routes = self._get_sheet_data(wb, 'Transport_Routes')
        for row in routes:
            if row.get('name'):
                Route.objects.get_or_create(
                    name=row['name'],
                    defaults={
                        'start_point': 'School',
                        'end_point': 'City',
                        'fare': 1500.00
                    }
                )
        
        self.stdout.write(f"  ✓ Routes: {Route.objects.count()}")

    def import_library(self, wb):
        self.stdout.write("\n📖 Importing Library...")
        books = self._get_sheet_data(wb, 'Library_Books')
        
        for row in books:
            if row.get('isbn'):
                author_name = row.get('author', 'Unknown Author')
                author, _ = Author.objects.get_or_create(name=author_name)
                
                # Strip hyphens for ISBN-13
                isbn = str(row['isbn']).replace('-', '')[:13]
                
                Book.objects.get_or_create(
                    isbn=isbn,
                    defaults={
                        'title': row.get('title', 'Untitled'),
                        'author': author,
                        'quantity': row.get('total_copies', 1),
                        'available_copies': row.get('available_copies', 1)
                    }
                )
        
        self.stdout.write(f"  ✓ Books: {Book.objects.count()}")

    def import_communication(self, wb):
        self.stdout.write("\n📢 Importing Communications...")
        notices = self._get_sheet_data(wb, 'Notices')
        
        for row in notices:
            if row.get('title'):
                Notice.objects.get_or_create(
                    title=row['title'],
                    defaults={
                        'content': row.get('content', ''),
                        'target_audience': row.get('target_audience', 'ALL'),
                        'is_published': True
                    }
                )
        
        self.stdout.write(f"  ✓ Notices: {Notice.objects.count()}")

    def import_exams(self, wb):
        self.stdout.write("\n📝 Importing Examinations...")
        from apps.examinations.models import GradeScale, Grade, ExamType, Examination, ExamSchedule, StudentMark
        from datetime import time

        # 1. Setup Defaults
        grade_scale, _ = GradeScale.objects.get_or_create(name='CBSE Standard')
        if not Grade.objects.filter(grade_scale=grade_scale).exists():
            Grade.objects.create(grade_scale=grade_scale, grade='A1', min_percentage=91, max_percentage=100, grade_point=10)
            Grade.objects.create(grade_scale=grade_scale, grade='A2', min_percentage=81, max_percentage=90, grade_point=9)
            Grade.objects.create(grade_scale=grade_scale, grade='B1', min_percentage=71, max_percentage=80, grade_point=8)
            Grade.objects.create(grade_scale=grade_scale, grade='B2', min_percentage=61, max_percentage=70, grade_point=7)
            # Add more grades as needed...

        exam_type, _ = ExamType.objects.get_or_create(
            code='TERM', 
            defaults={'name': 'Term Exam', 'exam_type': 'SUMMATIVE'}
        )
        
        academic_year = AcademicYear.objects.filter(is_current=True).first()
        if not academic_year: 
            return

        # 2. Import Exams
        exams_data = self._get_sheet_data(wb, 'Exams')
        exam_map = {}
        for row in exams_data:
            if row.get('name'):
                exam, _ = Examination.objects.get_or_create(
                    name=row['name'],
                    academic_year=academic_year,
                    defaults={
                        'exam_type': exam_type,
                        'grade_scale': grade_scale,
                        'start_date': date(2025, 3, 1),
                        'end_date': date(2025, 3, 15),
                        'status': 'COMPLETED',
                        'is_published': True
                    }
                )
                exam_map[row.get('id')] = exam
        
        self.stdout.write(f"  ✓ Exams: {Examination.objects.count()}")

        # 3. Import Results
        # Note: This requires complex mapping. We'll verify what we have first.
        results_data = self._get_sheet_data(wb, 'Exam_Results')
        
        self.stdout.write(f"  ℹ️  Exam Results import skipped (ID mismatch prevention). Structure created.")

    def verify_data(self):
        self.stdout.write("\n📊 Data Summary:")
        self.stdout.write(f"  • Academic Years: {AcademicYear.objects.count()}")
        self.stdout.write(f"  • Classes: {Class.objects.count()}")
        self.stdout.write(f"  • Sections: {Section.objects.count()}")
        self.stdout.write(f"  • Subjects: {Subject.objects.count()}")
        self.stdout.write(f"  • Students: {Student.objects.count()}")
        self.stdout.write(f"  • Staff: {StaffMember.objects.count()}")
        self.stdout.write(f"  • Enrollments: {StudentEnrollment.objects.count()}")
        self.stdout.write(f"  • Vehicles: {Vehicle.objects.count()}")
        self.stdout.write(f"  • Routes: {Route.objects.count()}")
        try:
            import apps.examinations.models as exams
            self.stdout.write(f"  • Exams: {exams.Examination.objects.count()}")
        except ImportError:
            pass
        self.stdout.write(f"  • Books: {Book.objects.count()}")
        self.stdout.write(f"  • Notices: {Notice.objects.count()}")
