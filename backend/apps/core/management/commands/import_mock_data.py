import os
import openpyxl
from datetime import datetime
from django.core.management.base import BaseCommand
from django.db import transaction, connection
from django.conf import settings
from apps.authentication.models import User
from apps.tenants.models import School
from apps.students.models import Student, StudentParent
from apps.staff.models import StaffMember
from apps.academics.models import AcademicYear, Board, Class, Section, Subject, StudentEnrollment

class Command(BaseCommand):
    help = 'Import mock data from Excel files'

    def add_arguments(self, parser):
        parser.add_argument('--school-name', type=str, default='Demo High School', help='Name of the school to import data into')
        parser.add_argument('--reset', action='store_true', help='Delete existing data before import')

    def handle(self, *args, **options):
        school_name = options['school_name']
        reset = options['reset']
        
        base_dir = r"G:\School Mgmt System\mock_data\files"
        students_file = os.path.join(base_dir, "students_mock_data.xlsx")
        staff_file = os.path.join(base_dir, "teachers_staff_mock_data.xlsx")

        if not os.path.exists(students_file) or not os.path.exists(staff_file):
            self.stdout.write(self.style.ERROR(f"Mock data files not found in {base_dir}"))
            return

        # 1. Setup School context
        school = self.setup_school(school_name)
        if not school:
            return

        # Explicitly set tenant for connection if using django-tenants
        if hasattr(connection, 'set_tenant'):
            connection.set_tenant(school)
        
        self.stdout.write(self.style.SUCCESS(f"Using School: {school.name} (Schema: {school.schema_name})"))

        with transaction.atomic():
            if reset:
                self.reset_data()

            # 2. Setup Academics (Year, Board, Classes, Sections)
            academic_year = self.setup_academics(students_file)
            
            # 3. Import Staff
            self.import_staff(staff_file)
            
            # 4. Import Students
            self.import_students(students_file, academic_year)
            
        self.stdout.write(self.style.SUCCESS("Import completed successfully!"))

    def setup_school(self, name):
        # Check if we should use public schema or specific tenant
        # For now, just find or create the school record
        school, created = School.objects.get_or_create(
            name=name,
            defaults={
                'code': 'DEMO01',
                'subdomain': 'demo',
                'schema_name': 'public', # Assuming development environment uses public
                'email': 'admin@demo.com',
                'primary_board': 'CBSE',
                'is_active': True
            }
        )
        if created:
            self.stdout.write(f"Created new school: {name}")
        else:
            self.stdout.write(f"Found existing school: {name}")
        return school

    def reset_data(self):
        self.stdout.write("Resetting data...")
        StudentEnrollment.objects.all().delete()
        Student.objects.all().delete()
        StaffMember.objects.all().delete()
        # Be careful deleting Users. Only delete non-superusers?
        # User.objects.exclude(is_superuser=True).delete() 
        # Actually safer to verify by user_type
        User.objects.filter(user_type__in=['STUDENT', 'TEACHER', 'PARENT']).delete()
        
    def setup_academics(self, students_file):
        self.stdout.write("Setting up Academics...")
        
        # Board
        board, _ = Board.objects.get_or_create(
            board_type='CBSE',
            defaults={
                'board_name': 'Central Board of Secondary Education',
                'board_code': 'CBSE',
                'is_active': True
            }
        )

        # Academic Year
        academic_year, _ = AcademicYear.objects.get_or_create(
            name='2024-2025',
            defaults={
                'start_date': '2024-04-01',
                'end_date': '2025-03-31',
                'is_current': True
            }
        )

        # Classes & Sections from Excel
        wb = openpyxl.load_workbook(students_file, read_only=True)
        ws = wb['All Students']
        headers = [cell.value for cell in ws[1]]
        
        class_idx = headers.index('Class')
        section_idx = headers.index('Section')
        
        # Collect unique class-section pairs
        pairs = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[class_idx] and row[section_idx]:
                c_val = str(row[class_idx]).strip()
                s_val = str(row[section_idx]).strip()
                pairs.add((c_val, s_val))
        
        unique_classes = sorted(list(set([p[0] for p in pairs])))
        
        def get_order(c):
            c = str(c).upper()
            if 'LKG' in c: return -2
            if 'UKG' in c: return -1
            if 'NURSERY' in c: return -3
            try:
                val = c.replace('CLASS', '').replace('_', '').replace(' ', '').replace('GRADE', '')
                return int(val)
            except:
                return 100

        # Sort based on order
        unique_classes.sort(key=get_order)

        # Assign orders with conflict resolution
        used_orders = set()
        
        for c_name in unique_classes:
            target_order = get_order(c_name)
            while target_order in used_orders:
                target_order += 1
            
            used_orders.add(target_order)
            
            # Create Class
            clean_name = str(c_name).replace('Class_', '')
            
            class_obj, _ = Class.objects.get_or_create(
                name=clean_name,
                board=board,
                defaults={
                    'display_name': c_name.replace('_', ' '),
                    'class_order': target_order
                }
            )
            
            # Create Sections for this class
            class_sections = sorted(list(set([p[1] for p in pairs if p[0] == c_name])))
            for sec_name in class_sections:
                Section.objects.get_or_create(
                    class_instance=class_obj,
                    name=sec_name,
                    academic_year=academic_year,
                    defaults={
                        'max_students': 40
                    }
                )

        self.stdout.write(f"Setup {len(unique_classes)} classes and sections.")
        return academic_year

    def import_staff(self, staff_file):
        self.stdout.write("Importing Staff...")
        wb = openpyxl.load_workbook(staff_file, read_only=True)
        ws = wb['All Staff']
        headers = [cell.value for cell in ws[1]]
        
        # Map headers
        h = {name: i for i, name in enumerate(headers)}
        
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            email = row[h['Email']]
            emp_id = row[h['Employee_ID']]
            
            if not email:
                continue
                
            if StaffMember.objects.filter(employee_id=emp_id).exists():
                continue
                
            # Create User
            user, created = User.objects.get_or_create(
                email=email,
                defaults={
                    'first_name': row[h['First_Name']],
                    'last_name': row[h['Last_Name']],
                    'phone': str(row[h['Phone']]).replace('-', '').replace(' ', '')[:10],
                    'user_type': 'TEACHER' if 'Teacher' in (row[h['Designation']] or '') else 'SCHOOL_ADMIN',
                    'is_staff': True
                }
            )
            if created:
                user.set_password('School@123')
                user.save()
            
            # Check if user already has staff profile (OneToOne constraint)
            if StaffMember.objects.filter(user=user).exists():
                self.stdout.write(f"Skipping duplicate staff profile for {email}")
                continue

            # Create Staff Profile
            StaffMember.objects.create(
                user=user,
                employee_id=emp_id,
                first_name=row[h['First_Name']],
                last_name=row[h['Last_Name']],
                email=email,
                phone_number=user.phone,
                designation=row[h['Designation']].upper().replace(' ', '_') if row[h['Designation']] else 'OTHER',
                department=row[h['Department']] or '',
                joining_date=row[h['Joining_Date']] or datetime.now().date(),
                date_of_birth=row[h['Date_of_Birth']] or datetime(1990, 1, 1).date(),
                gender=row[h['Gender']][0] if row[h['Gender']] else 'O',
                current_address_line1=row[h['Address']] or ''
            )
            count += 1
            
        self.stdout.write(f"Imported {count} staff members.")

    def import_students(self, students_file, academic_year):
        self.stdout.write("Importing Students...")
        wb = openpyxl.load_workbook(students_file, read_only=True)
        ws = wb['All Students']
        headers = [cell.value for cell in ws[1]]
        
        h = {name: i for i, name in enumerate(headers)}
        
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            adm_no = row[h['Admission_Number']]
            if not adm_no or Student.objects.filter(admission_number=adm_no).exists():
                continue
            
            # Generate student email (no Email column in file)
            email = f"{adm_no.lower()}@demo.com"
            
            # Create User
            try:
                user, created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        'first_name': row[h['First_Name']],
                        'last_name': row[h['Last_Name']],
                        'phone': str(row[h['Father_Phone']]).replace('-', '')[:10] if row[h['Father_Phone']] else '',
                        'user_type': 'STUDENT'
                    }
                )
            except Exception:
                # Handle dupe email by appending adm_no
                email = f"{adm_no.lower()}_{email}"
                user, created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        'first_name': row[h['First_Name']],
                        'last_name': row[h['Last_Name']],
                        'user_type': 'STUDENT'
                    }
                )

            if created:
                user.set_password('School@123')
                user.save()

            # Find Section
            c_name = str(row[h['Class']]).replace('Class_', '')
            s_name = row[h['Section']]
            
            try:
                # Need to find class by display name logic basically
                # Or query by name if I cleaned it
                # Logic used in setup_academics: Class_5 -> name='5'
                cls_obj = Class.objects.filter(name=c_name).first()
                if not cls_obj:
                    # fallback if logic mismatches
                    continue
                    
                section_obj = Section.objects.get(
                    class_instance=cls_obj,
                    name=s_name,
                    academic_year=academic_year
                )
            except Exception as e:
                print(f"Skipping {row[h['First_Name']]}: Section note found {c_name} {s_name}")
                continue

            # Create Student
            student = Student.objects.create(
                user=user,
                admission_number=adm_no,
                first_name=row[h['First_Name']],
                last_name=row[h['Last_Name']],
                date_of_birth=row[h['Date_of_Birth']] or datetime(2015,1,1).date(),
                gender=row[h['Gender']][0] if row[h['Gender']] else 'O',
                blood_group=row[h['Blood_Group']] or '',
                admission_date=datetime.now().date(),
                admission_status='ACTIVE',
                father_name=row[h['Father_Name']] or '',
                father_phone=str(row[h['Father_Phone']])[:10] if row[h['Father_Phone']] else '',
                father_email=row[h['Father_Email']] or '',
                mother_name=row[h['Mother_Name']] or '',
                current_address_line1=row[h['Address']] or '',
                current_city=row[h['City']] or '',
                current_state=row[h['State']] or '',
                current_pincode=str(row[h['PIN_Code']])[:6] if row[h['PIN_Code']] else ''
            )

            # Create Enrollment
            StudentEnrollment.objects.create(
                student=student,
                section=section_obj,
                academic_year=academic_year,
                roll_number=row[h['Roll_Number']] or '',
                enrollment_date=datetime.now().date(),
                is_active=True
            )
            count += 1
            
        self.stdout.write(f"Imported {count} students.")
