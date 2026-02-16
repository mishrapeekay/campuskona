"""
Partner Payout Report Generator

Generates detailed payout reports for partners in PDF and Excel formats.
"""

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Partner, Payout, Commission


class PayoutReportGenerator:
    """
    Generate payout reports for partners.
    """
    
    @staticmethod
    def generate_excel_report(payout_id):
        """
        Generate Excel payout report.
        
        Args:
            payout_id: UUID of the payout
            
        Returns:
            HttpResponse with Excel file
        """
        payout = Payout.objects.select_related('partner').get(id=payout_id)
        commissions = payout.commissions.select_related('lead', 'school').all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payout Report"
        
        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "PARTNER COMMISSION PAYOUT REPORT"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Payout Details
        row = 3
        ws[f'A{row}'] = "Payout Number:"
        ws[f'B{row}'] = payout.payout_number
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Partner Name:"
        ws[f'B{row}'] = payout.partner.name
        
        row += 1
        ws[f'A{row}'] = "Partner Code:"
        ws[f'B{row}'] = payout.partner.partner_code
        
        row += 1
        ws[f'A{row}'] = "Period:"
        ws[f'B{row}'] = f"{payout.period_start} to {payout.period_end}"
        
        row += 1
        ws[f'A{row}'] = "Payout Date:"
        ws[f'B{row}'] = str(payout.payout_date) if payout.payout_date else "Pending"
        
        row += 1
        ws[f'A{row}'] = "Payment Method:"
        ws[f'B{row}'] = payout.get_payment_method_display()
        
        # Commission Details Header
        row += 2
        headers = [
            'S.No', 'School Name', 'Lead Contact', 'Commission Type',
            'Subscription Amount', 'Commission Amount', 'Earned Date', 'Status'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Commission Data
        row += 1
        for idx, commission in enumerate(commissions, start=1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=commission.school.name)
            ws.cell(row=row, column=3, value=commission.lead.contact_person)
            ws.cell(row=row, column=4, value=commission.get_commission_type_display())
            ws.cell(row=row, column=5, value=float(commission.subscription_amount))
            ws.cell(row=row, column=6, value=float(commission.commission_amount))
            ws.cell(row=row, column=7, value=str(commission.earned_date))
            ws.cell(row=row, column=8, value=commission.get_status_display())
            
            # Apply border to all cells
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
            
            row += 1
        
        # Summary Section
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "PAYOUT SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row += 1
        ws[f'A{row}'] = "Total Commissions:"
        ws[f'B{row}'] = payout.commission_count
        
        row += 1
        ws[f'A{row}'] = "Gross Amount:"
        ws[f'B{row}'] = f"₹{payout.total_amount:,.2f}"
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = f"TDS ({payout.tds_percentage}%):"
        ws[f'B{row}'] = f"₹{payout.tds_amount:,.2f}"
        
        row += 1
        ws[f'A{row}'] = "Net Payable Amount:"
        ws[f'B{row}'] = f"₹{payout.net_amount:,.2f}"
        ws[f'B{row}'].font = Font(bold=True, size=12, color="006100")
        ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Bank Details
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "BANK DETAILS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row += 1
        ws[f'A{row}'] = "Bank Name:"
        ws[f'B{row}'] = payout.partner.bank_name or "N/A"
        
        row += 1
        ws[f'A{row}'] = "Account Number:"
        ws[f'B{row}'] = payout.partner.account_number or "N/A"
        
        row += 1
        ws[f'A{row}'] = "IFSC Code:"
        ws[f'B{row}'] = payout.partner.ifsc_code or "N/A"
        
        row += 1
        ws[f'A{row}'] = "PAN Number:"
        ws[f'B{row}'] = payout.partner.pan_number or "N/A"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="payout_{payout.payout_number}.xlsx"'
        
        return response
    
    @staticmethod
    def generate_partner_summary_report(partner_id, year=None):
        """
        Generate annual summary report for a partner.
        
        Args:
            partner_id: UUID of the partner
            year: Year for the report (default: current year)
            
        Returns:
            HttpResponse with Excel file
        """
        partner = Partner.objects.get(id=partner_id)
        
        if not year:
            year = datetime.now().year
        
        # Get all payouts for the year
        payouts = Payout.objects.filter(
            partner=partner,
            created_at__year=year
        ).order_by('created_at')
        
        # Get all commissions for the year
        commissions = Commission.objects.filter(
            partner=partner,
            earned_date__year=year
        )
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Summary {year}"
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"PARTNER ANNUAL SUMMARY REPORT - {year}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Partner Details
        row = 3
        ws[f'A{row}'] = "Partner Name:"
        ws[f'B{row}'] = partner.name
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Partner Code:"
        ws[f'B{row}'] = partner.partner_code
        
        row += 1
        ws[f'A{row}'] = "Email:"
        ws[f'B{row}'] = partner.email
        
        # Statistics
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "PERFORMANCE STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        
        row += 1
        stats = [
            ("Total Leads", partner.total_leads),
            ("Total Conversions", partner.total_conversions),
            ("Conversion Rate", f"{partner.conversion_rate}%"),
            ("Total Commission Earned", f"₹{partner.total_commission_earned:,.2f}"),
            ("Total Commission Paid", f"₹{partner.total_commission_paid:,.2f}"),
            ("Pending Commission", f"₹{partner.pending_commission:,.2f}"),
        ]
        
        for stat_name, stat_value in stats:
            ws[f'A{row}'] = stat_name
            ws[f'B{row}'] = stat_value
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        # Payouts Table
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "PAYOUTS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        
        row += 1
        payout_headers = ['Payout Number', 'Period', 'Gross Amount', 'TDS', 'Net Amount', 'Status']
        for col, header in enumerate(payout_headers, start=1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
        
        row += 1
        for payout in payouts:
            ws.cell(row=row, column=1, value=payout.payout_number)
            ws.cell(row=row, column=2, value=f"{payout.period_start} to {payout.period_end}")
            ws.cell(row=row, column=3, value=f"₹{payout.total_amount:,.2f}")
            ws.cell(row=row, column=4, value=f"₹{payout.tds_amount:,.2f}")
            ws.cell(row=row, column=5, value=f"₹{payout.net_amount:,.2f}")
            ws.cell(row=row, column=6, value=payout.get_status_display())
            row += 1
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="partner_summary_{partner.partner_code}_{year}.xlsx"'
        
        return response
