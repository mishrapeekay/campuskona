"""
CampusKona — School Onboarding API
Implements the "Magic Excel" strategy: bulk import of classes, sections,
fee structures, students and staff from a single master workbook.
"""
import io
import csv
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# Helper: parse rows from an openpyxl worksheet
# ─────────────────────────────────────────────────────────────

def _ws_to_rows(ws):
    """Convert an openpyxl worksheet to list-of-dicts (normalised keys)."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    headers = [
        str(h).strip().lower().replace(' ', '_') if h else ''
        for h in all_rows[0]
    ]
    rows = [
        dict(zip(headers, row))
        for row in all_rows[1:]
        if any(cell is not None and str(cell).strip() for cell in row)
    ]
    return headers, rows


def _parse_file(file_obj):
    """Parse uploaded Excel/CSV → (headers, rows). Returns None on error."""
    filename = file_obj.name.lower()
    try:
        if filename.endswith(('.xlsx', '.xls')):
            wb = load_workbook(file_obj, read_only=True, data_only=True)
            return _ws_to_rows(wb.active)
        elif filename.endswith('.csv'):
            content = file_obj.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            headers = [h.strip().lower().replace(' ', '_') for h in (reader.fieldnames or [])]
            return headers, list(reader)
        else:
            return None, None
    except Exception:
        return None, None


def _val(row, *keys):
    """Extract first non-empty value from row by trying multiple keys."""
    for key in keys:
        v = row.get(key)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ''


def _bool_val(raw, default=True):
    """'yes'/'no'/'true'/'false'/1/0 → bool."""
    if raw is None:
        return default
    s = str(raw).strip().lower()
    return s not in ('no', 'false', '0', 'n')


def _styled_header_row(ws, headers, color_hex):
    """Write headers to row 1 with coloured background."""
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    return ws


def _excel_response(wb, filename):
    """Return an HttpResponse with Excel content."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ─────────────────────────────────────────────────────────────
# Readiness Score
# ─────────────────────────────────────────────────────────────

PHASES = [
    (0,  'Phase 0 — Not started'),
    (10, 'Week 1 — Foundation (Academic Year)'),
    (20, 'Week 1 — Foundation (Classes)'),
    (30, 'Week 1 — Foundation (Sections)'),
    (40, 'Week 1 — Foundation (Subjects)'),
    (50, 'Week 2 — Staff Live'),
    (65, 'Week 3 — Students enrolled'),
    (75, 'Week 5 — Fee structure active'),
    (85, 'Week 3 — Timetable set up'),
    (95, 'Week 4 — Parents engaged'),
    (100, 'Week 6 — Fully operational'),
]


def _current_phase(score):
    phase = PHASES[0][1]
    for threshold, label in PHASES:
        if score >= threshold:
            phase = label
    return phase


def _grade(score):
    if score >= 90: return 'A'
    if score >= 70: return 'B'
    if score >= 50: return 'C'
    if score >= 30: return 'D'
    return 'F'


class OnboardingReadinessView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.academics.models import AcademicYear, Class, Section, Subject
        from apps.staff.models import StaffMember
        from apps.students.models import Student
        from apps.finance.models import FeeStructure
        from apps.timetable.models import ClassTimetable
        from apps.privacy.models import ParentalConsent

        score = 0
        checks = {}

        # 1. Academic Year (10 pts)
        ay = AcademicYear.objects.filter(is_current=True, is_deleted=False).first()
        if ay:
            score += 10
            checks['academic_year'] = {'status': 'ok', 'detail': ay.name}
        else:
            checks['academic_year'] = {'status': 'missing', 'detail': 'No current academic year set'}

        # 2. Classes (10 pts)
        class_count = Class.objects.filter(is_deleted=False, is_active=True).count()
        if class_count >= 1:
            score += 10
            checks['classes'] = {'status': 'ok', 'count': class_count}
        else:
            checks['classes'] = {'status': 'missing', 'count': 0, 'detail': 'No classes created'}

        # 3. Sections (10 pts)
        section_count = Section.objects.filter(is_deleted=False, is_active=True).count()
        if section_count >= 1:
            score += 10
            checks['sections'] = {'status': 'ok', 'count': section_count}
        else:
            checks['sections'] = {'status': 'missing', 'count': 0, 'detail': 'No sections created'}

        # 4. Subjects (5 pts)
        subject_count = Subject.objects.filter(is_deleted=False, is_active=True).count()
        if subject_count >= 1:
            score += 5
            checks['subjects'] = {'status': 'ok', 'count': subject_count}
        else:
            checks['subjects'] = {'status': 'missing', 'count': 0, 'detail': 'No subjects created'}

        # 5. Staff (10 pts)
        staff_count = StaffMember.objects.filter(is_deleted=False).count()
        if staff_count >= 1:
            score += 10
            checks['staff'] = {'status': 'ok', 'count': staff_count}
        else:
            checks['staff'] = {'status': 'missing', 'count': 0, 'detail': 'No staff uploaded'}

        # 6. Students (15 pts)
        student_count = Student.objects.filter(is_deleted=False).count()
        if student_count >= 1:
            score += 15
            checks['students'] = {'status': 'ok', 'count': student_count}
        else:
            checks['students'] = {'status': 'missing', 'count': 0, 'detail': 'No students uploaded'}

        # 7. Fee Structure (10 pts)
        fee_count = FeeStructure.objects.filter(is_active=True).count()
        if fee_count >= 1:
            score += 10
            checks['fee_structure'] = {'status': 'ok', 'count': fee_count}
        else:
            checks['fee_structure'] = {'status': 'missing', 'count': 0, 'detail': 'No fee structure defined'}

        # 8. Timetable (10 pts)
        tt_count = ClassTimetable.objects.filter(is_active=True).count()
        if tt_count >= 1:
            score += 10
            checks['timetable'] = {'status': 'ok', 'count': tt_count}
        else:
            checks['timetable'] = {'status': 'missing', 'count': 0, 'detail': 'No timetable slots created'}

        # 9. Parent consent (10 pts) — partial credit at 30%+
        consent_count = ParentalConsent.objects.filter(consent_given=True).count()
        if student_count > 0:
            consent_pct = (consent_count / student_count) * 100
            if consent_pct >= 50:
                score += 10
                checks['parent_consent'] = {'status': 'ok', 'rate': f'{consent_pct:.0f}%'}
            elif consent_pct >= 30:
                score += 5
                checks['parent_consent'] = {'status': 'warning', 'rate': f'{consent_pct:.0f}%', 'detail': 'Below 50% target'}
            else:
                checks['parent_consent'] = {'status': 'missing', 'rate': f'{consent_pct:.0f}%', 'detail': 'Consent collection not started'}
        else:
            checks['parent_consent'] = {'status': 'skipped', 'detail': 'No students yet'}

        # 10. DPDP (10 pts — simple proxy: any consent record exists)
        if consent_count >= 1 or student_count == 0:
            score += 10
            checks['dpdp_compliance'] = {'status': 'ok', 'detail': 'DPDP consent workflow active'}
        else:
            checks['dpdp_compliance'] = {'status': 'missing', 'detail': 'DPA and consent collection not started'}

        # Determine next action
        next_action, next_url = _determine_next_action(checks)

        return Response({
            'score': score,
            'grade': _grade(score),
            'phase': _current_phase(score),
            'ready_for_operations': score >= 65,
            'checks': checks,
            'next_action': next_action,
            'next_action_url': next_url,
        })


def _determine_next_action(checks):
    priorities = [
        ('academic_year', 'Set up the current academic year', '/settings'),
        ('classes', 'Upload classes and sections', '/onboarding'),
        ('sections', 'Upload sections for your classes', '/onboarding'),
        ('staff', 'Upload your teaching and admin staff', '/staff/bulk-upload'),
        ('students', 'Upload your student records', '/students/bulk-upload'),
        ('fee_structure', 'Define fee structure for your classes', '/onboarding'),
        ('timetable', 'Create the class timetable', '/timetable/manage'),
        ('parent_consent', 'Start parent consent collection via WhatsApp', '/communication/notices'),
        ('dpdp_compliance', 'Complete DPDP DPA signing process', '/compliance/dashboard'),
    ]
    for key, action, url in priorities:
        check = checks.get(key, {})
        if check.get('status') in ('missing', 'warning'):
            return action, url
    return 'Your school is fully operational!', '/dashboard'


# ─────────────────────────────────────────────────────────────
# Classes + Sections
# ─────────────────────────────────────────────────────────────

CS_HEADERS = [
    'class_name', 'class_order', 'board_name',
    'section_name', 'class_teacher_employee_id', 'max_students', 'room_number',
]

CS_SAMPLE_ROWS = [
    ('LKG',      -2,  'CBSE', 'A', '',        30, '001'),
    ('LKG',      -2,  'CBSE', 'B', '',        30, '002'),
    ('UKG',      -1,  'CBSE', 'A', '',        30, '003'),
    ('UKG',      -1,  'CBSE', 'B', '',        30, '004'),
    ('Class 1',   1,  'CBSE', 'A', 'EMP001',  40, '101'),
    ('Class 1',   1,  'CBSE', 'B', '',        40, '102'),
    ('Class 2',   2,  'CBSE', 'A', 'EMP002',  40, '201'),
    ('Class 2',   2,  'CBSE', 'B', '',        40, '202'),
    ('Class 3',   3,  'CBSE', 'A', '',        40, '301'),
    ('Class 4',   4,  'CBSE', 'A', '',        40, '401'),
    ('Class 5',   5,  'CBSE', 'A', '',        42, '501'),
    ('Class 6',   6,  'CBSE', 'A', 'EMP003',  42, '601'),
    ('Class 6',   6,  'CBSE', 'B', '',        42, '602'),
    ('Class 7',   7,  'CBSE', 'A', '',        42, '701'),
    ('Class 7',   7,  'CBSE', 'B', '',        42, '702'),
    ('Class 8',   8,  'CBSE', 'A', '',        42, '801'),
    ('Class 9',   9,  'CBSE', 'A', 'EMP004',  45, '901'),
    ('Class 9',   9,  'CBSE', 'B', '',        45, '902'),
    ('Class 10', 10,  'CBSE', 'A', 'EMP005',  45, '1001'),
    ('Class 10', 10,  'CBSE', 'B', '',        45, '1002'),
    ('Class 11', 11,  'CBSE', 'A', '',        40, '1101'),
    ('Class 11', 11,  'CBSE', 'B', '',        40, '1102'),
    ('Class 12', 12,  'CBSE', 'A', '',        40, '1201'),
    ('Class 12', 12,  'CBSE', 'B', '',        40, '1202'),
]


class ClassSectionTemplateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Classes_Sections'
        _styled_header_row(ws, CS_HEADERS, '1565C0')
        for row in CS_SAMPLE_ROWS:
            ws.append(row)
        return _excel_response(wb, 'classes_sections_template.xlsx')


def _process_class_section_rows(rows, request_or_schema=None):
    """Core logic shared by ClassSectionBulkUploadView and MasterUploadView."""
    from apps.academics.models import AcademicYear, Class, Section, Board
    from apps.staff.models import StaffMember
    from django.db import transaction

    schema = ''
    if hasattr(request_or_schema, 'tenant'):
        schema = request_or_schema.tenant.schema_name

    classes_created = 0
    classes_existing = 0
    sections_created = 0
    sections_existing = 0
    warnings = []
    errors = []

    ay = AcademicYear.objects.filter(is_current=True, is_deleted=False).first()
    if not ay:
        return {
            'classes_created': 0, 'classes_existing': 0,
            'sections_created': 0, 'sections_existing': 0,
            'warnings': [], 'errors': [{'error': 'No current academic year found. Create one first.'}],
        }

    # Resolve default board (first active or named)
    default_board = Board.objects.filter(is_deleted=False).first()

    for row_idx, row in enumerate(rows, start=2):
        class_name = _val(row, 'class_name', 'class', 'grade')
        section_name = _val(row, 'section_name', 'section', 'division')

        if not class_name or not section_name:
            errors.append({'row': row_idx, 'error': 'class_name and section_name are required'})
            continue

        # Resolve class_order
        try:
            class_order = int(_val(row, 'class_order', 'order') or 0)
        except ValueError:
            class_order = 0

        # Resolve board
        board_name = _val(row, 'board_name', 'board')
        board = default_board
        if board_name:
            b = Board.objects.filter(
                is_deleted=False
            ).filter(
                name__iexact=board_name
            ).first() or Board.objects.filter(
                is_deleted=False, board_code__iexact=board_name
            ).first()
            if b:
                board = b
            else:
                warnings.append({'row': row_idx, 'warning': f'Board "{board_name}" not found, using default board'})

        if not board:
            errors.append({'row': row_idx, 'error': 'No board available. Create a board first.'})
            continue

        try:
            with transaction.atomic():
                cls, created = Class.objects.get_or_create(
                    name=class_name,
                    board=board,
                    defaults={
                        'display_name': class_name,
                        'class_order': class_order,
                        'is_active': True,
                    }
                )
                if created:
                    classes_created += 1
                else:
                    classes_existing += 1

                max_students = 40
                try:
                    max_students = int(_val(row, 'max_students', 'strength') or 40)
                except ValueError:
                    pass

                room_number = _val(row, 'room_number', 'room')

                section, sec_created = Section.objects.get_or_create(
                    class_instance=cls,
                    name=section_name.upper(),
                    academic_year=ay,
                    defaults={
                        'max_students': max_students,
                        'room_number': room_number,
                        'is_active': True,
                    }
                )
                if sec_created:
                    sections_created += 1
                else:
                    sections_existing += 1

                # Assign class teacher if provided
                teacher_emp_id = _val(row, 'class_teacher_employee_id', 'class_teacher', 'teacher_id')
                if teacher_emp_id:
                    teacher = StaffMember.objects.filter(
                        employee_id=teacher_emp_id, is_deleted=False
                    ).first()
                    if teacher:
                        section.class_teacher = teacher
                        section.save(update_fields=['class_teacher'])
                    else:
                        warnings.append({
                            'row': row_idx,
                            'warning': f'Teacher with employee_id "{teacher_emp_id}" not found'
                        })

        except Exception as e:
            errors.append({'row': row_idx, 'error': str(e)})

    return {
        'classes_created': classes_created,
        'classes_existing': classes_existing,
        'sections_created': sections_created,
        'sections_existing': sections_existing,
        'warnings': warnings,
        'errors': errors,
    }


class ClassSectionBulkUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        headers, rows = _parse_file(file_obj)
        if rows is None:
            return Response(
                {'error': 'Unsupported file format. Use .xlsx or .csv'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not rows:
            return Response({'error': 'File is empty'}, status=status.HTTP_400_BAD_REQUEST)

        result = _process_class_section_rows(rows, request)
        return Response(result, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────────────────────
# Fee Structure
# ─────────────────────────────────────────────────────────────

FEE_HEADERS = [
    'fee_category_name', 'fee_category_code', 'class_name',
    'amount', 'frequency', 'due_day', 'is_mandatory',
]

FEE_SAMPLE_ROWS = [
    ('Tuition Fee',    'TUITION',     'Class 1',  2500, 'MONTHLY',      10, 'Yes'),
    ('Tuition Fee',    'TUITION',     'Class 2',  2500, 'MONTHLY',      10, 'Yes'),
    ('Tuition Fee',    'TUITION',     'Class 3',  2500, 'MONTHLY',      10, 'Yes'),
    ('Tuition Fee',    'TUITION',     'Class 4',  2800, 'MONTHLY',      10, 'Yes'),
    ('Tuition Fee',    'TUITION',     'Class 5',  2800, 'MONTHLY',      10, 'Yes'),
    ('Tuition Fee',    'TUITION',     'Class 6',  3000, 'MONTHLY',      10, 'Yes'),
    ('Transport Fee',  'TRANSPORT',   'ALL',       800, 'MONTHLY',       5, 'No'),
    ('Annual Fund',    'ANNUAL_FUND', 'ALL',      5000, 'ANNUAL',        '', 'Yes'),
    ('Computer Lab',   'COMPUTER',    'Class 6',   500, 'QUARTERLY',    15, 'Yes'),
    ('Library Fee',    'LIBRARY',     'ALL',       300, 'ANNUAL',        '', 'No'),
]

FREQ_MAP = {
    'monthly': 'MONTHLY', 'month': 'MONTHLY', 'mo': 'MONTHLY', 'm': 'MONTHLY',
    'quarterly': 'QUARTERLY', 'quarter': 'QUARTERLY', 'q': 'QUARTERLY',
    'half-yearly': 'HALF_YEARLY', 'halfyearly': 'HALF_YEARLY', 'hy': 'HALF_YEARLY',
    'half yearly': 'HALF_YEARLY', 'semi-annual': 'HALF_YEARLY',
    'annual': 'ANNUAL', 'yearly': 'ANNUAL', 'year': 'ANNUAL', 'y': 'ANNUAL',
    'one-time': 'ONE_TIME', 'onetime': 'ONE_TIME', 'once': 'ONE_TIME', 'ot': 'ONE_TIME',
    'one time': 'ONE_TIME',
}


class FeeStructureTemplateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Fee_Structure'
        _styled_header_row(ws, FEE_HEADERS, 'E65100')
        for row in FEE_SAMPLE_ROWS:
            ws.append(row)

        # Instruction note in last row
        ws.append([])
        note_cell = ws.cell(row=ws.max_row + 1, column=1,
                            value='TIP: Use class_name=ALL to apply a fee to every class.')
        note_cell.font = Font(italic=True, color='888888')

        return _excel_response(wb, 'fee_structure_template.xlsx')


def _process_fee_structure_rows(rows, request_or_schema=None):
    """Core logic for fee structure import."""
    from apps.academics.models import AcademicYear, Class
    from apps.finance.models import FeeCategory, FeeStructure
    from django.db import transaction

    categories_created = 0
    fee_structures_created = 0
    fee_structures_updated = 0
    warnings = []
    errors = []

    ay = AcademicYear.objects.filter(is_current=True, is_deleted=False).first()
    if not ay:
        return {
            'categories_created': 0, 'fee_structures_created': 0, 'fee_structures_updated': 0,
            'warnings': [], 'errors': [{'error': 'No current academic year found. Create one first.'}],
        }

    for row_idx, row in enumerate(rows, start=2):
        cat_name = _val(row, 'fee_category_name', 'category_name', 'fee_name')
        cat_code = _val(row, 'fee_category_code', 'category_code', 'code')
        class_name = _val(row, 'class_name', 'class', 'grade')
        amount_raw = _val(row, 'amount', 'fee_amount')
        freq_raw = _val(row, 'frequency', 'freq', 'payment_frequency')
        is_mandatory_raw = _val(row, 'is_mandatory', 'mandatory')

        if not cat_name or not cat_code or not class_name or not amount_raw:
            errors.append({'row': row_idx, 'error': 'fee_category_name, fee_category_code, class_name and amount are required'})
            continue

        # Amount
        try:
            amount = Decimal(str(amount_raw).replace(',', ''))
        except InvalidOperation:
            errors.append({'row': row_idx, 'error': f'Invalid amount: {amount_raw}'})
            continue

        # Frequency
        freq = FREQ_MAP.get(freq_raw.lower(), None)
        if not freq:
            freq = freq_raw.upper() if freq_raw.upper() in ('MONTHLY', 'QUARTERLY', 'HALF_YEARLY', 'ANNUAL', 'ONE_TIME') else None
        if not freq:
            errors.append({'row': row_idx, 'error': f'Unknown frequency: "{freq_raw}". Use MONTHLY/QUARTERLY/HALF_YEARLY/ANNUAL/ONE_TIME'})
            continue

        # Due day
        due_day = None
        due_day_raw = _val(row, 'due_day', 'due_date', 'payment_day')
        if due_day_raw:
            try:
                due_day = int(due_day_raw)
                if not (1 <= due_day <= 31):
                    due_day = None
            except ValueError:
                pass

        is_mandatory = _bool_val(is_mandatory_raw, default=True)

        try:
            with transaction.atomic():
                cat, cat_created = FeeCategory.objects.get_or_create(
                    code=cat_code.upper(),
                    defaults={
                        'name': cat_name,
                        'is_mandatory': is_mandatory,
                        'is_active': True,
                    }
                )
                if cat_created:
                    categories_created += 1

                # Resolve classes
                if class_name.upper() == 'ALL':
                    classes = list(Class.objects.filter(is_deleted=False, is_active=True))
                    if not classes:
                        warnings.append({'row': row_idx, 'warning': 'class_name=ALL but no classes exist yet. Upload classes first.'})
                        continue
                else:
                    cls = Class.objects.filter(is_deleted=False, name__iexact=class_name).first()
                    if not cls:
                        cls = Class.objects.filter(
                            is_deleted=False,
                            display_name__iexact=class_name
                        ).first()
                    if not cls:
                        errors.append({'row': row_idx, 'error': f'Class "{class_name}" not found. Upload classes first.'})
                        continue
                    classes = [cls]

                for cls in classes:
                    _, created = FeeStructure.objects.update_or_create(
                        academic_year=ay,
                        class_obj=cls,
                        fee_category=cat,
                        defaults={
                            'amount': amount,
                            'frequency': freq,
                            'due_day': due_day,
                            'is_active': True,
                        }
                    )
                    if created:
                        fee_structures_created += 1
                    else:
                        fee_structures_updated += 1

        except Exception as e:
            errors.append({'row': row_idx, 'error': str(e)})

    return {
        'categories_created': categories_created,
        'fee_structures_created': fee_structures_created,
        'fee_structures_updated': fee_structures_updated,
        'warnings': warnings,
        'errors': errors,
    }


class FeeStructureBulkUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        headers, rows = _parse_file(file_obj)
        if rows is None:
            return Response(
                {'error': 'Unsupported file format. Use .xlsx or .csv'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not rows:
            return Response({'error': 'File is empty'}, status=status.HTTP_400_BAD_REQUEST)

        result = _process_fee_structure_rows(rows, request)
        return Response(result, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────────────────────
# Master Workbook Template
# ─────────────────────────────────────────────────────────────

class MasterTemplateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()

        # ── Sheet 1: README ──────────────────────────────────
        ws_readme = wb.active
        ws_readme.title = 'README'

        title_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
        title_font = Font(color='FFFFFF', bold=True, size=14)
        header_fill = PatternFill(start_color='283593', end_color='283593', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        # Title
        ws_readme.merge_cells('A1:G1')
        title_cell = ws_readme['A1']
        title_cell.value = 'CampusKona — School Onboarding Workbook'
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_readme.row_dimensions[1].height = 30

        ws_readme.merge_cells('A2:G2')
        sub_cell = ws_readme['A2']
        sub_cell.value = 'Fill sheets 2–5 in order. Upload this file at /onboarding. Questions? support@campuskona.com'
        sub_cell.font = Font(italic=True, color='444444')
        sub_cell.alignment = Alignment(horizontal='center')

        # Instruction table
        ws_readme.append([])
        headers = ['Sheet', 'Sheet Name', 'What It Creates', 'Required Columns', 'Notes']
        for col_idx, h in enumerate(headers, 1):
            cell = ws_readme.cell(row=4, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        instructions = [
            ('2', 'Classes_Sections', 'Class + Section records', 'class_name, class_order, section_name', 'One row per section. Repeat class_name for each section.'),
            ('3', 'Fee_Structure', 'Fee Categories + Fee Structures', 'fee_category_code, class_name, amount, frequency', 'Use class_name=ALL for school-wide fees.'),
            ('4', 'Students', 'Student + User accounts', 'first_name', 'Admission numbers auto-generated if blank.'),
            ('5', 'Staff', 'Staff + User accounts', 'first_name', 'Employee IDs auto-generated if blank.'),
        ]
        for row_data in instructions:
            ws_readme.append(list(row_data))

        # DPDP note
        ws_readme.append([])
        ws_readme.append([])
        dpdp_cell = ws_readme.cell(
            row=ws_readme.max_row, column=1,
            value='⚠ DPDP Notice: Columns containing Aadhaar, Religion, Caste data in the Students sheet are flagged as SENSITIVE. Parental consent will be requested via WhatsApp after upload.'
        )
        dpdp_cell.font = Font(color='B71C1C', bold=True)
        ws_readme.merge_cells(f'A{ws_readme.max_row}:G{ws_readme.max_row}')

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_readme.column_dimensions[col].width = 25

        # ── Sheet 2: Classes_Sections ───────────────────────
        ws_cs = wb.create_sheet('Classes_Sections')
        _styled_header_row(ws_cs, CS_HEADERS, '1565C0')
        for row in CS_SAMPLE_ROWS:
            ws_cs.append(row)

        # ── Sheet 3: Fee_Structure ──────────────────────────
        ws_fee = wb.create_sheet('Fee_Structure')
        _styled_header_row(ws_fee, FEE_HEADERS, 'E65100')
        for row in FEE_SAMPLE_ROWS:
            ws_fee.append(row)

        # ── Sheet 4: Students ───────────────────────────────
        ws_stu = wb.create_sheet('Students')
        stu_headers = [
            'first_name', 'last_name', 'middle_name', 'date_of_birth',
            'gender', 'blood_group', 'phone_number', 'email',
            'father_name', 'mother_name',
            'admission_number', 'current_address_line1', 'current_city',
            'current_state', 'current_pincode',
        ]
        _styled_header_row(ws_stu, stu_headers, '4472C4')
        ws_stu.append([
            'Arjun', 'Sharma', '', '2010-05-15', 'M', 'O+',
            '9876543210', '', 'Rajesh Sharma', 'Priya Sharma',
            'ADM001', '123 MG Road', 'Pune', 'Maharashtra', '411001',
        ])
        ws_stu.append([
            'Priya', 'Verma', '', '2011-08-22', 'F', 'A+',
            '9765432100', '', 'Mohan Verma', 'Sunita Verma',
            'ADM002', '45 Shivaji Nagar', 'Nagpur', 'Maharashtra', '440001',
        ])

        # ── Sheet 5: Staff ──────────────────────────────────
        ws_sta = wb.create_sheet('Staff')
        sta_headers = [
            'first_name', 'last_name', 'email', 'phone_number', 'gender',
            'employee_id', 'joining_date', 'designation', 'employment_type',
        ]
        _styled_header_row(ws_sta, sta_headers, '2E7D32')
        ws_sta.append(['Meena', 'Iyer', 'meena@school.com', '9811223344', 'F', 'EMP001', '2024-06-01', 'TEACHER', 'PERMANENT'])
        ws_sta.append(['Ramesh', 'Gupta', '', '9922334455', 'M', 'EMP002', '2024-06-01', 'TEACHER', 'PERMANENT'])

        return _excel_response(wb, 'CampusKona_Onboarding_Workbook.xlsx')


# ─────────────────────────────────────────────────────────────
# Master Upload (dispatch to each processor)
# ─────────────────────────────────────────────────────────────

class MasterUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file_obj.name.lower()
        if not filename.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Master upload requires an .xlsx file (the 5-sheet workbook)'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = load_workbook(file_obj, read_only=True, data_only=True)
        except Exception as e:
            return Response({'error': f'Could not open workbook: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        sheet_names = [s.lower() for s in wb.sheetnames]
        results = {}
        total_records = 0

        # Helper to find sheet by name pattern
        def get_sheet(patterns):
            for pat in patterns:
                for sname in wb.sheetnames:
                    if pat in sname.lower():
                        return wb[sname]
            return None

        # ── Process Classes + Sections ──
        ws = get_sheet(['classes_sections', 'class_section', 'classes'])
        if ws:
            _, rows = _ws_to_rows(ws)
            if rows:
                res = _process_class_section_rows(rows, request)
                results['classes_sections'] = res
                total_records += res.get('classes_created', 0) + res.get('sections_created', 0)

        # ── Process Fee Structure ──
        ws = get_sheet(['fee_structure', 'fee structure', 'fees'])
        if ws:
            _, rows = _ws_to_rows(ws)
            if rows:
                res = _process_fee_structure_rows(rows, request)
                results['fee_structure'] = res
                total_records += res.get('fee_structures_created', 0)

        # ── Process Students ──
        ws = get_sheet(['students', 'student'])
        if ws and ws.title.lower() not in ('readme', 'instructions'):
            _, rows = _ws_to_rows(ws)
            if rows:
                res = _process_student_rows(rows, request)
                results['students'] = res
                total_records += res.get('imported', 0)

        # ── Process Staff ──
        ws = get_sheet(['staff'])
        if ws and ws.title.lower() not in ('readme', 'instructions'):
            _, rows = _ws_to_rows(ws)
            if rows:
                res = _process_staff_rows(rows, request)
                results['staff'] = res
                total_records += res.get('imported', 0)

        if not results:
            return Response(
                {'error': 'No recognisable sheets found. Expected sheets named: Classes_Sections, Fee_Structure, Students, Staff'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check readiness
        cs = results.get('classes_sections', {})
        sta = results.get('staff', {})
        ready = (
            (cs.get('classes_created', 0) + cs.get('classes_existing', 0)) >= 1
            and (cs.get('sections_created', 0) + cs.get('sections_existing', 0)) >= 1
            and (sta.get('imported', 0)) >= 1
        )

        return Response({
            **results,
            'total_records': total_records,
            'ready_for_operations': ready,
        }, status=status.HTTP_200_OK)


def _process_student_rows(rows, request):
    """Inline student processor (mirrors StudentViewSet.bulk_upload logic)."""
    import uuid
    from django.db import transaction
    from django.contrib.auth import get_user_model
    from django.utils import timezone
    from apps.students.models import Student

    User = get_user_model()
    schema = request.tenant.schema_name if hasattr(request, 'tenant') else 'public'

    FIELD_MAP = {
        'name': 'first_name', 'student_name': 'first_name', 'first_name': 'first_name',
        'last_name': 'last_name', 'surname': 'last_name',
        'middle_name': 'middle_name',
        'dob': 'date_of_birth', 'date_of_birth': 'date_of_birth', 'birth_date': 'date_of_birth',
        'gender': 'gender',
        'phone': 'phone_number', 'mobile': 'phone_number', 'phone_number': 'phone_number',
        'father_name': 'father_name', 'father': 'father_name',
        'mother_name': 'mother_name', 'mother': 'mother_name',
        'blood_group': 'blood_group',
        'admission_no': 'admission_number', 'admission_number': 'admission_number',
        'email': 'email',
        'address': 'current_address_line1', 'current_address_line1': 'current_address_line1',
        'city': 'current_city', 'current_city': 'current_city',
        'state': 'current_state', 'current_state': 'current_state',
        'pincode': 'current_pincode', 'current_pincode': 'current_pincode',
    }
    DPDP_SENSITIVE = {'aadhar_number', 'aadhaar_number', 'religion', 'caste', 'category'}
    headers = list(rows[0].keys()) if rows else []
    dpdp_flags = [
        {'column': col, 'message': f'Column "{col}" is SENSITIVE under DPDP Act 2023.'}
        for col in headers if col in DPDP_SENSITIVE
    ]

    imported, errors, warnings = 0, [], []

    for row_idx, row in enumerate(rows, start=2):
        mapped = {}
        for col, val in row.items():
            field = FIELD_MAP.get(col)
            if field and val is not None and str(val).strip():
                mapped[field] = str(val).strip()

        first_name = mapped.get('first_name', '')
        if not first_name:
            errors.append({'row': row_idx, 'error': 'Name is required'})
            continue

        admission_number = mapped.get('admission_number', '')
        if not admission_number:
            admission_number = f'AUTO-{uuid.uuid4().hex[:8].upper()}'
            warnings.append({'row': row_idx, 'warning': f'No admission number, auto-assigned: {admission_number}'})

        try:
            with transaction.atomic():
                pseudo_email = (
                    mapped.get('email') or
                    f'student.{admission_number.lower().replace(" ", "")}@{schema}.campuskona.internal'
                )
                if User.objects.filter(email=pseudo_email).exists():
                    warnings.append({'row': row_idx, 'warning': f'Student {admission_number} already exists, skipped'})
                    continue

                user = User.objects.create(email=pseudo_email, username=pseudo_email,
                                           first_name=first_name, last_name=mapped.get('last_name', ''), is_active=True)
                user.set_unusable_password()
                user.save(update_fields=['password'])

                gender_raw = mapped.get('gender', 'M').upper()
                gender = 'M' if gender_raw in ('M', 'MALE', 'BOY') else ('F' if gender_raw in ('F', 'FEMALE', 'GIRL') else 'O')

                student_data = {
                    'user': user, 'admission_number': admission_number,
                    'first_name': first_name, 'last_name': mapped.get('last_name', ''),
                    'middle_name': mapped.get('middle_name', ''),
                    'father_name': mapped.get('father_name', ''),
                    'mother_name': mapped.get('mother_name', ''),
                    'blood_group': mapped.get('blood_group', ''),
                    'current_address_line1': mapped.get('current_address_line1', ''),
                    'current_city': mapped.get('current_city', ''),
                    'current_state': mapped.get('current_state', ''),
                    'current_pincode': mapped.get('current_pincode', ''),
                    'email': mapped.get('email', '') or '',
                    'gender': gender,
                    'admission_date': timezone.now().date(),
                    'admission_status': 'ACTIVE',
                }

                if mapped.get('phone_number'):
                    ph = ''.join(filter(str.isdigit, mapped['phone_number']))[-10:]
                    if len(ph) == 10:
                        student_data['phone_number'] = ph

                if mapped.get('date_of_birth'):
                    from datetime import datetime
                    for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            student_data['date_of_birth'] = datetime.strptime(mapped['date_of_birth'], fmt).date()
                            break
                        except ValueError:
                            continue

                Student.objects.create(**student_data)
                imported += 1
        except Exception as e:
            errors.append({'row': row_idx, 'error': str(e)})

    return {
        'imported': imported, 'total_rows': len(rows),
        'warnings': warnings, 'errors': errors,
        'dpdp_flags': dpdp_flags, 'dpdp_flag_count': len(dpdp_flags),
        'message': f'Successfully imported {imported} of {len(rows)} students',
    }


def _process_staff_rows(rows, request):
    """Inline staff processor (mirrors StaffMemberViewSet.bulk_upload logic)."""
    import uuid
    from django.db import transaction
    from django.contrib.auth import get_user_model
    from apps.staff.models import StaffMember

    User = get_user_model()
    schema = request.tenant.schema_name if hasattr(request, 'tenant') else 'public'

    FIELD_MAP = {
        'name': 'first_name', 'staff_name': 'first_name', 'first_name': 'first_name',
        'last_name': 'last_name', 'surname': 'last_name',
        'email': 'email', 'phone': 'phone_number', 'phone_number': 'phone_number',
        'employee_id': 'employee_id', 'staff_id': 'employee_id',
        'joining_date': 'joining_date', 'date_of_joining': 'joining_date',
        'designation': 'designation', 'employment_type': 'employment_type',
        'gender': 'gender',
    }

    imported, errors, warnings = 0, [], []

    for row_idx, row in enumerate(rows, start=2):
        mapped = {}
        for col, val in row.items():
            field = FIELD_MAP.get(col)
            if field and val is not None and str(val).strip():
                mapped[field] = str(val).strip()

        first_name = mapped.get('first_name', '')
        if not first_name:
            errors.append({'row': row_idx, 'error': 'Name (first_name) is required'})
            continue

        employee_id = mapped.get('employee_id', '')
        if not employee_id:
            employee_id = f'EMP-{uuid.uuid4().hex[:6].upper()}'
            warnings.append({'row': row_idx, 'warning': f'No employee_id, auto-assigned: {employee_id}'})

        try:
            with transaction.atomic():
                pseudo_email = (
                    mapped.get('email') or
                    f'staff.{employee_id.lower().replace(" ", "")}@{schema}.campuskona.internal'
                )
                if User.objects.filter(email=pseudo_email).exists():
                    warnings.append({'row': row_idx, 'warning': f'Staff {employee_id} already exists, skipped'})
                    continue

                user = User.objects.create(email=pseudo_email, username=pseudo_email,
                                           first_name=first_name, last_name=mapped.get('last_name', ''), is_active=True)
                user.set_unusable_password()
                user.save(update_fields=['password'])

                gender_raw = mapped.get('gender', 'M').upper()
                gender = 'M' if gender_raw in ('M', 'MALE') else ('F' if gender_raw in ('F', 'FEMALE') else 'O')

                joining_date = None
                if mapped.get('joining_date'):
                    from datetime import datetime, date
                    for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y']:
                        try:
                            joining_date = datetime.strptime(mapped['joining_date'], fmt).date()
                            break
                        except ValueError:
                            continue
                if not joining_date:
                    from datetime import date
                    joining_date = date.today()

                StaffMember.objects.create(
                    user=user, employee_id=employee_id,
                    first_name=first_name, last_name=mapped.get('last_name', ''),
                    gender=gender, joining_date=joining_date,
                    designation=mapped.get('designation', ''),
                    employment_type=mapped.get('employment_type', 'PERMANENT').upper(),
                    phone_number=mapped.get('phone_number', ''),
                )
                imported += 1
        except Exception as e:
            errors.append({'row': row_idx, 'error': str(e)})

    return {
        'imported': imported, 'total_rows': len(rows),
        'warnings': warnings, 'errors': errors,
        'message': f'Successfully imported {imported} of {len(rows)} staff members',
    }
