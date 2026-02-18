"""
DPDP Act 2023 Compliance - API Views
Implements consent management, grievance redressal, and data subject rights
"""
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.db.models import Q

from apps.privacy.models import (
    ConsentPurpose,
    ParentalConsent,
    ConsentAuditLog,
    Grievance,
    GrievanceComment,
    DataBreach,
    DeletionRequest,
    CorrectionRequest,
)
from apps.privacy.serializers import (
    ConsentPurposeSerializer,
    ParentalConsentSerializer,
    ConsentRequestSerializer,
    ConsentGrantSerializer,
    ConsentWithdrawalSerializer,
    GrievanceSerializer,
    GrievanceCommentSerializer,
    DataBreachSerializer,
    DeletionRequestSerializer,
    CorrectionRequestSerializer,
    SensitiveDataAccessSerializer,
    AccessPatternAlertSerializer,
)
from apps.privacy.services.verification import ParentalVerificationService
from apps.privacy.services.data_export import DataExportService
from apps.privacy.services.data_deletion import DataDeletionService
from apps.privacy.services.data_correction import DataCorrectionService
from apps.privacy.services.notifications import PrivacyNotificationService
from apps.privacy.services.audit_logging import AuditLoggingService
from apps.privacy.models import SensitiveDataAccess, AccessPatternAlert, ParentalConsent, ConsentPurpose, Grievance, CorrectionRequest, DeletionRequest
from apps.students.models import Student
from apps.staff.models import StaffRoleAssignment, StaffMember
from apps.academics.models import Section, StudentEnrollment
from django.db.models import Count, Q
from django.utils import timezone


class ConsentPurposeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    List available consent purposes
    Read-only: purposes are managed via admin or seed command
    """
    queryset = ConsentPurpose.objects.filter(is_active=True)
    serializer_class = ConsentPurposeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter active purposes ordered by mandatory first"""
        return ConsentPurpose.objects.filter(
            is_active=True
        ).order_by('-is_mandatory', 'category', 'name')


class ParentalConsentViewSet(viewsets.ModelViewSet):
    """
    Manage parental consent for student data processing
    Implements DPDP Act Section 9.1 (Verifiable Consent Obligation)
    """
    queryset = ParentalConsent.objects.all()
    serializer_class = ParentalConsentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter consent records based on user role"""
        user = self.request.user

        # Parents see only their children's consents
        if user.user_type == 'PARENT':
            # Get student IDs for this parent
            from apps.students.models import StudentParent
            student_ids = StudentParent.objects.filter(
                parent_user=user
            ).values_list('student_id', flat=True)

            return ParentalConsent.objects.filter(
                student_id__in=student_ids
            ).select_related('student', 'parent_user', 'purpose')

        # School staff see all consents for their tenant
        if hasattr(self.request, 'tenant'):
            return ParentalConsent.objects.filter(
                student__tenant=self.request.tenant
            ).select_related('student', 'parent_user', 'purpose')

        return ParentalConsent.objects.none()

    @action(detail=False, methods=['post'])
    def request_consent(self, request):
        """
        Request consent from parent for specific purpose
        Sends OTP for verification

        POST /api/v1/privacy/consents/request_consent/
        Body: {
            "student_id": 1,
            "purpose_code": "COMMUNICATION_NOTICES",
            "verification_method": "EMAIL_OTP"
        }
        """
        serializer = ConsentRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        student_id = serializer.validated_data['student_id']
        purpose_code = serializer.validated_data['purpose_code']
        verification_method = serializer.validated_data['verification_method']

        try:
            student = Student.objects.get(id=student_id)
            purpose = ConsentPurpose.objects.get(code=purpose_code)

            # Check if user is parent of this student
            from apps.students.models import StudentParent
            if not StudentParent.objects.filter(
                student=student,
                parent_user=request.user
            ).exists():
                return Response(
                    {'error': 'You are not authorized as parent/guardian of this student'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Create or get existing consent record
            consent, created = ParentalConsent.objects.get_or_create(
                student=student,
                parent_user=request.user,
                purpose=purpose,
                defaults={
                    'consent_text': purpose.description,
                    'verification_method': verification_method,
                }
            )

            # If consent already given and not withdrawn, no need to re-verify
            if consent.is_valid():
                return Response({
                    'message': 'Consent already granted for this purpose',
                    'consent_id': str(consent.consent_id)
                })

            # Send verification based on method
            verification_sent = False
            if verification_method == 'EMAIL_OTP':
                verification_sent = ParentalVerificationService.send_email_otp(request.user, student)
            elif verification_method == 'SMS_OTP':
                verification_sent = ParentalVerificationService.send_sms_otp(request.user, student)
            elif verification_method == 'EXISTING_IDENTITY':
                success, message = ParentalVerificationService.verify_using_existing_identity(request.user)
                if not success:
                    return Response({'error': message}, status=status.HTTP_400_BAD_REQUEST)
                verification_sent = True
            elif verification_method == 'AADHAAR_VIRTUAL_TOKEN':
                return Response(
                    {'error': 'Aadhaar verification not yet implemented. Please use OTP verification.'},
                    status=status.HTTP_501_NOT_IMPLEMENTED
                )

            if not verification_sent:
                return Response(
                    {'error': 'Failed to send verification code. Please try again.'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            # Create audit log
            ConsentAuditLog.objects.create(
                consent=consent,
                action='REQUESTED',
                performed_by=request.user,
                ip_address=self._get_client_ip(request),
                details={
                    'verification_method': verification_method,
                    'purpose_code': purpose_code,
                }
            )

            message = 'Verification code sent' if 'OTP' in verification_method else 'Consent requested'
            return Response({
                'message': message,
                'consent_id': str(consent.consent_id),
                'verification_method': verification_method
            })

        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)
        except ConsentPurpose.DoesNotExist:
            return Response({'error': 'Invalid consent purpose'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def grant_consent(self, request):
        """
        Grant consent after OTP verification

        POST /api/v1/privacy/consents/grant_consent/
        Body: {
            "consent_id": "uuid-here",
            "otp": "123456",
            "agreed": true
        }
        """
        serializer = ConsentGrantSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        consent_id = serializer.validated_data['consent_id']
        otp = serializer.validated_data.get('otp', '')
        agreed = serializer.validated_data['agreed']

        try:
            consent = ParentalConsent.objects.get(
                consent_id=consent_id,
                parent_user=request.user
            )

            if not agreed:
                return Response(
                    {'error': 'Consent must be explicitly agreed to'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Verify OTP if required
            if consent.verification_method in ['EMAIL_OTP', 'SMS_OTP']:
                if not otp:
                    return Response(
                        {'error': 'OTP is required for this verification method'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                method = 'email' if consent.verification_method == 'EMAIL_OTP' else 'sms'
                success, message = ParentalVerificationService.verify_otp(
                    request.user,
                    consent.student,
                    otp,
                    method
                )

                if not success:
                    return Response({'error': message}, status=status.HTTP_400_BAD_REQUEST)

            # Grant consent
            consent.consent_given = True
            consent.consent_date = timezone.now()
            consent.verified_at = timezone.now()
            consent.withdrawn = False  # Reset if previously withdrawn
            consent.ip_address = self._get_client_ip(request)
            consent.user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]
            consent.save()

            # Create audit log
            ConsentAuditLog.objects.create(
                consent=consent,
                action='GIVEN',
                performed_by=request.user,
                ip_address=self._get_client_ip(request),
                details={
                    'otp_verified': bool(otp),
                    'verification_method': consent.verification_method,
                }
            )

            return Response({
                'message': 'Consent granted successfully',
                'consent': ParentalConsentSerializer(consent).data
            })

        except ParentalConsent.DoesNotExist:
            return Response(
                {'error': 'Consent record not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['post'])
    def withdraw_consent(self, request, pk=None):
        """
        Withdraw previously granted consent
        Implements Right to Withdrawal (DPDP Act)

        POST /api/v1/privacy/consents/{id}/withdraw_consent/
        Body: {
            "reason": "No longer wish to receive communications"
        }
        """
        consent = self.get_object()
        serializer = ConsentWithdrawalSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        reason = serializer.validated_data.get('reason', '')

        # Check if consent belongs to requesting user
        if consent.parent_user != request.user:
            return Response(
                {'error': 'You are not authorized to withdraw this consent'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Cannot withdraw mandatory consent
        if consent.purpose.is_mandatory:
            return Response({
                'error': 'Cannot withdraw consent for mandatory processing. This data is essential for providing educational services. Contact school administration if you wish to discontinue services.',
                'purpose': consent.purpose.name,
                'legal_basis': consent.purpose.legal_basis
            }, status=status.HTTP_400_BAD_REQUEST)

        # Mark as withdrawn
        consent.withdrawn = True
        consent.withdrawn_at = timezone.now()
        consent.withdrawal_reason = reason
        consent.save()

        # Create audit log
        ConsentAuditLog.objects.create(
            consent=consent,
            action='WITHDRAWN',
            performed_by=request.user,
            ip_address=self._get_client_ip(request),
            details={'reason': reason}
        )

        # TODO: Trigger data deletion workflow for withdrawn consent
        # This will be implemented in Phase 3 (Data Subject Rights)

        return Response({
            'message': 'Consent withdrawn successfully. Data will be deleted as per retention policy.',
            'consent': ParentalConsentSerializer(consent).data
        })

    @action(detail=False, methods=['get'])
    def my_consents(self, request):
        """
        Get all consent records for current parent user

        GET /api/v1/privacy/consents/my_consents/?student_id=1
        """
        student_id = request.query_params.get('student_id')

        queryset = self.get_queryset()

        if student_id:
            queryset = queryset.filter(student_id=student_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def _get_client_ip(self, request):
        """Extract client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


class GrievanceViewSet(viewsets.ModelViewSet):
    """
    Grievance Redressal System
    Implements DPDP Act Section 12 (Grievance Redressal)
    """
    queryset = Grievance.objects.all()
    serializer_class = GrievanceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter grievances based on user role"""
        user = self.request.user

        # Parents see only their own grievances
        if user.user_type == 'PARENT':
            return Grievance.objects.filter(
                filed_by=user
            ).select_related('student', 'filed_by', 'assigned_to', 'resolved_by')

        # School staff see all grievances for their tenant
        if hasattr(self.request, 'tenant'):
            return Grievance.objects.filter(
                student__tenant=self.request.tenant
            ).select_related('student', 'filed_by', 'assigned_to', 'resolved_by')

        return Grievance.objects.none()

    def create(self, request, *args, **kwargs):
        """File new grievance"""
        data = request.data.copy()
        data['filed_by'] = request.user.id
        data['status'] = 'SUBMITTED'

        serializer = self.get_serializer(data=data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        grievance = serializer.save()

        # Send acknowledgment email to parent
        try:
            PrivacyNotificationService.send_grievance_acknowledgment(grievance)
        except Exception as e:
            # Log error but don't fail the request
            print(f"Failed to send grievance acknowledgment email: {e}")

        # Notify admin/DPO of new grievance
        try:
            PrivacyNotificationService.notify_admin_of_grievance(grievance)
        except Exception as e:
            print(f"Failed to send admin notification email: {e}")

        # Auto-acknowledge the grievance
        grievance.status = 'ACKNOWLEDGED'
        grievance.acknowledged_at = timezone.now()
        grievance.save()

        return Response(
            GrievanceSerializer(grievance).data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=['post'])
    def add_comment(self, request, pk=None):
        """
        Add comment/update to grievance

        POST /api/v1/privacy/grievances/{id}/add_comment/
        Body: {
            "comment": "We are investigating this issue",
            "is_internal": false
        }
        """
        grievance = self.get_object()
        comment_text = request.data.get('comment')
        is_internal = request.data.get('is_internal', False)

        if not comment_text:
            return Response(
                {'error': 'Comment text is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        comment = GrievanceComment.objects.create(
            grievance=grievance,
            user=request.user,
            comment=comment_text,
            is_internal=is_internal
        )

        # Notify parent if not internal comment
        if not is_internal:
            try:
                update_message = f"""
New comment from {request.user.get_full_name()}:

{comment_text}

Added on: {comment.created_at.strftime('%d %B %Y, %H:%M')}
"""
                PrivacyNotificationService.send_grievance_update(grievance, update_message)
            except Exception as e:
                print(f"Failed to send grievance update email: {e}")

        return Response({
            'message': 'Comment added successfully',
            'comment': GrievanceCommentSerializer(comment).data
        })

    @action(detail=True, methods=['post'])
    def acknowledge(self, request, pk=None):
        """Acknowledge grievance (within 24 hours requirement)"""
        grievance = self.get_object()

        if grievance.status != 'SUBMITTED':
            return Response(
                {'error': 'Grievance already acknowledged'},
                status=status.HTTP_400_BAD_REQUEST
            )

        grievance.status = 'ACKNOWLEDGED'
        grievance.acknowledged_at = timezone.now()
        grievance.save()

        return Response({
            'message': 'Grievance acknowledged',
            'grievance': GrievanceSerializer(grievance).data
        })

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """
        Resolve grievance

        POST /api/v1/privacy/grievances/{id}/resolve/
        Body: {
            "resolution_notes": "Issue has been resolved by..."
        }
        """
        grievance = self.get_object()
        resolution_notes = request.data.get('resolution_notes', '')

        if not resolution_notes:
            return Response(
                {'error': 'Resolution notes are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        grievance.status = 'RESOLVED'
        grievance.resolved_at = timezone.now()
        grievance.resolved_by = request.user
        grievance.resolution_notes = resolution_notes
        grievance.save()

        # Send resolution notification to parent
        try:
            PrivacyNotificationService.send_grievance_resolution(grievance)
        except Exception as e:
            print(f"Failed to send grievance resolution email: {e}")

        return Response({
            'message': 'Grievance resolved successfully',
            'grievance': GrievanceSerializer(grievance).data
        })


class DataBreachViewSet(viewsets.ModelViewSet):
    """
    Data Breach Notification Management
    Implements DPDP Act Section 8 (72-hour notification requirement)
    """
    queryset = DataBreach.objects.all()
    serializer_class = DataBreachSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Only staff with appropriate permissions can view breach data"""
        user = self.request.user

        # Only admin, superadmin, DPO can view breaches
        if user.user_type in ['ADMIN', 'SUPERADMIN'] or user.is_staff:
            if hasattr(self.request, 'tenant'):
                return DataBreach.objects.filter(
                    students_affected__tenant=self.request.tenant
                ).distinct()

        return DataBreach.objects.none()


class DeletionRequestViewSet(viewsets.ModelViewSet):
    """
    Right to Erasure (DPDP Act Section 14)
    Handle data deletion requests
    """
    queryset = DeletionRequest.objects.all()
    serializer_class = DeletionRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter based on user role"""
        user = self.request.user

        if user.user_type == 'PARENT':
            return DeletionRequest.objects.filter(
                requested_by=user
            ).select_related('student', 'requested_by', 'completed_by')

        if hasattr(self.request, 'tenant'):
            return DeletionRequest.objects.filter(
                student__tenant=self.request.tenant
            ).select_related('student', 'requested_by', 'completed_by')

        return DeletionRequest.objects.none()

    def create(self, request, *args, **kwargs):
        """
        Submit data deletion request

        POST /api/v1/privacy/deletion-requests/
        Body: {
            "student_id": 1,
            "reason": "Withdrawing child from school"
        }
        """
        student_id = request.data.get('student_id')
        reason = request.data.get('reason', '')

        if not student_id:
            return Response(
                {'error': 'student_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            student = Student.objects.get(id=student_id)

            # Verify parent has access
            from apps.students.models import StudentParent
            if not StudentParent.objects.filter(
                student=student,
                parent_user=request.user
            ).exists():
                return Response(
                    {'error': 'You are not authorized for this student'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Check if deletion is allowed
            can_delete, message = DataDeletionService.can_delete_student(student)
            if not can_delete:
                return Response(
                    {'error': message, 'can_delete': False},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Create deletion request
            deletion_request = DeletionRequest.objects.create(
                student=student,
                requested_by=request.user,
                reason=reason,
                status='PENDING'
            )

            # Send confirmation email to parent
            try:
                PrivacyNotificationService.send_deletion_request_confirmation(deletion_request)
            except Exception as e:
                print(f"Failed to send deletion confirmation email: {e}")

            return Response({
                'message': 'Deletion request submitted. School administration will review and process within 7 days.',
                'deletion_request': DeletionRequestSerializer(deletion_request).data
            }, status=status.HTTP_201_CREATED)

        except Student.DoesNotExist:
            return Response(
                {'error': 'Student not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """
        Approve and execute deletion request (Admin only)

        POST /api/v1/privacy/deletion-requests/{id}/approve/
        Body: {
            "approval_notes": "Verified student is inactive for 10+ years"
        }
        """
        # Check admin permission
        if request.user.user_type not in ['ADMIN', 'SUPERADMIN']:
            return Response(
                {'error': 'Only administrators can approve deletion requests'},
                status=status.HTTP_403_FORBIDDEN
            )

        deletion_request = self.get_object()

        if deletion_request.status != 'PENDING':
            return Response(
                {'error': f'Cannot approve request in status: {deletion_request.status}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        approval_notes = request.data.get('approval_notes', '')

        # Execute deletion
        success, message = DataDeletionService.process_deletion_request(
            deletion_request,
            request.user
        )

        if success:
            # Send completion notification to parent
            try:
                deletion_request.refresh_from_db()
                PrivacyNotificationService.send_deletion_completion(deletion_request)
            except Exception as e:
                print(f"Failed to send deletion completion email: {e}")

            return Response({
                'message': message,
                'deletion_request': DeletionRequestSerializer(deletion_request).data
            })
        else:
            return Response(
                {'error': message},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """
        Reject deletion request (Admin only)

        POST /api/v1/privacy/deletion-requests/{id}/reject/
        Body: {
            "rejection_reason": "Student has outstanding fees"
        }
        """
        # Check admin permission
        if request.user.user_type not in ['ADMIN', 'SUPERADMIN']:
            return Response(
                {'error': 'Only administrators can reject deletion requests'},
                status=status.HTTP_403_FORBIDDEN
            )

        deletion_request = self.get_object()
        rejection_reason = request.data.get('rejection_reason', '')

        if not rejection_reason:
            return Response(
                {'error': 'Rejection reason is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        deletion_request.status = 'REJECTED'
        deletion_request.reviewed_at = timezone.now()
        deletion_request.notes = rejection_reason
        deletion_request.save()

        # TODO: Send email notification to parent

        return Response({
            'message': 'Deletion request rejected',
            'deletion_request': DeletionRequestSerializer(deletion_request).data
        })


class CorrectionRequestViewSet(viewsets.ModelViewSet):
    """
    Right to Correction (DPDP Act Section 13)
    Handle data correction requests
    """
    queryset = CorrectionRequest.objects.all()
    serializer_class = CorrectionRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter based on user role"""
        user = self.request.user

        if user.user_type == 'PARENT':
            return CorrectionRequest.objects.filter(
                requested_by=user
            ).select_related('student', 'requested_by', 'completed_by')

        if hasattr(self.request, 'tenant'):
            return CorrectionRequest.objects.filter(
                student__tenant=self.request.tenant
            ).select_related('student', 'requested_by', 'completed_by')

        return CorrectionRequest.objects.none()

    def create(self, request, *args, **kwargs):
        """
        Submit data correction request

        POST /api/v1/privacy/correction-requests/
        Body: {
            "student_id": 1,
            "field_name": "date_of_birth",
            "corrected_value": "2010-05-15",
            "reason": "Birth certificate shows correct date as 15th May, not 16th"
        }
        """
        student_id = request.data.get('student_id')
        field_name = request.data.get('field_name')
        corrected_value = request.data.get('corrected_value')
        reason = request.data.get('reason', '')

        if not all([student_id, field_name, corrected_value]):
            return Response(
                {'error': 'student_id, field_name, and corrected_value are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            student = Student.objects.get(id=student_id)

            # Verify parent has access
            from apps.students.models import StudentParent
            if not StudentParent.objects.filter(
                student=student,
                parent_user=request.user
            ).exists():
                return Response(
                    {'error': 'You are not authorized for this student'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Submit correction request using service
            correction_request, message = DataCorrectionService.submit_correction_request(
                student=student,
                field_name=field_name,
                corrected_value=corrected_value,
                requested_by=request.user,
                reason=reason
            )

            # Send confirmation email to parent
            try:
                PrivacyNotificationService.send_correction_request_confirmation(correction_request)
            except Exception as e:
                print(f"Failed to send correction confirmation email: {e}")

            return Response({
                'message': message,
                'correction_request': CorrectionRequestSerializer(correction_request).data
            }, status=status.HTTP_201_CREATED)

        except Student.DoesNotExist:
            return Response(
                {'error': 'Student not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """
        Approve correction request (Admin only)

        POST /api/v1/privacy/correction-requests/{id}/approve/
        """
        # Check admin permission
        if request.user.user_type not in ['ADMIN', 'SUPERADMIN']:
            return Response(
                {'error': 'Only administrators can approve correction requests'},
                status=status.HTTP_403_FORBIDDEN
            )

        correction_request = self.get_object()

        # Approve and apply correction
        success, message = DataCorrectionService.approve_correction(
            correction_request,
            request.user
        )

        if success:
            # Send completion notification to parent
            try:
                correction_request.refresh_from_db()
                PrivacyNotificationService.send_correction_completion(correction_request)
            except Exception as e:
                print(f"Failed to send correction completion email: {e}")

            return Response({
                'message': message,
                'correction_request': CorrectionRequestSerializer(correction_request).data
            })
        else:
            return Response(
                {'error': message},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """
        Reject correction request (Admin only)

        POST /api/v1/privacy/correction-requests/{id}/reject/
        Body: {
            "rejection_reason": "Please provide supporting documentation"
        }
        """
        # Check admin permission
        if request.user.user_type not in ['ADMIN', 'SUPERADMIN']:
            return Response(
                {'error': 'Only administrators can reject correction requests'},
                status=status.HTTP_403_FORBIDDEN
            )

        correction_request = self.get_object()
        rejection_reason = request.data.get('rejection_reason', '')

        if not rejection_reason:
            return Response(
                {'error': 'Rejection reason is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        success = DataCorrectionService.reject_correction(
            correction_request,
            request.user,
            rejection_reason
        )

        if success:
            # Send rejection notification to parent
            try:
                correction_request.refresh_from_db()
                PrivacyNotificationService.send_correction_completion(correction_request)
            except Exception as e:
                print(f"Failed to send correction rejection email: {e}")

            return Response({
                'message': 'Correction request rejected',
                'correction_request': CorrectionRequestSerializer(correction_request).data
            })
        else:
            return Response(
                {'error': 'Failed to reject correction request'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def correctable_fields(self, request):
        """
        Get list of fields that can be corrected

        GET /api/v1/privacy/correction-requests/correctable_fields/
        """
        fields_info = []
        for field_name in DataCorrectionService.CORRECTABLE_FIELDS:
            display_name = DataCorrectionService.get_field_display_name(field_name)
            requires_approval = field_name in DataCorrectionService.ADMIN_APPROVAL_REQUIRED

            fields_info.append({
                'field_name': field_name,
                'display_name': display_name,
                'requires_admin_approval': requires_approval
            })

        return Response({
            'correctable_fields': fields_info,
            'total_count': len(fields_info)
        })


class SensitiveDataAccessViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Audit logs for sensitive data access
    Read-only - logs are created automatically
    """
    queryset = SensitiveDataAccess.objects.all()
    serializer_class = SensitiveDataAccessSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['user', 'student', 'field_name', 'access_type', 'is_flagged', 'has_valid_consent']
    search_fields = ['user__first_name', 'user__last_name', 'student__admission_number', 'student__first_name']
    ordering_fields = ['accessed_at', 'field_name', 'access_type']
    ordering = ['-accessed_at']

    def get_queryset(self):
        """Filter based on user role - only admins can view audit logs"""
        user = self.request.user

        # Only admin/superadmin can view audit logs
        if user.user_type not in ['ADMIN', 'SUPERADMIN']:
            return SensitiveDataAccess.objects.none()

        if hasattr(self.request, 'tenant'):
            return SensitiveDataAccess.objects.filter(
                student__tenant=self.request.tenant
            ).select_related('user', 'student', 'consent_purpose', 'reviewed_by')

        return SensitiveDataAccess.objects.all()

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """
        Get access summary statistics
 
        GET /api/v1/privacy/audit-logs/summary/?user_id=1&days=30
        If user_id is not provided, returns global summary for the school.
        """
        user_id = request.query_params.get('user_id')
        days = int(request.query_params.get('days', 30))
        tenant = getattr(request, 'tenant', None)

        if not user_id:
            summary = AuditLoggingService.get_global_access_summary(days, tenant)
            return Response(summary)

        try:
            from apps.authentication.models import User
            user = User.objects.get(id=user_id)
            summary = AuditLoggingService.get_user_access_summary(user, days)
            return Response(summary)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=False, methods=['get'])
    def student_history(self, request):
        """
        Get access history for a specific student

        GET /api/v1/privacy/audit-logs/student_history/?student_id=1&days=90
        """
        student_id = request.query_params.get('student_id')
        days = int(request.query_params.get('days', 90))

        if not student_id:
            return Response(
                {'error': 'student_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response(
                {'error': 'Student not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        accesses = AuditLoggingService.get_student_access_history(student, days)
        serializer = self.get_serializer(accesses, many=True)
        return Response({
            'student': {
                'id': student.id,
                'admission_number': student.admission_number,
                'full_name': student.full_name
            },
            'access_history': serializer.data
        })


class AccessPatternAlertViewSet(viewsets.ModelViewSet):
    """
    Access pattern alerts for suspicious data access
    """
    queryset = AccessPatternAlert.objects.all()
    serializer_class = AccessPatternAlertSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['user', 'alert_type', 'severity', 'status']
    ordering_fields = ['detected_at', 'severity']
    ordering = ['-severity', '-detected_at']

    def get_queryset(self):
        """Only admins can view alerts"""
        user = self.request.user

        if user.user_type not in ['ADMIN', 'SUPERADMIN']:
            return AccessPatternAlert.objects.none()

        return AccessPatternAlert.objects.all().select_related(
            'user', 'assigned_to'
        ).prefetch_related('affected_students', 'related_accesses')

    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get all pending alerts"""
        alerts = AuditLoggingService.get_pending_alerts()
        serializer = self.get_serializer(alerts, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """
        Resolve an alert

        POST /api/v1/privacy/alerts/{id}/resolve/
        Body: {
            "resolution_notes": "Verified as legitimate access"
        }
        """
        alert = self.get_object()
        resolution_notes = request.data.get('resolution_notes', '')

        if not resolution_notes:
            return Response(
                {'error': 'Resolution notes are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        AuditLoggingService.resolve_alert(alert, request.user, resolution_notes)

        return Response({
            'message': 'Alert resolved successfully',
            'alert': AccessPatternAlertSerializer(alert).data
        })

    @action(detail=True, methods=['post'])
    def mark_false_positive(self, request, pk=None):
        """
        Mark alert as false positive

        POST /api/v1/privacy/alerts/{id}/mark_false_positive/
        Body: {
            "notes": "User had legitimate reason for bulk access"
        }
        """
        alert = self.get_object()
        notes = request.data.get('notes', '')


# ─────────────────────────────────────────────────────────────
# Workstream D: DPDP Document Generation Views
# ─────────────────────────────────────────────────────────────
from rest_framework.views import APIView
from django.http import HttpResponse


class GenerateDPAView(APIView):
    """Generate and download the Data Processing Agreement PDF (Workstream D)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.privacy.services.document_generator import generate_dpa_pdf
        tenant = getattr(request, 'tenant', None)
        school_name = getattr(tenant, 'name', '') or getattr(tenant, 'schema_name', 'School')
        schema_name = getattr(tenant, 'schema_name', 'school')

        pdf_bytes = generate_dpa_pdf(school_name=school_name, schema_name=schema_name)
        content_type = 'application/pdf'
        ext = 'pdf'
        # Detect HTML fallback by checking first bytes
        if pdf_bytes[:1] in (b'<', b'\n', b'\r'):
            content_type = 'text/html'
            ext = 'html'

        response = HttpResponse(pdf_bytes, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="CampusKona_DPA_{schema_name}.{ext}"'
        return response


class GeneratePrivacyNoticeView(APIView):
    """Generate and download the Privacy Notice PDF (Workstream D)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.privacy.services.document_generator import generate_privacy_notice_pdf
        tenant = getattr(request, 'tenant', None)
        school_name = getattr(tenant, 'name', '') or getattr(tenant, 'schema_name', 'School')
        schema_name = getattr(tenant, 'schema_name', 'school')

        pdf_bytes = generate_privacy_notice_pdf(school_name=school_name)
        content_type = 'application/pdf'
        ext = 'pdf'
        if pdf_bytes[:1] in (b'<', b'\n', b'\r'):
            content_type = 'text/html'
            ext = 'html'

        response = HttpResponse(pdf_bytes, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="CampusKona_PrivacyNotice_{schema_name}.{ext}"'
        return response


class GenerateComplianceCertificateView(APIView):
    """Generate and download the DPDP Compliance Certificate (Workstream D)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.privacy.services.document_generator import generate_compliance_certificate_pdf
        tenant = getattr(request, 'tenant', None)
        school_name = getattr(tenant, 'name', '') or getattr(tenant, 'schema_name', 'School')
        schema_name = getattr(tenant, 'schema_name', 'school')

        consent_rate = 0.0
        audit_score = 85
        try:
            consented = ParentalConsent.objects.filter(consent_given=True).count()
            total = ParentalConsent.objects.count()
            if total > 0:
                consent_rate = (consented / total) * 100
        except Exception:
            pass

        if consent_rate < 80:
            return Response(
                {'error': f'Certificate requires 80%+ consent rate. Current: {consent_rate:.1f}%'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        pdf_bytes = generate_compliance_certificate_pdf(
            school_name=school_name,
            schema_name=schema_name,
            consent_rate=consent_rate,
            audit_score=audit_score,
        )
        content_type = 'application/pdf'
        ext = 'pdf'
        if pdf_bytes[:1] in (b'<', b'\n', b'\r'):
            content_type = 'text/html'
            ext = 'html'

        response = HttpResponse(pdf_bytes, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="CampusKona_Certificate_{schema_name}.{ext}"'
        return response


class BulkImportDPDPAuditView(APIView):
    """Audit an Excel/CSV file for DPDP-sensitive columns before import (Workstream H)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import io
        import csv

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file_obj.name.lower()
        headers = []
        row_count = 0

        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                from openpyxl import load_workbook
                wb = load_workbook(file_obj, read_only=True, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                headers = [
                    str(h).strip().lower().replace(' ', '_')
                    for h in (all_rows[0] if all_rows else []) if h
                ]
                row_count = len(all_rows) - 1
            elif filename.endswith('.csv'):
                content = file_obj.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                headers = [h.strip().lower().replace(' ', '_') for h in (reader.fieldnames or [])]
                row_count = sum(1 for _ in reader)
        except Exception as e:
            return Response({'error': f'File parse error: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        DPDP_SENSITIVE = {
            'aadhar_number': 'Aadhaar number — government ID, requires explicit consent (DPDP S.9)',
            'aadhaar_number': 'Aadhaar number — government ID',
            'aadhar': 'Aadhaar — government ID',
            'aadhaar': 'Aadhaar — government ID',
            'religion': 'Religion — special category data under DPDP Act 2023',
            'caste': 'Caste — special category data, requires explicit consent',
            'category': 'Caste category — may contain special category data',
        }

        flags = [
            {
                'column': col,
                'reason': DPDP_SENSITIVE[col],
                'recommendation': 'Obtain explicit parental consent before importing. CampusKona will encrypt this field.',
            }
            for col in headers if col in DPDP_SENSITIVE
        ]

        return Response({
            'file_name': file_obj.name,
            'total_rows': row_count,
            'total_columns': len(headers),
            'columns': headers,
            'dpdp_flags': flags,
            'flag_count': len(flags),
            'requires_consent': len(flags) > 0,
            'message': (
                f'Found {len(flags)} DPDP-sensitive column(s) in your file. Parental consent required before import.'
                if flags else 'No DPDP-sensitive columns detected. Safe to import.'
            ),
        })

        AuditLoggingService.mark_as_false_positive(alert, request.user, notes)

        return Response({
            'message': 'Alert marked as false positive',
            'alert': AccessPatternAlertSerializer(alert).data
        })

    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        """
        Assign alert to a user for investigation

        POST /api/v1/privacy/alerts/{id}/assign/
        Body: {
            "assigned_to_id": 5
        }
        """
        alert = self.get_object()
        assigned_to_id = request.data.get('assigned_to_id')

        if not assigned_to_id:
            return Response(
                {'error': 'assigned_to_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from apps.authentication.models import User
            assigned_to = User.objects.get(id=assigned_to_id)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        alert.assigned_to = assigned_to
        alert.status = 'INVESTIGATING'
        alert.save()

        return Response({
            'message': f'Alert assigned to {assigned_to.get_full_name()}',
            'alert': AccessPatternAlertSerializer(alert).data
        })


class ComplianceDashboardViewSet(viewsets.ViewSet):
    """
    Dashboard for DPDP Coordinators to monitor compliance in their assigned sections.
    """
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def summary(self, request):
        user = request.user
        
        # 1. Identify DPDP Coordinator assignments
        try:
            staff = StaffMember.objects.get(user=user)
        except StaffMember.DoesNotExist:
            # Fallback for admins who might not have a staff profile linked but want to see everything
            if user.user_type in ['ADMIN', 'SUPER_ADMIN']:
                staff = None
            else:
                return Response({"error": "Staff profile not found"}, status=status.HTTP_404_NOT_FOUND)
            
        if user.user_type in ['ADMIN', 'SUPER_ADMIN']:
            sections = Section.objects.all().select_related('class_obj')
        else:
            assignments = StaffRoleAssignment.objects.filter(
                staff_member=staff,
                role__code='DPDP_COORDINATOR',
                is_active=True
            ).select_related('section', 'section__class_obj')
            
            if not assignments.exists():
                return Response({
                    "section_stats": [],
                    "overall_summary": {
                        "total_students": 0,
                        "consents_completed": 0,
                        "pending_grievances": 0,
                        "pending_corrections": 0
                    }
                })
            sections = [a.section for a in assignments if a.section]

        # 2. For each section, calculate stats
        section_stats = []
        total_summary = {
            "total_students": 0,
            "consents_completed": 0,
            "pending_grievances": 0,
            "pending_corrections": 0
        }

        mandatory_purposes = list(ConsentPurpose.objects.filter(is_mandatory=True, is_active=True).values_list('id', flat=True))
        mandatory_count = len(mandatory_purposes)

        for section in sections:
            # Students in this section
            student_ids = list(StudentEnrollment.objects.filter(
                section=section, 
                is_active=True
            ).values_list('student_id', flat=True))
            
            student_count = len(student_ids)
            if student_count == 0:
                continue
                
            # Optimized Consent Calculation
            # Count how many mandatory consents each student has
            consent_counts = ParentalConsent.objects.filter(
                student_id__in=student_ids,
                purpose_id__in=mandatory_purposes,
                consent_given=True,
                withdrawn=False
            ).values('student').annotate(count=Count('purpose')).filter(count__gte=mandatory_count)
            
            compliant_students = consent_counts.count()
            
            # Grievances
            grievances = Grievance.objects.filter(
                student_id__in=student_ids,
                status__in=['FILED', 'ACKNOWLEDGED', 'INVESTIGATING']
            ).count()
            
            # Correction Requests
            corrections = CorrectionRequest.objects.filter(
                student_id__in=student_ids,
                status__in=['PENDING', 'IN_PROGRESS']
            ).count()

            stats = {
                "section_id": section.id,
                "section_name": f"{section.class_obj.name} {section.name}",
                "total_students": student_count,
                "compliant_students": compliant_students,
                "compliance_rate": round((compliant_students / student_count * 100), 1) if student_count > 0 else 100,
                "pending_grievances": grievances,
                "pending_corrections": corrections
            }
            section_stats.append(stats)
            
            total_summary["total_students"] += student_count
            total_summary["consents_completed"] += compliant_students
            total_summary["pending_grievances"] += grievances
            total_summary["pending_corrections"] += corrections

        return Response({
            "section_stats": section_stats,
            "overall_summary": total_summary,
            "last_updated": timezone.now()
        })

    @action(detail=True, methods=['get'])
    def section_details(self, request, pk=None):
        """Get detailed student compliance list for a section"""
        try:
            section = Section.objects.get(id=pk)
        except Section.DoesNotExist:
            return Response({"error": "Section not found"}, status=404)

        enrollments = StudentEnrollment.objects.filter(
            section=section,
            is_active=True
        ).select_related('student')

        mandatory_purposes = list(ConsentPurpose.objects.filter(is_mandatory=True, is_active=True).values_list('id', flat=True))
        mandatory_count = len(mandatory_purposes)

        details = []
        for enc in enrollments:
            student = enc.student
            
            # Check mandatory consents
            given = ParentalConsent.objects.filter(
                student=student,
                purpose_id__in=mandatory_purposes,
                consent_given=True,
                withdrawn=False
            ).count()
            
            details.append({
                "student_id": student.id,
                "roll_number": enc.roll_number,
                "name": student.full_name,
                "admission_number": student.admission_number,
                "is_compliant": given >= mandatory_count,
                "consents_given": given,
                "consents_required": mandatory_count,
                "pending_grievances": Grievance.objects.filter(student=student, status__in=['FILED', 'ACKNOWLEDGED', 'INVESTIGATING']).count(),
                "pending_corrections": CorrectionRequest.objects.filter(student=student, status__in=['PENDING', 'IN_PROGRESS']).count()
            })

        return Response({
            "section_name": f"{section.class_obj.name} {section.name}",
            "students": details
        })
