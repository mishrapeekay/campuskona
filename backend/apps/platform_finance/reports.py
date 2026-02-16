"""
Platform Finance Report Generators
==================================
Excel and PDF report generation for investor dashboard and financial ledger
"""

import os
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.core.files.base import ContentFile

from .models import InvestorMetric, FinancialLedger, FinancialSnapshot
from .services import InvestorMetricsService


class InvestorReportGenerator:
    """Generate investor dashboard reports"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        
        # Styling
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.subheader_font = Font(name='Arial', size=12, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        self.bold_font = Font(name='Arial', size=10, bold=True)
        
        self.header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        self.subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        self.total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_investor_dashboard_report(self, start_date, end_date, report_format='EXCEL'):
        """Generate comprehensive investor dashboard report"""
        if report_format == 'EXCEL':
            return self._generate_excel_dashboard(start_date, end_date)
        else:
            raise NotImplementedError(f"Format {report_format} not implemented yet")
    
    def _generate_excel_dashboard(self, start_date, end_date):
        """Generate Excel investor dashboard report"""
        self.workbook = Workbook()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create sheets
        self._create_summary_sheet(start_date, end_date)
        self._create_metrics_trend_sheet(start_date, end_date)
        self._create_regional_breakdown_sheet()
        self._create_financial_health_sheet()
        
        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f'investor_dashboard_{start_date}_{end_date}.xlsx'
        
        return ContentFile(output.read(), name=filename)
    
    def _create_summary_sheet(self, start_date, end_date):
        """Create summary sheet with key metrics"""
        ws = self.workbook.create_sheet('Summary')
        
        # Get latest metrics
        latest_metric = InvestorMetric.objects.filter(
            snapshot_date__lte=end_date
        ).order_by('-snapshot_date').first()
        
        if not latest_metric:
            latest_metric = InvestorMetricsService.create_daily_snapshot()
        
        # Title
        ws['A1'] = 'INVESTOR DASHBOARD - SUMMARY'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f'Period: {start_date} to {end_date}'
        ws['A2'].font = self.subheader_font
        ws.merge_cells('A2:D2')
        
        # Revenue Metrics
        row = 4
        ws[f'A{row}'] = 'REVENUE METRICS'
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        metrics_data = [
            ('Monthly Recurring Revenue (MRR)', f'₹{latest_metric.mrr:,.2f}'),
            ('Annual Recurring Revenue (ARR)', f'₹{latest_metric.arr:,.2f}'),
            ('Average Revenue Per School', f'₹{latest_metric.mrr / latest_metric.active_schools if latest_metric.active_schools > 0 else 0:,.2f}'),
        ]
        
        for label, value in metrics_data:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            ws[f'A{row}'].font = self.normal_font
            ws[f'C{row}'].font = self.bold_font
            row += 1
        
        # School Metrics
        row += 1
        ws[f'A{row}'] = 'SCHOOL METRICS'
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        school_data = [
            ('Total Schools', latest_metric.total_schools),
            ('Active Schools', latest_metric.active_schools),
            ('New Schools This Month', latest_metric.new_schools_this_month),
            ('Churned Schools This Month', latest_metric.churned_schools_this_month),
        ]
        
        for label, value in school_data:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            ws[f'A{row}'].font = self.normal_font
            ws[f'C{row}'].font = self.bold_font
            row += 1
        
        # Growth Metrics
        row += 1
        ws[f'A{row}'] = 'GROWTH & HEALTH METRICS'
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        growth_data = [
            ('Churn Rate', f'{latest_metric.churn_rate}%'),
            ('Growth Rate (MoM)', f'{latest_metric.growth_rate}%'),
            ('Customer Acquisition Cost (CAC)', f'₹{latest_metric.cac:,.2f}'),
            ('Lifetime Value (LTV)', f'₹{latest_metric.ltv:,.2f}'),
            ('LTV:CAC Ratio', f'{latest_metric.ltv_cac_ratio:.2f}'),
        ]
        
        for label, value in growth_data:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            ws[f'A{row}'].font = self.normal_font
            ws[f'C{row}'].font = self.bold_font
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['C'].width = 20
    
    def _create_metrics_trend_sheet(self, start_date, end_date):
        """Create metrics trend sheet"""
        ws = self.workbook.create_sheet('Trends')
        
        # Get historical data
        metrics = InvestorMetric.objects.filter(
            snapshot_date__gte=start_date,
            snapshot_date__lte=end_date
        ).order_by('snapshot_date')
        
        # Headers
        headers = ['Date', 'MRR', 'ARR', 'Active Schools', 'Churn Rate %', 'Growth Rate %', 'CAC', 'LTV', 'LTV:CAC']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.border = self.thin_border
        
        # Data rows
        for row, metric in enumerate(metrics, start=2):
            ws.cell(row=row, column=1, value=metric.snapshot_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=float(metric.mrr))
            ws.cell(row=row, column=3, value=float(metric.arr))
            ws.cell(row=row, column=4, value=metric.active_schools)
            ws.cell(row=row, column=5, value=float(metric.churn_rate))
            ws.cell(row=row, column=6, value=float(metric.growth_rate))
            ws.cell(row=row, column=7, value=float(metric.cac))
            ws.cell(row=row, column=8, value=float(metric.ltv))
            ws.cell(row=row, column=9, value=float(metric.ltv_cac_ratio))
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_regional_breakdown_sheet(self):
        """Create regional breakdown sheet"""
        ws = self.workbook.create_sheet('Regional Breakdown')
        
        # Get latest metrics
        latest_metric = InvestorMetric.objects.order_by('-snapshot_date').first()
        
        if not latest_metric:
            return
        
        # Title
        ws['A1'] = 'REGIONAL DISTRIBUTION'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:C1')
        
        # Headers
        ws['A3'] = 'State/Region'
        ws['B3'] = 'Number of Schools'
        ws['C3'] = 'Percentage'
        
        for cell in ['A3', 'B3', 'C3']:
            ws[cell].font = self.subheader_font
            ws[cell].fill = self.subheader_fill
        
        # Data
        total_schools = sum(latest_metric.region_distribution.values())
        row = 4
        
        for region, count in sorted(latest_metric.region_distribution.items(), 
                                    key=lambda x: x[1], reverse=True):
            ws[f'A{row}'] = region
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f'{(count / total_schools * 100):.1f}%' if total_schools > 0 else '0%'
            row += 1
        
        # Total
        ws[f'A{row}'] = 'TOTAL'
        ws[f'B{row}'] = total_schools
        ws[f'C{row}'] = '100%'
        
        for cell in [f'A{row}', f'B{row}', f'C{row}']:
            ws[cell].font = self.bold_font
            ws[cell].fill = self.total_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def _create_financial_health_sheet(self):
        """Create financial health indicators sheet"""
        ws = self.workbook.create_sheet('Financial Health')
        
        # Get latest snapshot
        snapshot = FinancialSnapshot.objects.order_by('-snapshot_date').first()
        
        if not snapshot:
            return
        
        # Title
        ws['A1'] = 'FINANCIAL HEALTH INDICATORS'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        # Profitability
        row = 3
        ws[f'A{row}'] = 'PROFITABILITY'
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        profit_data = [
            ('Platform Revenue (Total)', f'₹{snapshot.platform_revenue_total:,.2f}'),
            ('Platform Expenses (Total)', f'₹{snapshot.platform_expenses_total:,.2f}'),
            ('Gross Profit', f'₹{snapshot.gross_profit:,.2f}'),
            ('Net Profit', f'₹{snapshot.net_profit:,.2f}'),
            ('Profit Margin', f'{(snapshot.net_profit / snapshot.platform_revenue_total * 100) if snapshot.platform_revenue_total > 0 else 0:.2f}%'),
        ]
        
        for label, value in profit_data:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            ws[f'A{row}'].font = self.normal_font
            ws[f'C{row}'].font = self.bold_font
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['C'].width = 20


class LedgerReportGenerator:
    """Generate financial ledger reports"""
    
    def generate_ledger_export(self, queryset):
        """Export ledger entries to Excel"""
        workbook = Workbook()
        ws = workbook.active
        ws.title = 'Financial Ledger'
        
        # Styling
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        
        # Headers
        headers = [
            'Sequence', 'Date', 'Transaction Type', 'Category',
            'Amount', 'Currency', 'Tenant', 'Description',
            'Reference Type', 'Reference ID', 'Created By', 'Hash'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        for row, entry in enumerate(queryset, start=2):
            ws.cell(row=row, column=1, value=entry.sequence_number)
            ws.cell(row=row, column=2, value=entry.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=3, value=entry.get_transaction_type_display())
            ws.cell(row=row, column=4, value=entry.get_category_display())
            ws.cell(row=row, column=5, value=float(entry.amount))
            ws.cell(row=row, column=6, value=entry.currency)
            ws.cell(row=row, column=7, value=entry.tenant_schema)
            ws.cell(row=row, column=8, value=entry.description)
            ws.cell(row=row, column=9, value=entry.reference_type)
            ws.cell(row=row, column=10, value=str(entry.reference_id) if entry.reference_id else '')
            ws.cell(row=row, column=11, value=entry.created_by.email if entry.created_by else '')
            ws.cell(row=row, column=12, value=entry.current_hash[:16] + '...')
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'ledger_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return ContentFile(output.read(), name=filename)
