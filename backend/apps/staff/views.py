from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from apps.authentication.permissions import IsPrincipal, IsTeacher, IsOwnerOrAdmin
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone

from apps.staff.models import (
    StaffMember,
    StaffDocument,
    # StaffAttendance and StaffLeave moved to attendance app
    StaffQualification,
    StaffExperience
)
from apps.staff.serializers import (
    StaffMemberSerializer,
    StaffMemberCreateSerializer,
    StaffMemberUpdateSerializer,
    StaffMemberListSerializer,
    StaffDocumentSerializer,
    # StaffAttendanceSerializer and StaffLeaveSerializer moved to attendance app
    StaffQualificationSerializer,
    StaffExperienceSerializer
)
from apps.core.models import AuditLog


class StaffMemberViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Staff Member CRUD operations
    """
    permission_classes = [IsPrincipal]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['employment_status', 'employment_type', 'designation', 'department', 'gender']
    search_fields = [
        'employee_id',
        'first_name',
        'last_name',
        'email',
        'phone_number',
        'designation'
    ]
    ordering_fields = ['employee_id', 'first_name', 'joining_date', 'created_at']
    ordering = ['-joining_date']

    def get_queryset(self):
        """Filter out soft-deleted staff, with optional subject filtering for timetable."""
        queryset = StaffMember.objects.filter(is_deleted=False)
        # Filter teachers by subject they teach (used by timetable cascade dropdown)
        subject_id = self.request.query_params.get('subject_id')
        if subject_id:
            queryset = queryset.filter(subjects_taught__contains=[str(subject_id)])
        return queryset

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return StaffMemberListSerializer
        elif self.action == 'create':
            return StaffMemberCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return StaffMemberUpdateSerializer
        return StaffMemberSerializer

    def perform_create(self, serializer):
        """Create staff member with audit log"""
        staff_member = serializer.save()

        AuditLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='StaffMember',
            object_id=str(staff_member.id),
            changes=serializer.validated_data,
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    def perform_update(self, serializer):
        """Update staff member with audit log"""
        old_data = StaffMemberSerializer(serializer.instance).data
        staff_member = serializer.save()

        AuditLog.objects.create(
            user=self.request.user,
            action='UPDATE',
            model_name='StaffMember',
            object_id=str(staff_member.id),
            changes={
                'old': old_data,
                'new': StaffMemberSerializer(staff_member).data
            },
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    def perform_destroy(self, instance):
        """Soft delete staff member"""
        instance.delete(user=self.request.user)

        AuditLog.objects.create(
            user=self.request.user,
            action='DELETE',
            model_name='StaffMember',
            object_id=str(instance.id),
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    @action(detail=True, methods=['get'])
    def profile(self, request, pk=None):
        """
        Get complete staff profile with all related data
        """
        staff_member = self.get_object()

        profile_data = StaffMemberSerializer(staff_member).data
        profile_data['documents'] = StaffDocumentSerializer(
            staff_member.documents.filter(is_deleted=False),
            many=True
        ).data
        profile_data['qualifications'] = StaffQualificationSerializer(
            staff_member.qualifications.filter(is_deleted=False),
            many=True
        ).data
        profile_data['work_experiences'] = StaffExperienceSerializer(
            staff_member.work_experiences.filter(is_deleted=False),
            many=True
        ).data

        return Response(profile_data)

    @action(detail=True, methods=['post'], url_path='assign-subjects')
    def assign_subjects(self, request, pk=None):
        """
        Assign subjects to a staff member.
        Expects: { "subjects": ["uuid1", "uuid2", ...] }
        """
        from apps.academics.models import Subject

        staff_member = self.get_object()
        subject_ids = request.data.get('subjects', [])

        if not isinstance(subject_ids, list):
            return Response(
                {'error': 'subjects must be a list of subject IDs'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate that all subject IDs exist
        valid_subjects = Subject.objects.filter(
            id__in=subject_ids, is_deleted=False, is_active=True
        )
        valid_ids = [str(s.id) for s in valid_subjects]

        invalid_ids = [sid for sid in subject_ids if str(sid) not in valid_ids]
        if invalid_ids:
            return Response(
                {'error': f'Invalid subject IDs: {invalid_ids}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update the staff member's subjects_taught field
        staff_member.subjects_taught = [str(sid) for sid in subject_ids]
        staff_member.save(update_fields=['subjects_taught', 'updated_at'])

        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            model_name='StaffMember',
            object_id=str(staff_member.id),
            changes={'subjects_taught': staff_member.subjects_taught},
            ip_address=request.META.get('REMOTE_ADDR')
        )

        # Return subject details for frontend display
        subjects_data = [
            {'id': str(s.id), 'name': s.name, 'code': s.code}
            for s in valid_subjects
        ]

        return Response({
            'staff_id': str(staff_member.id),
            'subjects': subjects_data,
        })

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Get staff statistics for dashboard.
        """
        queryset = self.get_queryset()

        total = queryset.count()
        active = queryset.filter(employment_status='ACTIVE').count()
        on_leave = queryset.filter(employment_status='ON_LEAVE').count()
        resigned = queryset.filter(employment_status='RESIGNED').count()
        terminated = queryset.filter(employment_status='TERMINATED').count()

        # Gender distribution
        male = queryset.filter(gender='M').count()
        female = queryset.filter(gender='F').count()
        other = queryset.filter(gender='O').count()

        # Employment type distribution
        permanent = queryset.filter(employment_type='PERMANENT').count()
        contract = queryset.filter(employment_type='CONTRACT').count()
        probation = queryset.filter(employment_type='PROBATION').count()

        return Response({
            'total': total,
            'by_status': {
                'active': active,
                'on_leave': on_leave,
                'resigned': resigned,
                'terminated': terminated,
            },
            'by_gender': {
                'male': male,
                'female': female,
                'other': other,
            },
            'by_employment_type': {
                'permanent': permanent,
                'contract': contract,
                'probation': probation,
            }
        })

    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """
        Get dashboard statistics.
        If user is ADMIN, returns general staff stats.
        If user is TEACHER, returns teacher-specific stats.
        """
        user = request.user
        
        if user.user_type == 'TEACHER':
            try:
                staff_member = user.staff_profile
            except StaffMember.DoesNotExist:
                return Response({"error": "Staff profile not found"}, status=404)
                
            # Imports inside method to avoid circular imports if any
            from apps.academics.models import Section, StudentEnrollment
            from apps.attendance.models import StudentAttendance 
            from datetime import date
            
            # 1. Assigned Class & Student Count
            # Find active sections where this teacher is the class teacher
            managed_sections = Section.objects.filter(
                class_teacher=staff_member,
                is_active=True
            )
            
            assigned_class_name = "No Class Assigned"
            total_students = 0
            section_ids = []
            
            if managed_sections.exists():
                # For simplicity, if multiple sections, show the first one or a summary
                # Usually a teacher is class teacher of one section
                primary_section = managed_sections.first()
                assigned_class_name = f"{primary_section.class_instance.name} - {primary_section.name}"
                
                # Count total students across all managed sections
                for section in managed_sections:
                    count = StudentEnrollment.objects.filter(
                        section=section,
                        is_active=True,
                        enrollment_status='ENROLLED'
                    ).count()
                    total_students += count
                    section_ids.append(section.id)

            # 2. Pending Attendance
            # Check if attendance marked for today for the managed sections
            # We assume attendance is marked if at least one record exists for the section today
            attendance_marked = False
            if section_ids:
                today = date.today()
                # Find students in these sections
                from apps.students.models import Student
                student_ids = StudentEnrollment.objects.filter(
                    section_id__in=section_ids,
                    is_active=True
                ).values_list('student_id', flat=True)
                
                # Check for attendance records
                attendance_count = StudentAttendance.objects.filter(
                    date=today,
                    student_id__in=student_ids
                ).count()
                
                # If we have some records, we assume it's started/done
                # Logic can be refined to check if ALL are marked
                if attendance_count > 0:
                    attendance_marked = True

            return Response({
                'role': 'TEACHER',
                'assigned_class': assigned_class_name,
                'total_students': total_students,
                'attendance_marked': attendance_marked,
                'today_classes_count': 0, # Placeholder until Timetable is ready
            })

        # ADMIN STATS (Existing logic)
        queryset = self.get_queryset()
        
        total = queryset.count()
        active = queryset.filter(employment_status='ACTIVE').count()
        on_leave = queryset.filter(employment_status='ON_LEAVE').count()
        resigned = queryset.filter(employment_status='RESIGNED').count()
        
        # Employment type distribution
        permanent = queryset.filter(employment_type='PERMANENT').count()
        contract = queryset.filter(employment_type='CONTRACT').count()
        temporary = queryset.filter(employment_type='TEMPORARY').count()
        
        # Gender distribution
        male = queryset.filter(gender='M').count()
        female = queryset.filter(gender='F').count()
        other = queryset.filter(gender='O').count()
        
        return Response({
            'role': 'ADMIN',
            'total': total,
            'by_status': {
                'active': active,
                'on_leave': on_leave,
                'resigned': resigned,
            },
            'by_employment_type': {
                'permanent': permanent,
                'contract': contract,
                'temporary': temporary,
            },
            'by_gender': {
                'male': male,
                'female': female,
                'other': other,
            }
        })

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """
        Bulk upload staff from Excel/CSV file (Workstream A).
        Accepts .xlsx, .xls, .csv files.
        """
        import io
        import csv
        from openpyxl import load_workbook
        from django.db import transaction
        from django.contrib.auth import get_user_model
        User = get_user_model()

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file_obj.name.lower()
        rows = []
        headers = []

        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = load_workbook(file_obj, read_only=True, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    return Response({'error': 'Empty file'}, status=status.HTTP_400_BAD_REQUEST)
                headers = [
                    str(h).strip().lower().replace(' ', '_') if h else ''
                    for h in all_rows[0]
                ]
                rows = [dict(zip(headers, row)) for row in all_rows[1:] if any(row)]
            elif filename.endswith('.csv'):
                content = file_obj.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                headers = [h.strip().lower().replace(' ', '_') for h in (reader.fieldnames or [])]
                rows = list(reader)
            else:
                return Response(
                    {'error': 'Unsupported file format. Use .xlsx or .csv'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response({'error': f'File parse error: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        FIELD_MAP = {
            'name': 'first_name', 'staff_name': 'first_name', 'first_name': 'first_name',
            'last_name': 'last_name', 'surname': 'last_name',
            'email': 'email',
            'phone': 'phone_number', 'mobile': 'phone_number', 'phone_number': 'phone_number',
            'employee_id': 'employee_id', 'staff_id': 'employee_id',
            'joining_date': 'joining_date', 'date_of_joining': 'joining_date',
            'designation': 'designation',
            'employment_type': 'employment_type',
            'gender': 'gender',
        }

        imported = 0
        errors = []
        warnings = []
        schema = request.tenant.schema_name if hasattr(request, 'tenant') else 'public'

        for row_idx, row in enumerate(rows, start=2):
            mapped = {}
            for col, val in row.items():
                field = FIELD_MAP.get(str(col).lower().strip())
                if field and val is not None and str(val).strip():
                    mapped[field] = str(val).strip()

            first_name = mapped.get('first_name', '')
            if not first_name:
                errors.append({'row': row_idx, 'error': 'Name is required'})
                continue

            employee_id = mapped.get('employee_id', '')
            if not employee_id:
                import uuid
                employee_id = f'EMP-{uuid.uuid4().hex[:6].upper()}'
                warnings.append({'row': row_idx, 'warning': f'Auto-assigned employee_id: {employee_id}'})

            email = mapped.get('email') or f'staff.{employee_id.lower()}@{schema}.campuskona.internal'

            try:
                with transaction.atomic():
                    if User.objects.filter(email=email).exists():
                        warnings.append({'row': row_idx, 'warning': f'User {email} already exists, skipped'})
                        continue

                    user = User.objects.create(
                        email=email,
                        username=email,
                        first_name=first_name,
                        last_name=mapped.get('last_name', ''),
                        is_active=True,
                    )
                    user.set_unusable_password()
                    user.save(update_fields=['password'])

                    joining_date = timezone.now().date()
                    if mapped.get('joining_date'):
                        from datetime import datetime as _dt
                        for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y']:
                            try:
                                joining_date = _dt.strptime(mapped['joining_date'], fmt).date()
                                break
                            except ValueError:
                                continue

                    gender_raw = mapped.get('gender', 'M').upper()
                    gender = 'M' if gender_raw in ('M', 'MALE') else ('F' if gender_raw in ('F', 'FEMALE') else 'O')

                    StaffMember.objects.create(
                        user=user,
                        employee_id=employee_id,
                        first_name=first_name,
                        last_name=mapped.get('last_name', ''),
                        email=email,
                        joining_date=joining_date,
                        designation=mapped.get('designation', 'TEACHER'),
                        employment_type=mapped.get('employment_type', 'PERMANENT'),
                        employment_status='ACTIVE',
                        gender=gender,
                    )
                    imported += 1
            except Exception as e:
                errors.append({'row': row_idx, 'error': str(e)})

        return Response({
            'imported': imported,
            'total_rows': len(rows),
            'warnings': warnings,
            'errors': errors,
            'message': f'Successfully imported {imported} of {len(rows)} staff members',
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='bulk_template')
    def bulk_template(self, request):
        """Download sample Excel template for bulk staff upload."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = 'Staff'
        headers = [
            'first_name', 'last_name', 'email', 'phone_number', 'gender',
            'employee_id', 'joining_date', 'designation', 'employment_type',
        ]
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 20
        ws.append(['Priya', 'Mehta', 'priya@school.com', '9876543210', 'F', 'EMP001', '2024-06-01', 'TEACHER', 'PERMANENT'])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="staff_bulk_upload_template.xlsx"'
        return response


class StaffDocumentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Staff Documents
    """
    serializer_class = StaffDocumentSerializer
    permission_classes = [IsOwnerOrAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['staff_member', 'document_type', 'is_verified']
    ordering_fields = ['created_at', 'issue_date']
    ordering = ['-created_at']

    def get_queryset(self):
        """Filter documents based on user type"""
        queryset = StaffDocument.objects.filter(is_deleted=False)

        user = self.request.user
        if user.user_type == 'TEACHER':
            # Staff can only see their own documents
            queryset = queryset.filter(staff_member__user=user)

        return queryset.select_related('staff_member', 'verified_by')

    def perform_create(self, serializer):
        """Create document with audit log"""
        document = serializer.save()

        AuditLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='StaffDocument',
            object_id=str(document.id),
            ip_address=self.request.META.get('REMOTE_ADDR')
        )

    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """
        Verify a staff document
        """
        document = self.get_object()

        document.is_verified = True
        document.verified_by = request.user
        document.verified_at = timezone.now()
        document.save()

        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            model_name='StaffDocument',
            object_id=str(document.id),
            changes={'verified': True},
            ip_address=request.META.get('REMOTE_ADDR')
        )

        return Response({
            'message': 'Document verified successfully',
            'document': StaffDocumentSerializer(document).data
        })


# NOTE: StaffAttendanceViewSet and StaffLeaveViewSet have been moved to apps.attendance.views

class StaffQualificationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Staff Qualifications
    """
    serializer_class = StaffQualificationSerializer
    permission_classes = [IsOwnerOrAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['staff_member', 'qualification_type']
    ordering_fields = ['year_of_completion']
    ordering = ['-year_of_completion']

    def get_queryset(self):
        """Filter based on user type"""
        queryset = StaffQualification.objects.filter(is_deleted=False)

        user = self.request.user
        if user.user_type == 'TEACHER':
            queryset = queryset.filter(staff_member__user=user)

        return queryset.select_related('staff_member')

    def perform_create(self, serializer):
        """Create qualification with audit log"""
        qualification = serializer.save()

        AuditLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='StaffQualification',
            object_id=str(qualification.id),
            ip_address=self.request.META.get('REMOTE_ADDR')
        )


class StaffExperienceViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Staff Work Experience
    """
    serializer_class = StaffExperienceSerializer
    permission_classes = [IsOwnerOrAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['staff_member', 'is_current']
    ordering_fields = ['from_date']
    ordering = ['-from_date']

    def get_queryset(self):
        """Filter based on user type"""
        queryset = StaffExperience.objects.filter(is_deleted=False)

        user = self.request.user
        if user.user_type == 'TEACHER':
            queryset = queryset.filter(staff_member__user=user)

        return queryset.select_related('staff_member')

    def perform_create(self, serializer):
        """Create experience with audit log"""
        experience = serializer.save()

        AuditLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='StaffExperience',
            object_id=str(experience.id),
            ip_address=self.request.META.get('REMOTE_ADDR')
        )
