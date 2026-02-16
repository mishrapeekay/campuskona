from django.db.models import Count, Q
from apps.students.models import Student
from apps.staff.models import StaffMember
from apps.academics.models import StudentEnrollment, AcademicYear, Class
from apps.government_reports.models import UDISECodeMapping
from django.utils import timezone
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

class UDISEPlusService:
    def generate_report_data(self, academic_year_str):
        # 1. Get Academic Year
        try:
             academic_year = AcademicYear.objects.get(name=academic_year_str)
        except AcademicYear.DoesNotExist:
             # Fallback: Try to find by partial match or get current
             academic_year = AcademicYear.objects.filter(is_current=True).first()
             if not academic_year:
                 raise ValueError(f"Academic Year {academic_year_str} not found and no current year set")

        return {
            "academic_year": academic_year.name,
            "school_profile": self._get_school_profile(),
            "student_stats": self._get_student_stats(academic_year),
            "staff_stats": self._get_staff_stats(academic_year),
            "infrastructure": self._get_infrastructure_data()
        }

    def generate_excel(self, data):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UDISE+ Summary"

        # Styles
        header_font = Font(bold=True, size=14)
        sub_header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"UDISE+ Report Summary - {data['academic_year']}"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        # School Profile
        row = 3
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "School Profile"
        ws[f'A{row}'].font = sub_header_font
        row += 1
        
        profile = data['school_profile']
        for key, value in profile.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Student Stats
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "Student Statistics"
        ws[f'A{row}'].font = sub_header_font
        row += 1

        stats = data['student_stats']
        ws[f'A{row}'] = "Total Enrollment"
        ws[f'B{row}'] = stats['total_enrollment']
        row += 1

        ws[f'A{row}'] = "Gender Breakdown"
        row += 1
        for gender, count in stats['by_gender'].items():
            ws[f'B{row}'] = gender.title()
            ws[f'C{row}'] = count
            row += 1
        
        ws[f'A{row}'] = "Social Category Breakdown"
        row += 1
        for category, count in stats['by_category'].items():
            ws[f'B{row}'] = category.title()
            ws[f'C{row}'] = count
            row += 1

        row += 1
        ws[f'A{row}'] = "Class-wise Enrollment"
        row += 1
        headers = ["Class", "Total", "Boys", "Girls", "CWSN"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = border
        
        row += 1
        for cls_stat in stats['by_class']:
            ws.cell(row=row, column=1, value=cls_stat['class']).border = border
            ws.cell(row=row, column=2, value=cls_stat['total']).border = border
            ws.cell(row=row, column=3, value=cls_stat['boys']).border = border
            ws.cell(row=row, column=4, value=cls_stat['girls']).border = border
            ws.cell(row=row, column=5, value=cls_stat['cwsn']).border = border
            row += 1

        # Staff Stats
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "Staff Statistics"
        ws[f'A{row}'].font = sub_header_font
        row += 1
        
        staff = data['staff_stats']
        for key, value in staff.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _get_student_stats(self, academic_year):
        # Base QuerySet: Active enrollments for the year
        enrollments = StudentEnrollment.objects.filter(
            academic_year=academic_year,
            is_active=True,
            enrollment_status='ENROLLED'
        ).select_related('student', 'section__class_instance')

        # Use mappings if needed, for now we stick to raw counts but map keys for display if we want
        # Ideally, we would loop through UDISECodeMapping objects to form the keys,
        # but for aggregation, static keys are easier. 
        
        stats = {
            "total_enrollment": enrollments.count(),
            "by_gender": {
                "boys": enrollments.filter(student__gender='M').count(),
                "girls": enrollments.filter(student__gender='F').count(),
                "transgender": enrollments.filter(student__gender='O').count(),
            },
            "by_category": {
                "general": enrollments.filter(student__category='GENERAL').count(),
                "sc": enrollments.filter(student__category='SC').count(),
                "st": enrollments.filter(student__category='ST').count(),
                "obc": enrollments.filter(student__category='OBC').count(),
                "ews": enrollments.filter(student__category='EWS').count(),
            },
            "cwsn": enrollments.filter(student__is_differently_abled=True).count(),
            "by_class": []
        }
        
        # Breakdown by Class
        classes = Class.objects.filter(is_active=True).order_by('class_order')
        for cls in classes:
            cls_enrollments = enrollments.filter(section__class_instance=cls)
            stats["by_class"].append({
                "class": cls.name,
                "total": cls_enrollments.count(),
                "boys": cls_enrollments.filter(student__gender='M').count(),
                "girls": cls_enrollments.filter(student__gender='F').count(),
                "cwsn": cls_enrollments.filter(student__is_differently_abled=True).count(),
            })
            
        return stats

    def _get_staff_stats(self, academic_year):
        # Staff is usually not tied to academic year strictly in the same way as students (re-enrollment), 
        # but we check who was active.
        
        active_staff = StaffMember.objects.filter(
            Q(employment_status='ACTIVE') | Q(employment_status='ON_LEAVE')
        )
        
        # Broad filter for "Teacher" designation or explicit role
        teachers = active_staff.filter(
            Q(designation__icontains='TEACHER') | 
            Q(designation__in=['PRINCIPAL', 'VICE_PRINCIPAL', 'HEAD_TEACHER', 'PRT', 'TGT', 'PGT'])
        )
        
        return {
            "total_staff": active_staff.count(),
            "teaching_staff": teachers.count(),
            "non_teaching_staff": active_staff.exclude(id__in=teachers.values('id')).count(),
            "trained_teachers": teachers.filter(highest_qualification__in=['B_ED', 'M_ED', 'D_ED', 'DIPLOMA']).count()
        }

    def _get_school_profile(self):
        # Verification pending on where to get School Name/UDISE Code
        # Usually in TenantSettings or similar.
        # Returning placeholders for now.
        return {
            "school_name": "Demo School", # Replace with dynamic fetch
            "udise_code": "NOT_SET",
            "school_category": "Primary with Upper Primary",
            "school_management": "Private Unaided"
        }

    def _get_infrastructure_data(self):
        return {
            "status": "Infrastructure data not integrated",
            "classrooms": 0, # Placeholder
            "toilets": 0, # Placeholder
            "labs": 0 # Placeholder
        }
