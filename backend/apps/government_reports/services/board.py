from apps.examinations.models import ExamResult, Examination
from apps.academics.models import Class, AcademicYear
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

class BoardReportService:
    def generate_cbse_format(self, academic_year_str, class_id=None, exam_id=None):
        try:
             academic_year = AcademicYear.objects.get(name=academic_year_str)
        except AcademicYear.DoesNotExist:
             raise ValueError(f"Academic Year {academic_year_str} not found")

        # Base Query
        results = ExamResult.objects.filter(
            examination__academic_year=academic_year
        ).select_related('student', 'class_obj', 'section', 'examination')

        if class_id:
            results = results.filter(class_obj__id=class_id)
        
        if exam_id:
            results = results.filter(examination__id=exam_id)
        
        # Sort by Class, Section, Student Name
        results = results.order_by('class_obj__class_order', 'section__name', 'student__first_name')

        return self._create_excel(results, academic_year.name)

    def _create_excel(self, results, year):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Board Report"

        # Styles
        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        headers = [
            "Admission No", "Student Name", "Class", "Section", 
            "Exam Name", "Total Marks", "Max Marks", "Percentage", 
            "Grade", "Rank", "Result"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.border = border

        row = 2
        for result in results:
            ws.cell(row=row, column=1, value=result.student.admission_number).border = border
            ws.cell(row=row, column=2, value=result.student.get_full_name()).border = border
            ws.cell(row=row, column=3, value=result.class_obj.name).border = border
            ws.cell(row=row, column=4, value=result.section.name).border = border
            ws.cell(row=row, column=5, value=result.examination.name).border = border
            ws.cell(row=row, column=6, value=result.total_marks_obtained).border = border
            ws.cell(row=row, column=7, value=result.total_max_marks).border = border
            ws.cell(row=row, column=8, value=result.percentage).border = border
            ws.cell(row=row, column=9, value=result.overall_grade).border = border
            ws.cell(row=row, column=10, value=result.rank if result.rank else "-").border = border
            ws.cell(row=row, column=11, value="PASS" if result.is_passed else "FAIL").border = border
            row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
