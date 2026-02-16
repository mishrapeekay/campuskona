from apps.students.models import Student
from apps.academics.models import StudentEnrollment, Class, AcademicYear
from apps.government_reports.models import RTEComplianceRecord
from django.db.models import Sum
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

class RTEComplianceService:
    def get_compliance_status(self, academic_year_str):
        # 1. Get Academic Year
        try:
             academic_year = AcademicYear.objects.get(name=academic_year_str)
        except AcademicYear.DoesNotExist:
             raise ValueError(f"Academic Year {academic_year_str} not found")

        # 2. Determine Entry Level Class (Usually Class 1 or KG)
        # We'll assume the lowest ordered class is the entry level
        entry_class = Class.objects.filter(is_active=True).order_by('class_order').first()
        if not entry_class:
            return {"error": "No active classes found defined in the system"}

        # 3. Calculate Total Intake Capacity
        # Sum of max_students of all sections in entry class for this academic year
        total_seats = entry_class.sections.filter(
            academic_year=academic_year,
            is_active=True
        ).aggregate(total=Sum('max_students'))['total'] or 0

        reserved_seats = int(total_seats * 0.25)
        
        # 4. Calculate Actual RTE Admissions
        enrollments = StudentEnrollment.objects.filter(
            academic_year=academic_year,
            section__class_instance=entry_class,
            is_active=True,
            enrollment_status='ENROLLED'
        )
        
        # RTE Students typically marked as EWS or specific quota
        # Using 'EWS' category as primary indicator for now.
        rte_filled = enrollments.filter(student__category='EWS').count()
        
        # 5. Update or Create Compliance Record
        compliance_record, created = RTEComplianceRecord.objects.get_or_create(
            academic_year=academic_year.name,
            defaults={
                'total_intake_capacity': total_seats,
                'reserved_seats_rte': reserved_seats,
                'seats_filled_rte': rte_filled
            }
        )
        
        if not created:
            compliance_record.total_intake_capacity = total_seats
            compliance_record.reserved_seats_rte = reserved_seats
            compliance_record.seats_filled_rte = rte_filled
            compliance_record.save()
            
        return {
            "academic_year": academic_year.name,
            "entry_level_class": entry_class.display_name,
            "total_intake_capacity": total_seats,
            "mandated_rte_reservation": "25%",
            "reserved_seats_target": reserved_seats,
            "actual_rte_admissions": rte_filled,
            "shortfall": max(0, reserved_seats - rte_filled),
            "compliance_status": "COMPLIANT" if rte_filled >= reserved_seats else "NON_COMPLIANT",
            "compliance_percentage": compliance_record.compliance_percentage
        }

    def generate_excel(self, data):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RTE Compliance"

        # Styles
        header_font = Font(bold=True, size=14)
        key_font = Font(bold=True)
        
        ws.merge_cells('A1:B1')
        ws['A1'] = f"RTE Compliance Report - {data.get('academic_year', 'Unknown')}"
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 3
        for key, value in data.items():
            if key == 'error':
                ws[f'A{row}'] = "Error"
                ws[f'B{row}'] = value
                continue
            
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'A{row}'].font = key_font
            ws[f'B{row}'] = value
            row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
